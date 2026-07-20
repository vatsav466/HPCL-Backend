import urdhva_base
import time
import requests
import datetime
import hpcl_ceg_model
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import orchestrator.notification_manager.notification_factory as notification_factory

logger = urdhva_base.Logger.getInstance("va_day_end_closure")


async def send_closure_notification(
    html_table, attachment_file, total_alerts, sap_id, location_name
):

    ins = await notification_factory.get_notification_module("email")

    closure_date = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime(
        "%d-%b-%Y"
    )

    body = body = f"""
                    <html>
                    <head>
                    <style>
                    body {{
                        font-family: Arial, Helvetica, sans-serif;
                        font-size: 14px;
                        color: #333333;
                    }}

                    .location-table {{
                        border-collapse: collapse;
                        margin-bottom: 15px;
                    }}

                    .location-table td {{
                        border: 1px solid #d9d9d9;
                        padding: 6px 10px;
                    }}

                    .summary-table {{
                        border-collapse: collapse;
                        width: 90%;
                        border: 1px solid #808080;
                    }}

                    .summary-table th {{
                        background-color: #4472C4;
                        color: white;
                        border: 1px solid #808080;
                        padding: 8px;
                        text-align: center;
                    }}

                    .summary-table td {{
                        border: 1px solid #808080;
                        padding: 8px;
                    }}

                    .summary-table tr:nth-child(even) {{
                        background-color: #F5F5F5;
                    }}

                    .total-row {{
                        background-color: #E2F0D9;
                        font-weight: bold;
                    }}

                    .footer {{
                        margin-top: 20px;
                    }}
                    </style>
                    </head>

                    <body>

                    <p>Hi Team,</p>

                    <p>
                    The following VA alerts created on
                    <b>{closure_date}</b> have been automatically closed
                    as part of the Day End Closure activity.
                    </p>

                    <table class="location-table">
                    <tr>
                        <td><b>Location ID</b></td>
                        <td>{sap_id}</td>
                    </tr>
                    <tr>
                        <td><b>Location Name</b></td>
                        <td>{location_name}</td>
                    </tr>
                    </table>

                    {html_table}

                    <div class="footer">
                    Regards,<br>
                    NOVEX System
                    </div>

                    </body>
                    </html>
                    """
    await ins.publish_message(
        subject=f"VA Day End Alert Closure Summary - {sap_id} - {location_name}",
        recipients=["sreedhar.maddipati@algofusiontech.com "],
        cc_recipients=["venu@algofusiontech.com", "moufikali@algofusiontech.com"],
        bcc_recipients=[
            "yesu.p@algofusiontech.com",
            "poojitha.gumma@algofusiontech.com",
        ],
        html_content=True,
        body=body,
        force_send=True,
        inline_images={},
        attachments=[attachment_file],
    )


async def get_process_instance_id(business_key, camunda_url):
    process_instance_id = ""
    params = {"businessKey": business_key}

    url = f"{camunda_url}/engine-rest/process-instance"

    try:
        response = requests.get(
            url, params=params, timeout=(15, 15)  # (connect timeout, read timeout)
        )

        response.raise_for_status()

        instances = response.json()

        if instances:
            process_instance_id = instances[0]["id"]
            return process_instance_id

        print(f"Camunda flow not found for business key: {business_key}")

    except Exception as e:
        print(f"Unexpected error while fetching process instance: {str(e)}")

    return process_instance_id


async def _close_camunda_workflow(alert_data=None):
    # camunda_url = await helpers.get_alert_camunda_url(self.params['alert_id'], "error")
    camunda_url = alert_data.get("workflow_url", "")
    MAX_RETRIES = 3
    RETRY_DELAY = 10
    headers = {"Content-Type": "application/json"}
    business_key = alert_data.get("unique_id")
    instance_id = await get_process_instance_id(business_key, camunda_url)

    if not instance_id:
        instance_id = alert_data.get("workflow_instance_id")

    url = f"{camunda_url}/engine-rest/process-instance/{instance_id}"
    for attempt in range(MAX_RETRIES):
        try:
            response = requests.delete(
                url, headers=headers, timeout=(15, 15)
            )  # (connect timeout, read timeout)

            if response.status_code == 204:  # Success in Camunda
                print(f"{instance_id} Deleted successfully.")
                logger.info(f"{instance_id} Deleted successfully.")
                break
            else:
                print(
                    f"Error Deleting {alert_data['id']} {instance_id} {camunda_url} (attempt {attempt + 1}): {response.status_code} - {response.text}"
                )
                logger.info(
                    f"Error Deleting {alert_data['id']} {camunda_url} {instance_id} (attempt {attempt + 1}): {response.status_code} - {response.text}"
                )

        except requests.RequestException as e:
            print(
                f"Request error for {camunda_url} {instance_id} {alert_data['id']} (attempt {attempt + 1}): {e}"
            )
            logger.info(
                f"Request error for {camunda_url} {instance_id} {alert_data['id']} (attempt {attempt + 1}): {e}"
            )

        # Retry logic with exponential backoff
        if attempt < MAX_RETRIES - 1:
            time.sleep(RETRY_DELAY)
        else:
            print(
                f"Failed to Deleting {camunda_url} {instance_id} after {MAX_RETRIES} retries."
            )
            logger.info(
                f"Failed to Deleting {camunda_url} {instance_id} after {MAX_RETRIES} retries."
            )
            return False
    return True


async def day_end_closure():
    query = f"""SELECT
                    *
                FROM alerts
                WHERE alert_section = 'VA'
                AND bu = 'LPG'
                AND alert_status != 'Close'
                AND interlock_name IN (
                    'Fire Extinguisher Non Compliance (TT)',
                    'PPE non compliance',
                    'Wheel choke non compliance (TT)',
                    'TT Crew non avaibaility near TT',
                    'Position of Truck on weigh bridge',
                    'Detection of rolling of cylinders'
                )
                AND (created_at AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Kolkata')::date =
                    ((CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Kolkata')::date - 1)
                """

    result = await hpcl_ceg_model.Alerts.get_aggr_data(query, limit=0)

    print(f"Total alerts to be closed: {len(result.get('data', []))}")

    if result.get("data", []):
        alert_counts = result["data"]
        for alert in alert_counts:
            alert["alert_status"] = "Close"
            alert["alert_state"] = "Resolved"
            alert["closed_at"] = datetime.datetime.utcnow()
            alert["auto_close"] = True
            alert["alert_history"] = (alert.get("alert_history") or []) + [
                {
                    "action_type": "Resolved",
                    "action_msg": "Auto closed at the end of the day",
                    "alert_status": "Close",
                    "action_by": "SYSTEM",
                    "processed_time": datetime.datetime.now(
                        datetime.timezone.utc
                    ).isoformat(),
                }
            ]
            alert["alert_closure_reason"] = "DAY_END"
            await hpcl_ceg_model.Alerts(**alert).modify()

            try:
                status = await _close_camunda_workflow(alert_data=alert)
                if not status:
                    logger.error(
                        f"Failed to close Camunda workflow for alert {alert.get('id')}"
                    )
            except Exception as e:
                logger.exception(
                    f"Unexpected error while closing Camunda workflow for alert "
                    f"{alert.get('id')}: {str(e)}"
                )

        location_wise_alerts = defaultdict(list)

        for alert in alert_counts:
            sap_id = alert.get("sap_id", "")
            location_wise_alerts[sap_id].append(alert)

        for sap_id, location_alerts in location_wise_alerts.items():
            try:
                location_name = location_alerts[0].get(
                    "location_name", location_alerts[0].get("location", "")
                )
                summary = defaultdict(int)

                for alert in location_alerts:
                    summary[alert["interlock_name"]] += 1

                total_alerts = sum(summary.values())

                html_table = """
                <table class="summary-table">
                <tr>
                    <th>S.No</th>
                    <th>Interlock Name</th>
                    <th>Alerts Closed</th>
                </tr>
                """

                for idx, (interlock, count) in enumerate(summary.items(), start=1):

                    html_table += f"""
                    <tr>
                        <td align="center">{idx}</td>
                        <td>{interlock}</td>
                        <td align="center">{count}</td>
                    </tr>
                    """

                html_table += f"""
                <tr class="total-row">
                    <td colspan="2" align="center"><b>Total</b></td>
                    <td align="center"><b>{total_alerts}</b></td>
                </tr>
                </table>
                """

                safe_sap_id = str(sap_id).replace("/", "_")
                xlsx_file = f"/tmp/VA_Day_End_Closure_{safe_sap_id}.xlsx"

                wb = Workbook()
                ws = wb.active
                ws.title = "VA Alert Closure"

                # Title
                ws.merge_cells("A1:C1")

                title_cell = ws["A1"]
                title_cell.value = "VA Day End Alert Closure Summary"
                title_cell.font = Font(bold=True, size=14, color="FFFFFF")
                title_cell.fill = PatternFill("solid", fgColor="003366")
                title_cell.alignment = Alignment(horizontal="center")

                # Location Details
                ws["A3"] = "Location ID"
                ws["B3"] = sap_id

                ws["A4"] = "Location Name"
                ws["B4"] = location_name

                ws["A5"] = "Generated On"
                ws["B5"] = datetime.datetime.now().strftime("%d-%b-%Y %H:%M:%S")

                header_row = 7

                headers = ["S.No", "Interlock Name", "Alerts Closed"]

                header_fill = PatternFill("solid", fgColor="4472C4")
                header_font = Font(color="FFFFFF", bold=True)

                thin_border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

                for col_num, header in enumerate(headers, start=1):
                    cell = ws.cell(row=header_row, column=col_num)
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = thin_border

                row_num = header_row + 1

                for idx, (interlock, count) in enumerate(summary.items(), start=1):

                    ws.cell(row=row_num, column=1, value=idx)
                    ws.cell(row=row_num, column=2, value=interlock)
                    ws.cell(row=row_num, column=3, value=count)

                    for col in range(1, 4):
                        ws.cell(row=row_num, column=col).border = thin_border

                    row_num += 1

                # Total Row
                ws.merge_cells(
                    start_row=row_num, start_column=1, end_row=row_num, end_column=2
                )

                total_cell = ws.cell(row=row_num, column=1)
                total_cell.value = "Total"
                total_cell.fill = PatternFill("solid", fgColor="E2F0D9")
                total_cell.font = Font(bold=True)
                total_cell.alignment = Alignment(horizontal="center")
                total_cell.border = thin_border

                count_cell = ws.cell(row=row_num, column=3)
                count_cell.value = total_alerts
                count_cell.fill = PatternFill("solid", fgColor="E2F0D9")
                count_cell.font = Font(bold=True)
                count_cell.alignment = Alignment(horizontal="center")
                count_cell.border = thin_border

                # Column Widths
                ws.column_dimensions["A"].width = 12
                ws.column_dimensions["B"].width = 70
                ws.column_dimensions["C"].width = 18

                try:
                    wb.save(xlsx_file)
                except Exception as e:
                    logger.exception(f"Failed to generate Excel for {sap_id}: {str(e)}")
                    continue

                try:
                    await send_closure_notification(
                        html_table=html_table,
                        attachment_file=xlsx_file,
                        total_alerts=total_alerts,
                        sap_id=sap_id,
                        location_name=location_name,
                    )
                    logger.info(
                        f"Mail sent successfully for {sap_id} - {location_name}"
                    )
                except Exception as e:
                    logger.exception(
                        f"Mail sending failed for {sap_id} - {location_name}: {str(e)}"
                    )
            except Exception as e:
                logger.exception(f"Location processing failed for {sap_id}: {str(e)}")
                continue


if __name__ == "__main__":
    import asyncio

    asyncio.run(day_end_closure())
