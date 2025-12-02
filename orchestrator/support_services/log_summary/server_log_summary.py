import urdhva_base
import re
import os
import asyncio
import traceback
import paramiko
import csv
import psycopg2
from psycopg2 import Error
from datetime import datetime, time, timedelta, timezone
from pathlib import Path
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import orchestrator.notification_manager.notification_factory
import orchestrator.dbconnector.credential_loader as credential_loader

creds = credential_loader.get_credentials("HPCL_DEV")

LOG_DIR = "/var/log/ceg_logs"

SLOTS = [
    (time(8,0),"8AM"),
    (time(12,0),"12PM"),
    (time(17,0),"5PM")
]

DATETIME_PATTERN = re.compile(r'^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})')
ERROR_PATTERN = re.compile(r'(ERROR|logger\.error)', re.IGNORECASE)

DB_CONFIG = {
    'host': creds['host'],
    'database': creds['database'],
    'user': creds['user'],
    'password': creds['password'],
    'port': creds['port']
}

log_blocks = []
csv_data = []
error_frequency = {}   # 🔥 NEW for sorting most frequent errors


# -----------------------------------------------------------------------
# SLOT LABEL
# -----------------------------------------------------------------------
def get_current_slot_label():
    now_ist = datetime.now(ZoneInfo("Asia/Kolkata"))
    t = now_ist.time()

    if t >= time(17,0) or t < time(8,0):
        return "5PM"
    elif t >= time(12,0):
        return "5PM"
    elif t >= time(8,0):
        return "12PM"
    else:
        return "8AM"


# -----------------------------------------------------------------------
# CLEANUP OLD DATA
# -----------------------------------------------------------------------
async def cleanup_old_data():
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        cur.execute("""
            DELETE FROM log_summary
            WHERE created_at < (CURRENT_TIMESTAMP - INTERVAL '7 days')
        """)
        conn.commit()

    except Exception as e:
        print("Cleanup error:", e)
    finally:
        try: cur.close()
        except: pass
        try: conn.close()
        except: pass


# -----------------------------------------------------------------------
# INSERT CSV DATA
# -----------------------------------------------------------------------
async def insert_csv_to_database():
    if not csv_data:
        return

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        cur.execute("""
        CREATE TABLE IF NOT EXISTS log_summary (
            id SERIAL PRIMARY KEY,
            server VARCHAR(10),
            log_file VARCHAR(255),
            status VARCHAR(10),
            errors TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""")

        cur.executemany("""
            INSERT INTO log_summary (server, log_file, status, errors)
            VALUES (%s,%s,%s,%s)
        """, [(x["Server"], x["Log File"], x["Status"], x["errors"]) for x in csv_data])

        conn.commit()

    except Exception as e:
        print("Insert DB error:", e)
    finally:
        try: cur.close()
        except: pass
        try: conn.close()
        except: pass


# -----------------------------------------------------------------------
# PARSE LOCAL LOGS
# -----------------------------------------------------------------------
async def log_summary():
    now_ist = datetime.now(ZoneInfo("Asia/Kolkata"))
    today = now_ist.date()

    # SLOT START TIME CALCULATION
    slot_start = None
    slot_label = ""
    for i in range(len(SLOTS)):
        if now_ist.time() < SLOTS[i][0]:
            slot_label = SLOTS[i][1]
            slot_start = datetime.combine(today, SLOTS[i-1][0], tzinfo=ZoneInfo("Asia/Kolkata")) if i > 0 \
                          else datetime.combine(today - timedelta(days=1), SLOTS[-1][0], tzinfo=ZoneInfo("Asia/Kolkata"))
            break
    else:
        slot_label = SLOTS[-1][1]
        slot_start = datetime.combine(today, SLOTS[-1][0], tzinfo=ZoneInfo("Asia/Kolkata"))

    if slot_label == "8AM":
        slot_start = datetime.combine(today - timedelta(days=1), time(17,0), tzinfo=ZoneInfo("Asia/Kolkata"))

    slot_start_utc = slot_start.astimezone(timezone.utc)

    for log_file in Path(LOG_DIR).glob("*.log"):
        errors = []
        try:
            with open(log_file, "r", encoding="utf-8", errors="ignore") as f:
                for line in f:
                    m = DATETIME_PATTERN.match(line)
                    if m:
                        try:
                            t = datetime.strptime(m.group(1), "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
                            if t >= slot_start_utc and ERROR_PATTERN.search(line):
                                errors.append(line.strip())
                                error_frequency[line.strip()] = error_frequency.get(line.strip(), 0) + 1
                        except:
                            continue

            status = "FAIL" if errors else "PASS"

            block = f"Server : 211\nLog File : {log_file.name}\nStatus : {status}\n"
            block += "\n".join(errors) if errors else "(No errors found)"
            log_blocks.append(block)

            unique_errors = list(set(errors))
            csv_data.append({
                "Server": "211",
                "Log File": log_file.name,
                "Status": status,
                "errors": "\n".join(unique_errors) if unique_errors else "(No errors found)"
            })

        except Exception as e:
            msg = f"Could not read file: {e}"
            csv_data.append({"Server": "211", "Log File": log_file.name, "Status": "FAIL", "errors": msg})


# -----------------------------------------------------------------------
# PARSE REMOTE LOGS
# -----------------------------------------------------------------------
async def monitor_remote_logs(server):
    host = f"10.90.38.{server}"
    usr = urdhva_base.settings.novex_user
    pwd = urdhva_base.settings.novex_password
    rlog = "/var/log/ceg_logs"

    now_ist = datetime.now(ZoneInfo("Asia/Kolkata"))
    today = now_ist.date()

    # SLOT CALC
    slot_start = None
    slot_label = ""
    for i in range(len(SLOTS)):
        if now_ist.time() < SLOTS[i][0]:
            slot_label = SLOTS[i][1]
            slot_start = datetime.combine(today, SLOTS[i-1][0], tzinfo=ZoneInfo("Asia/Kolkata")) if i > 0 \
                          else datetime.combine(today - timedelta(days=1), SLOTS[-1][0], tzinfo=ZoneInfo("Asia/Kolkata"))
            break
    else:
        slot_label = SLOTS[-1][1]
        slot_start = datetime.combine(today, SLOTS[-1][0], tzinfo=ZoneInfo("Asia/Kolkata"))

    if slot_label == "8AM":
        slot_start = datetime.combine(today - timedelta(days=1), time(17,0), tzinfo=ZoneInfo("Asia/Kolkata"))

    slot_start_utc = slot_start.astimezone(timezone.utc)

    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    client.connect(host, username=usr, password=pwd)
    sftp = client.open_sftp()

    try:
        for fa in sftp.listdir_attr(rlog):
            if not fa.filename.endswith(".log"):
                continue

            try:
                fp = sftp.file(f"{rlog}/{fa.filename}", "r")
                lines = fp.read().decode(errors="ignore").splitlines()

                errors = []
                for line in lines:
                    m = DATETIME_PATTERN.match(line)
                    if m:
                        try:
                            t = datetime.strptime(m.group(1), "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
                            if t >= slot_start_utc and ERROR_PATTERN.search(line):
                                errors.append(line.strip())
                                error_frequency[line.strip()] = error_frequency.get(line.strip(), 0) + 1
                        except:
                            continue

                status = "FAIL" if errors else "PASS"

                block = f"Server : {server}\nLog File : {fa.filename}\nStatus : {status}\n"
                block += "\n".join(errors) if errors else "(No errors found)"
                log_blocks.append(block)

                csv_data.append({
                    "Server": server,
                    "Log File": fa.filename,
                    "Status": status,
                    "errors": "\n".join(list(set(errors))) if errors else "(No errors found)"
                })

            except Exception as e:
                csv_data.append({
                    "Server": server,
                    "Log File": fa.filename,
                    "Status": "FAIL",
                    "errors": f"Could not read: {e}"
                })

    finally:
        sftp.close()
        client.close()


# -----------------------------------------------------------------------
# GENERATE EXCEL FILE
# -----------------------------------------------------------------------
def generate_excel():
    excel_path = "/tmp/log_summary.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Log Summary"

    header = ["Server", "Log File", "Status", "Errors"]
    ws.append(header)

    # Colors
    red = PatternFill(start_color="FF0000", fill_type="solid")
    green = PatternFill(start_color="00CC00", fill_type="solid")

    for row in csv_data:
        status = row["Status"]
        excel_row = [row["Server"], row["Log File"], row["Status"], row["errors"]]
        ws.append(excel_row)

        fill = green if status == "PASS" else red
        for col in range(1, 5):
            ws.cell(ws.max_row, col).fill = fill

    wb.save(excel_path)
    return excel_path


# -----------------------------------------------------------------------
# SEND EMAIL
# -----------------------------------------------------------------------
async def send_mail(excel_path):
    now_ist = datetime.now(ZoneInfo("Asia/Kolkata"))
    txt_path = "/tmp/log_summary.txt"
    csv_path = "/tmp/log_summary.csv"

    # TXT
    with open(txt_path, "w") as f:
        f.write("\n\n".join(log_blocks))

    # CSV
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["Server","Log File","Status","errors"])
        w.writeheader()
        w.writerows(csv_data)

    formatted = now_ist.strftime("%d-%m-%Y %I:%M %p IST")

    html_body = """
    <html><body>
    <p>Hello,</p>
    <p>Attached is the daily log summary (TXT, CSV, Excel).</p>
    </body></html>
    """

    ins = await orchestrator.notification_manager.notification_factory.get_notification_module("email")

    await ins.publish_message(
        subject=f"Production Servers Log Summary - {formatted}",
        recipients=[
            "sreedhar.maddipati@algofusiontech.com",
            "bala@algofusiontech.com",
            "venu@algofusiontech.com",
            "moufikali@algofusiontech.com",
            "shrihari.b@algofusiontech.com",
            "keerthesrep@algofusiontech.com",
            "vamsi.c@urdhvapay.com",
            "yesu.p@algofusiontech.com",
            "manohar.v@algofusiontech.com",
            "mohith.p@algofusiontech.com",
            "pawann.k@algofusiontech.com"
        ],
        html_content=True,
        body=html_body,
        attachments=[txt_path, csv_path, excel_path],
        force_send=True
    )


# -----------------------------------------------------------------------
# MAIN
# -----------------------------------------------------------------------
async def main():
    await cleanup_old_data()

    servers = {
        "211": "local",
        "212": "remote",
        "217": "remote",
        "218": "remote",
        "219": "remote",
        "222": "remote"
    }

    for srv, typ in servers.items():
        if typ == "local":
            await log_summary()
        else:
            await monitor_remote_logs(srv)

    # Sort errors by frequency (most common on top)
    global csv_data
    csv_data = sorted(csv_data, key=lambda x: (-sum(err in x["errors"] for err in error_frequency), x["Status"]))

    excel_path = generate_excel()

    await insert_csv_to_database()
    await send_mail(excel_path)


if __name__ == "__main__":
    asyncio.run(main())