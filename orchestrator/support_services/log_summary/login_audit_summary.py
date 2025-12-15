import urdhva_base
import asyncio
import traceback
import pandas as pd
import charts_actions
import dashboard_studio_model
import orchestrator.notification_manager.notification_factory as notification_factory
import datetime
import pytz
import hpcl_ceg_model

# openpyxl imports for styling
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class LogAuditSummaryReport:

    EXCEL_FILENAME = '/tmp/login_audit_summary.xlsx'
    EXCEL_FILENAME_SBU_Open_Alerts = '/tmp/sbu_wise_open_or_escalated_alerts_count.xlsx'
    EXCEL_FILENAME_RO_Plant_Wise_Open_Alerts = '/tmp/ro_plant_wise_open_or_escalated_alerts_count.xlsx'
    EXCEL_FILENAME_TAS_Plant_Wise_Open_Alerts = '/tmp/tas_plant_wise_open_or_escalated_alerts_count.xlsx'
    EXCEL_FILENAME_LPG_Plant_Wise_Open_Alerts = '/tmp/lpg_plant_wise_open_or_escalated_alerts_count.xlsx'
    EXCEL_FILENAME_COMBINED_REPORT = '/tmp/user_login_summary_report.xlsx'

    @staticmethod
    def ensure_list(v):
        """Return a list for grouping/exploding. Empty list for NA."""
        if isinstance(v, (list, tuple, set)):
            return list(v)
        if pd.isna(v):
            return []
        return [v]

    async def html_template(self, audit_resp):
        # ---------- Build summary by BU & action_required ----------
        audit_resp['bu_list'] = audit_resp['bu'].apply(self.ensure_list)
        exploded = audit_resp.explode('bu_list', ignore_index=True).rename(columns={'bu_list': 'bu_norm'})
        exploded['bu_norm'] = exploded['bu_norm'].fillna('AdminUsers')

        summary_df = (
            exploded.groupby(['bu_norm', 'action_required'], dropna=False)
            .size()
            .reset_index(name='number_of_users_logged_in')
            .rename(columns={'bu_norm': 'bu'})
            .sort_values(['bu', 'action_required'])
        )

        # Build inline-styled HTML table (no <style> block)
        table_style = (
            "border-collapse:collapse; min-width:500px; font-family:Arial, sans-serif; font-size:14px;"
        )
        th_style = "border:1px solid #e0e0e0; padding:8px 12px; background:#2e7d32; color:#ffffff; font-weight:600; text-align:center;"
        td_style = "border:1px solid #e0e0e0; padding:8px 12px; text-align:center;"
        pill_yes = "display:inline-block;padding:2px 8px;border-radius:12px;font-size:12px;font-weight:600;background:#e8f5e9;color:#1b5e20;border:1px solid #c8e6c9;"
        pill_no = "display:inline-block;padding:2px 8px;border-radius:12px;font-size:12px;font-weight:600;background:#ffebee;color:#b71c1c;border:1px solid #ffcdd2;"

        # header
        html_rows = []
        html_rows.append(f"<table style='{table_style}'>")
        html_rows.append("<thead><tr>")
        html_rows.append(f"<th style='{th_style}'>BU</th>")
        html_rows.append(f"<th style='{th_style}'>Action Required</th>")
        html_rows.append(f"<th style='{th_style}'>Number of Users Logged In</th>")
        html_rows.append("</tr></thead>")
        html_rows.append("<tbody>")

        # rows
        for _, r in summary_df.iterrows():
            bu = str(r['bu'])
            action = str(r['action_required'])
            count = int(r['number_of_users_logged_in'])
            pill = f"<span style='{pill_yes}'>Yes</span>" if action.lower() == "yes" else f"<span style='{pill_no}'>No</span>"
            html_rows.append("<tr>")
            html_rows.append(f"<td style='{td_style}'>{bu}</td>")
            html_rows.append(f"<td style='{td_style}'>{pill}</td>")
            html_rows.append(f"<td style='{td_style}'>{count}</td>")
            html_rows.append("</tr>")

        html_rows.append("</tbody></table>")
        summary_html_table = "".join(html_rows)

        return summary_html_table, summary_df

    def _apply_title(self, ws, title_text: str, ncols: int):
        """Write a merged title in row 1 and style it."""
        title_cell = ws.cell(row=1, column=1, value=title_text)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        # fill for title (light gold)
        title_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        # merge across columns (if more than 1 column)
        if ncols > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)

    def _style_header(self, ws, header_row: int, ncols: int):
        """Style header row (bold, colored background, centered)."""
        header_fill = PatternFill(start_color="184481", end_color="184481", fill_type="solid")  # dark blue
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for col_idx in range(1, ncols + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            # If pandas left the header cell blank (rare), ensure the cell exists
            if cell.value is None:
                # Nothing to do - header text should have been written by pandas.
                pass
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        # set header row height
        ws.row_dimensions[header_row].height = 22

    def _apply_zebra_and_borders(self, ws, first_data_row: int, last_data_row: int, ncols: int):
        """Apply zebra fill and thin borders to data area. Assumes data written already."""
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        fill_even = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        fill_odd = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")  # aliceblue

        if last_data_row < first_data_row:
            # no data rows, still apply border to header-bottom row (single blank row)
            for col_idx in range(1, ncols + 1):
                c = ws.cell(row=first_data_row, column=col_idx)
                c.border = border
            return

        for r_idx in range(first_data_row, last_data_row + 1):
            # choose fill
            is_even = ((r_idx - first_data_row) % 2) == 1
            fill = fill_even if is_even else fill_odd
            for c_idx in range(1, ncols + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.border = border
                # preserve header-like cells if any, else apply fill
                # Only apply fill to data rows (not headers)
                cell.fill = fill

    def _auto_adjust_column_widths(self, ws, df: pd.DataFrame, ncols: int, max_width: int = 60):
        """Auto-set column widths based on header + data (capped)."""
        for i in range(ncols):
            col_letter = get_column_letter(i + 1)
            # compute max length from header and data
            header = str(df.columns[i]) if i < len(df.columns) else ""
            max_len = len(header)
            if len(df) > 0:
                try:
                    max_data_len = df.iloc[:, i].astype(str).map(len).max()
                    if pd.notna(max_data_len):
                        max_len = max(max_len, int(max_data_len))
                except Exception:
                    pass
            # add padding
            width = max_len + 4
            if width > max_width:
                width = max_width
            ws.column_dimensions[col_letter].width = width
    
    async def save_all_reports_to_one_excel(self, audit_resp, sbu_alerts_df, plant_alerts_dict, out_file):
        """
        Write multiple styled sheets using openpyxl via pandas.ExcelWriter(engine='openpyxl').

        USER LOGIN SUMMARY sheet layout:
         - Row 1: Title (merged)
         - Row 2: blank
         - Row 3: summary_df header
         - Row 4+: summary_df rows
         - One blank row
         - Then audit_resp header + rows
        """
        try:
            # Ensure inputs are DataFrames / dict
            if audit_resp is None:
                audit_resp = pd.DataFrame()
            if sbu_alerts_df is None:
                sbu_alerts_df = pd.DataFrame()
            if plant_alerts_dict is None:
                plant_alerts_dict = {}

            # Build summary_df safely by exploding any list-like 'bu' values
            # (same approach as html_template)
            audit_for_summary = audit_resp.copy()
            # ensure column 'bu' exists
            if 'bu' not in audit_for_summary.columns:
                audit_for_summary['bu'] = pd.NA
            audit_for_summary['bu_list'] = audit_for_summary['bu'].apply(self.ensure_list)
            exploded = audit_for_summary.explode('bu_list', ignore_index=True).rename(columns={'bu_list': 'bu_norm'})
            exploded['bu_norm'] = exploded['bu_norm'].fillna('AdminUsers')

            summary_df = (
                exploded.groupby(['bu_norm', 'action_required'], dropna=False)
                .size()
                .reset_index(name='number_of_users_logged_in')
                .rename(columns={'bu_norm': 'bu'})
                .sort_values(['bu', 'action_required'])
            )
            audit_resp.drop(columns=['bu_list'], inplace=True, errors='ignore')
            # start writing sheets
            startrow = 2  # pandas startrow (0-based) so header lands on Excel row 3
            with pd.ExcelWriter(out_file, engine='openpyxl') as writer:
                # ---------- USER LOGIN SUMMARY REPORT ----------
                sheet_name = 'User Login Summary Report'

                # Write summary_df first at startrow
                summary_df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
                ws = writer.sheets[sheet_name]

                ncols_summary = max(1, summary_df.shape[1]) if not summary_df.empty else max(1, audit_resp.shape[1] if not audit_resp.empty else 1)
                title_text = f"User Login Summary Report - {datetime.datetime.now(pytz.timezone('Asia/Kolkata')).strftime('%d-%m-%Y')}"
                self._apply_title(ws, title_text, ncols_summary)

                header_row_summary = startrow + 1  # excel 1-based header row
                self._style_header(ws, header_row_summary, ncols_summary)
                first_data_row_summary = header_row_summary + 1
                last_data_row_summary = header_row_summary + max(0, len(summary_df))
                self._apply_zebra_and_borders(ws, first_data_row_summary, last_data_row_summary, ncols_summary)
                # Auto width based on summary_df (if empty, use audit_resp columns)
                if not summary_df.empty:
                    self._auto_adjust_column_widths(ws, summary_df, ncols_summary)
                else:
                    df_for_width = audit_resp if not audit_resp.empty else pd.DataFrame(columns=[' '])
                    self._auto_adjust_column_widths(ws, df_for_width, ncols_summary)

                # Leave one blank row after summary, then write audit_resp
                audit_startrow = last_data_row_summary + 2
                audit_resp.to_excel(writer, sheet_name=sheet_name, startrow=audit_startrow, index=False, header=True)
                # Note: pandas will create header at audit_startrow (0-based), which corresponds to Excel row audit_startrow+1

                ncols_audit = max(1, audit_resp.shape[1]) if not audit_resp.empty else ncols_summary
                header_row_audit = audit_startrow + 1
                self._style_header(ws, header_row_audit, ncols_audit)
                first_data_row_audit = header_row_audit + 1
                last_data_row_audit = header_row_audit + max(0, len(audit_resp))
                self._apply_zebra_and_borders(ws, first_data_row_audit, last_data_row_audit, ncols_audit)
                # adjust widths to accommodate both tables: compute combined max per-col where possible
                # If audit and summary have same columns count, merge widths; otherwise call auto for audit too.
                try:
                    if not audit_resp.empty:
                        self._auto_adjust_column_widths(ws, audit_resp, ncols_audit)
                except Exception:
                    pass

                # Freeze panes so header of audit_resp remains visible (below title+summary)
                ws.freeze_panes = f"A{first_data_row_audit}"

                # ---------- SBU sheet ----------
                sbu_sheet = 'SBU Wise Open or Escalated Alerts Count'
                sbu_alerts_df.to_excel(writer, sheet_name=sbu_sheet, startrow=startrow, index=False)
                ws_sbu = writer.sheets[sbu_sheet]
                ncols_sbu = max(1, sbu_alerts_df.shape[1]) if not sbu_alerts_df.empty else 1
                self._apply_title(ws_sbu, "SBU Wise Open or Escalated Alerts Count", ncols_sbu)
                header_row_sbu = startrow + 1
                self._style_header(ws_sbu, header_row_sbu, ncols_sbu)
                first_data_row_sbu = header_row_sbu + 1
                last_data_row_sbu = header_row_sbu + max(0, len(sbu_alerts_df))
                self._apply_zebra_and_borders(ws_sbu, first_data_row_sbu, last_data_row_sbu, ncols_sbu)
                self._auto_adjust_column_widths(ws_sbu, sbu_alerts_df if not sbu_alerts_df.empty else pd.DataFrame(columns=sbu_alerts_df.columns), ncols_sbu)
                ws_sbu.freeze_panes = f"A{first_data_row_sbu}"

                # ---------- PLANT SHEETS ----------
                for sbu, df in plant_alerts_dict.items():
                    df = df if df is not None else pd.DataFrame()
                    title_base = f"{sbu} Plant Wise Open or Escalated Alerts Count"
                    sheet_name_plant = title_base[:31]
                    df.to_excel(writer, sheet_name=sheet_name_plant, startrow=startrow, index=False)
                    ws_plant = writer.sheets[sheet_name_plant]
                    ncols_plant = max(1, df.shape[1]) if not df.empty else 1
                    self._apply_title(ws_plant, title_base, ncols_plant)
                    header_row_plant = startrow + 1
                    self._style_header(ws_plant, header_row_plant, ncols_plant)
                    first_data_row_plant = header_row_plant + 1
                    last_data_row_plant = header_row_plant + max(0, len(df))
                    self._apply_zebra_and_borders(ws_plant, first_data_row_plant, last_data_row_plant, ncols_plant)
                    self._auto_adjust_column_widths(ws_plant, df if not df.empty else pd.DataFrame(columns=df.columns), ncols_plant)
                    ws_plant.freeze_panes = f"A{first_data_row_plant}"

            print(f"All reports saved in one Excel file: {out_file}")

        except Exception as e:
            # give clearer debugging information
            print(f"Failed to save combined Excel file: {e}")
            print(traceback.format_exc())



    async def send_notification(self, body, attachment_path):
        """
        Sends the login audit summary report via email with an Excel attachment.
        """
        try:
            ins = await notification_factory.get_notification_module("email")
            recipients = [["cvmallinath@hpcl.in","sreedhar.maddipati@algofusiontech.com","purushm@hpcl.in",
                           "sachinkwarghane@hpcl.in","bala@algofusiontech.com","venu@algofusiontech.com","yesu.p@algofusiontech.com"]]
            IST = pytz.timezone("Asia/Kolkata")
            today_ist = datetime.datetime.now(IST).strftime("%d-%m-%Y")
            for recipient in recipients:
                await ins.publish_message(
                    subject=f"Users Login Daily Report {today_ist}",
                    recipients=recipient,
                    html_content=True,   # no HTML body
                    body=body,
                    force_send=True,
                    attachments=[attachment_path]
                )
            print("Email notification with attachment sent successfully.")
        except Exception as e:
            print("Exception occurred while sending email notification")
            print(e)
            print("Traceback %s" % traceback.format_exc())

    async def sbu_wise_open_alerts_count(self):
        try:
            query = ("""
                     SELECT bu, COUNT(*) AS open_or_escalated_alerts_count
                     FROM alerts
                     WHERE alert_status != 'Close'
                     AND assigned_user_roles != '{}'
                     GROUP BY bu;""")
            dashboard_studio_model.Charts_Connection_Vault_RoutingParams.connection_id = 1
            dashboard_studio_model.Charts_Connection_Vault_RoutingParams.action = 'execute_query'
            function = await charts_actions.charts_connection_vault_routing(dashboard_studio_model.Charts_Connection_Vault_RoutingParams)
            sbu_wise_open_alerts_count_resp = await function(query=query)
            sbu_wise_open_alerts_count_resp = pd.DataFrame(sbu_wise_open_alerts_count_resp)
            return sbu_wise_open_alerts_count_resp
        except Exception as e:
            print("Exception occurred while processing login audit summary")
            print(e)
            print("Traceback %s" % traceback.format_exc())

    async def plant_wise_open_escalated_alerts_count(self, sbu):
        try:
            query = f"""
                        SELECT  
                            sap_id,
                            MAX(location_name) AS location_name,
                            COUNT(*) AS open_or_escalated_alerts_count
                        FROM alerts
                        WHERE alert_status != 'Close'
                        AND bu = '{sbu}'
                        AND assigned_user_roles != '{{}}'
                        GROUP BY sap_id
                        ORDER BY sap_id;
                    """
            dashboard_studio_model.Charts_Connection_Vault_RoutingParams.connection_id = 1
            dashboard_studio_model.Charts_Connection_Vault_RoutingParams.action = 'execute_query'
            function = await charts_actions.charts_connection_vault_routing(dashboard_studio_model.Charts_Connection_Vault_RoutingParams)
            plant_wise_open_alerts_count_resp = await function(query=query)
            plant_wise_open_alerts_count_resp = pd.DataFrame(plant_wise_open_alerts_count_resp)
            return plant_wise_open_alerts_count_resp
        except Exception as e:
            print("Exception occurred while processing login audit summary")
            print(e)
            print("Traceback %s" % traceback.format_exc())

    async def log_audit_summary(self):
        """
        Fetches the login audit summary, saves as Excel, and sends an email notification with attachment.
        """
        try:
            query = ("""
                        SELECT 
                            ula.employee_id,
                            ula.email,
                            u.first_name,
                            u.last_name,
                            u.contact_number,
                            ula.role,
                            u.bu,
                            u.sap_id,
                            u.region,
                            u.zone,
                            u.state,
                            u.sales_area,
                            COUNT(*) AS login_count
                        FROM user_login_audit ula
                        JOIN users u ON ula.employee_id = u.employee_id
                        WHERE
                            (ula.created_at AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Kolkata')::DATE = 
                                (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Kolkata')::DATE
                            AND ula.employee_id NOT IN ('admin', 'Admin', 'superadmin')
                            AND ula.employee_id NOT LIKE '9405\\_%'
                        GROUP BY 
                            ula.employee_id,
                            ula.email,
                            u.first_name,
                            u.last_name,
                            u.contact_number,
                            ula.role,
                            u.bu,
                            u.sap_id,
                            u.region,
                            u.zone,
                            u.state,
                            u.sales_area
                        ORDER BY login_count DESC;
                    """)

            dashboard_studio_model.Charts_Connection_Vault_RoutingParams.connection_id = 1
            dashboard_studio_model.Charts_Connection_Vault_RoutingParams.action = 'execute_query'
            function = await charts_actions.charts_connection_vault_routing(dashboard_studio_model.Charts_Connection_Vault_RoutingParams)
            login_audit_resp = await function(query=query)
            audit_resp = pd.DataFrame(login_audit_resp)

            for idx, record in audit_resp.iterrows():
                conditions = []
                if record.get("bu") and len(record["bu"]) > 0:
                    conditions.append(f"bu = '{record['bu'][0]}'")
                if record.get("sap_id") and len(record["sap_id"]) > 0:
                    conditions.append(f"sap_id = '{record['sap_id'][0]}'")
                if record.get("region") and len(record["region"]) > 0:
                    conditions.append(f"region = '{record['region'][0]}'")
                if record.get("zone") and len(record["zone"]) > 0:
                    conditions.append(f"zone = '{record['zone'][0]}'")
                if record.get("state") and len(record["state"]) > 0:
                    conditions.append(f"state = '{record['state'][0]}'")
                if record.get("sales_area") and len(record["sales_area"]) > 0:
                    conditions.append(f"sales_area = '{record['sales_area'][0]}'")
                conditions.append("alert_status != 'Close'")

                # Join all conditions with ' and '
                where_clause = " and ".join(conditions)

                # Compose the full query string
                query = f"select distinct(assigned_user_roles) from alerts where {where_clause}"

                audit_alerts_data = await hpcl_ceg_model.Alerts.get_aggr_data(query, limit=1000)
                roles = []
                if len(audit_alerts_data.get('data', [])):
                    for data in audit_alerts_data['data']:
                        if len(data.get('assigned_user_roles', [])):
                            for role in data['assigned_user_roles']:
                                if role not in roles:
                                    roles.append(role)

                if record.get('role'):
                    for role in roles:
                        if role in record['role']:
                            audit_resp.at[idx, 'action_required'] = 'yes'
                            break
                    else:
                        audit_resp.at[idx, 'action_required'] = 'No'
                else:
                    audit_resp.at[idx, 'action_required'] = 'No'

            print("Audit response preview:", audit_resp.head())

            if audit_resp.empty:
                print("No login audit data found for today.")
                return

            # Save a simple copy of the audit as well (optional)
            audit_resp.to_excel(self.EXCEL_FILENAME, index=False)

            summary_html_table, summary_df = await self.html_template(audit_resp)
            IST = pytz.timezone("Asia/Kolkata")
            today_ist = datetime.datetime.now(IST).strftime("%d-%m-%Y")
            #html_intro = f"<p>Please find attached the list of users who logged into <b>Novex</b> on <b>{today_ist}</b>.</p>"
            html_intro = f"<p>Please find the attached for the list of users whoever logged in to <b>Novex</b> on <b>{today_ist}</b>.</p>"
            body_html = f"{html_intro}<p><b>Summary by BU &amp; Action Required:</b></p>{summary_html_table}"
            #body_html = f"{html_intro}<p><b>Summary by BU &amp; Action Required:</b></p>{summary_html_table}"

            sbu_alerts_df = await self.sbu_wise_open_alerts_count()
            sbu_lists = ["TAS", "LPG", "RO"]
            plant_alerts = {}
            for sbu in sbu_lists:
                df = await self.plant_wise_open_escalated_alerts_count(sbu)
                plant_alerts[sbu] = df

            await self.save_all_reports_to_one_excel(audit_resp, sbu_alerts_df, plant_alerts, self.EXCEL_FILENAME_COMBINED_REPORT)
            await self.send_notification(body_html, self.EXCEL_FILENAME_COMBINED_REPORT)

        except Exception as e:
            print("Exception occurred while processing login audit summary")
            print(e)
            print("Traceback %s" % traceback.format_exc())


if __name__ == "__main__":
    login_audit_summary = LogAuditSummaryReport()
    asyncio.run(login_audit_summary.log_audit_summary())
