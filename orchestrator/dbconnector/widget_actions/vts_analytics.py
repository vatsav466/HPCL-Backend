import urdhva_base
import io
import os
import json
import asyncio
import traceback
import numpy as np
import pandas as pd
import polars as pl
import hpcl_ceg_model
import mysql.connector
from fastapi import Request
import dashboard_studio_model
from datetime import datetime
import polars.selectors as cs
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from hpcl_ceg_model import DeviceInstallation
from fastapi.responses import StreamingResponse
from dateutil.relativedelta import relativedelta
from fastapi.responses import JSONResponse, FileResponse
from orchestrator.dbconnector.widget_actions import widget_actions
import orchestrator.dbconnector.widget_actions.vts_query as vts_query
import orchestrator.dbconnector.credential_loader as credential_loader
import orchestrator.sync_services.vts.vts_ongoing_trips as vts_ongoing_trips


async def generate_cross_filter(cross_filters):
    _filters, daterange = [], None
    try:
        if cross_filters:
            for f in cross_filters:
                if "DATE" in f.key:
                    start = f.value.split(",")[0]
                    end = f.value.split(",")[-1]
                    daterange = f"'{start} 00:00:00' AND '{end} 23:59:59'"
                else:
                    _filters.append({f.key: f.value})
        return _filters, daterange
    except Exception as e:
        print("--- Exception in cross filters ---")
        print("Exception :", str(e))
        return _filters, daterange

async def get_drill_down_filter(filters, query):
    try:
        conditions = []
        _key = None
        if filters:
            for rec in filters:
                if (rec.key).lower().replace('"', '') in ["rejection_type"]:
                    _key = (rec.value).lower().replace('"', '')
                    continue
                values = rec.value.split(",")
                if len(values) == 1:
                    conditions.append(f'{rec.key} = \'{values[0]}\'')
                else:
                    conditions.append(f"{rec.key} IN {tuple(values)}")
        if conditions:
            if "where" in query.lower():
                query += " AND " + " AND ".join(conditions)
            else:
                query += " WHERE " + " AND ".join(conditions)
        if _key:
            return query, _key
        return query
    except Exception as e:
        print("--- Exception in drill down filters ---")
        print("Exception :", str(e))
        return query

async def filter_data(df, _filters):
    try:        
        if _filters:
            print("-"*30)
            print("_filters :", _filters)
            print("data columns :", df.columns)
            print("length of data :", len(df))
            mask = pd.Series(True, index=df.index)
            for _filter in _filters:
                for key, value in _filter.items():
                    key = key.replace('"','')
                    mask = mask & (df[key].fillna('') == value)
            df = df[mask]
            print("length of filtered data :", len(df))
            print("-"*30)
        return df
    except Exception as e:
        print("Exception in filtering data :", str(e))
    return df

async def safe_json(df):
    return df.replace([np.inf, -np.inf, np.nan], None).to_dict(orient="records")

def extract_invoice_number(invoice_str):
    # Not requried
    """Extract base invoice number: 9017293614-ZF23-1992 -> 9017293614"""
    if pd.isna(invoice_str) or invoice_str is None:
        return None
    invoice_str = str(invoice_str).strip()
    if '-' in invoice_str:
        return invoice_str.split('-')[0]
    return invoice_str


async def get_location_master():
    query = """
        SELECT sap_id, name
        FROM location_master
    """
    df = await VTSAnalyticsActions.execute_query(query, engine="polars")
    return df

async def get_shortage_data(filters, cross_filters, bu_ic, violation_types):
    """Fetch ALL shortage data from sales_trips_till_date"""
    shortage_query = """
                     SELECT vehicle_id, \
                            invoice_no, \
                            plant_nm, \
                            zone_nm, \
                            invoice_date, \
                            destination_code as shortage_destination_code, \
                            COALESCE(NULLIF(TRIM(qty_shortage), 'NaN'), '0.0')::NUMERIC AS qty_shortage, \
                            material_group_nm
                     FROM sales_trips_till_date
                     WHERE load_status = '6' \
                     """

    conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, shortage_query)
    shortage_query = VTSAnalyticsActions.apply_conditions_to_query(shortage_query, conditions)
    df_shortage = await VTSAnalyticsActions.execute_query(shortage_query, engine="polars")

    if df_shortage.is_empty():
        return pl.DataFrame()
    
    # Convert qty_shortage to numeric
    df_shortage = df_shortage.with_columns(
        pl.col('qty_shortage')
        .cast(pl.Float64, strict=False)
        .fill_null(0.0)
        .alias('qty_shortage')
    )

    # Filter out zero shortages for non-I&C BUs when not in default mode
    if not bu_ic and violation_types:
        df_shortage = df_shortage.filter(pl.col("qty_shortage") != 0)
        if df_shortage.is_empty():
            return pl.DataFrame()

    # Create standardized column names for merging
    df_shortage = df_shortage.with_columns(
        invoice_match_key=pl.col("invoice_no").cast(pl.String).str.strip_chars(),
        tl_number=pl.col("vehicle_id"),
        invoice_number=pl.col("invoice_no")
    )

    return df_shortage

async def get_truck_master():
    """Get truck master data for all trucks"""
    truck_query = """
                  SELECT DISTINCT truck_no, \
                                  transporter_name, \
                                  location_name, \
                                  zone
                  FROM vts_truck_master \
                  """
    return await VTSAnalyticsActions.execute_query(truck_query, engine="polars")

def create_shortage_detail(df_shortage: pl.DataFrame, bu_ic: bool, violation_types) -> pl.DataFrame:
    # 1) Clean qty_shortage and material_group_nm
    df_clean = df_shortage.with_columns(
        qty_shortage=(
            pl.col("qty_shortage")
            .cast(pl.Float64, strict=False)  # invalid -> null
            .fill_null(0.0)  # null -> 0
        ),
        material_group_nm=(
            pl.when(pl.col("material_group_nm").is_not_null())
            .then(pl.col("material_group_nm").cast(pl.String).str.strip_chars())
            .otherwise(pl.lit("0"))
        ),
    )

    # 2) Sum shortage per (tl_number, invoice_match_key, material_group_nm)
    per_mat = (
        df_clean
        .group_by(["tl_number", "invoice_match_key", "material_group_nm"])
        .agg(pl.col("qty_shortage").sum().alias("shortage"))
    )

    # 3) Apply "include zeros or only non-zero" rule
    if bu_ic and not violation_types:
        per_mat_filtered = per_mat
    else:
        per_mat_filtered = per_mat.filter(pl.col("shortage") != 0)

    # 4) Build "mat_group:shortage" strings
    per_mat_labeled = per_mat_filtered.with_columns(
        detail=(
                pl.col("material_group_nm").cast(pl.String)
                + pl.lit(":")
                + pl.col("shortage").cast(pl.String)
        )
    )

    # 5) Concatenate details per (tl_number, invoice_match_key)
    df_shortage_grouped = (
        per_mat_labeled
        .group_by(["tl_number", "invoice_match_key"])
        .agg(
            pl.col("detail")
            .str.concat(delimiter=", ")  # join list into a single string
            .alias("qty_shortage_detail")
        )
    )

    return df_shortage_grouped

def determine_merge_strategy(violation_types, bu_ic, bu_tas, bu_lpg):
    """
    Determine merge strategy based on BU and violation_types:
    Returns: (merge_how, base_is_shortage)
    """
    has_violation_filter = bool(violation_types)

    if bu_ic:
        if has_violation_filter:
            return ('inner', True)
        else:
            return ('left', True)
    elif bu_tas or bu_lpg:
        if has_violation_filter:
            return ('left', False)
        else:
            return ('outer', False)
    else:
        return ('left', False)

def merge_shortage_with_violations(
        df_violations, df_shortage, df_truck_master,
        violation_types, bu_ic, bu_tas, bu_lpg,
        all_violations, aggregate=True
):
    """
    Merge shortage and violations based on BU-specific logic
    """
    merge_how, base_is_shortage = determine_merge_strategy(violation_types, bu_ic, bu_tas, bu_lpg)

    if df_shortage.is_empty() and df_violations.is_empty():
        return pl.DataFrame()

    if df_shortage.is_empty():
        if base_is_shortage:
            # If shortage is base and empty, return empty
            return pl.DataFrame()
        else:
            # If violations are base, add empty shortage column
            if aggregate:
                df_violations = df_violations.with_columns(
                    pl.lit(0).cast(pl.Int64).alias("qty_shortage")
                )
            else:
                df_violations = df_violations.with_columns(
                    pl.lit('').alias("qty_shortage_detail")
                )

            # Map with truck master
            if not df_truck_master.is_empty() and not df_violations.is_empty():
                # Prepare lookup: truck_no -> transporter_name
                truck_lookup = (
                    df_truck_master
                    .rename({"truck_no": "tl_number"})  # align join key name
                    .select(["tl_number", "transporter_name"])
                )

                # Ensure transporter_name column exists (like get(..., empty Series))
                if "transporter_name" not in df_violations.columns:
                    df_violations = df_violations.with_columns(
                        transporter_name=pl.lit(None, dtype=pl.String)
                    )

                # Check if we need to fill (any nulls)
                needs_fill = df_violations[
                                 "transporter_name"].null_count() > 0  # cheap metadata op [[Missing data](https://docs.pola.rs/user-guide/expressions/missing-data/#missing-data)]

                if needs_fill:
                    # Join on tl_number to bring in transporter_name from truck master
                    df_violations = (
                        df_violations
                        .join(truck_lookup, on="tl_number", how="left", suffix="_truck")
                        .with_columns(
                            transporter_name=pl.coalesce(
                                "transporter_name",  # existing values
                                "transporter_name_truck",  # from lookup
                            )
                        )
                        .drop("transporter_name_truck")
                    )

            return df_violations

    if df_violations.is_empty():
        if not base_is_shortage:
            # Violations are base but empty
            return pl.DataFrame()

    # Prepare match keys
    if not df_violations.is_empty() and 'invoice_number' in df_violations.columns:
        cleaned_col_expr = pl.col('invoice_number').cast(pl.String).str.strip_chars()
        df_violations = df_violations.with_columns(
            invoice_match_key=pl.when(
                cleaned_col_expr.is_null()
            ).then(
                pl.lit(None)
            ).when(
                cleaned_col_expr.str.contains("-")
            ).then(
                cleaned_col_expr.str.split('-').list.get(0)
            ).otherwise(
                cleaned_col_expr
            )
        )

    df_shortage = df_shortage.with_columns(
        invoice_match_key=pl.col('invoice_no').cast(pl.String).str.strip_chars()
    )

    if aggregate:
        # Aggregate shortage by vehicle + invoice
        df_shortage_agg = (
            df_shortage
            .group_by(["tl_number", "invoice_match_key"])
            .agg(
                pl.col("qty_shortage").sum(),
                pl.col("plant_nm").first(),
                pl.col("zone_nm").first(),
                pl.col("invoice_number").first(),
                pl.col("invoice_date").first(),
                pl.col("shortage_destination_code").first()
            )
        )

        # Perform merge based on strategy
        if base_is_shortage:
            # Shortage is base (I&C)
            if df_violations.is_empty() or 'invoice_number' not in df_violations.columns:
                df_merged = df_shortage_agg.clone()
            else:
                df_merged = df_shortage_agg.join(
                    df_violations,
                    left_on=['tl_number', 'invoice_match_key'],
                    right_on=['tl_number', 'invoice_match_key'],
                    how=merge_how,
                    suffix='_shortage',
                    coalesce=True
                )
                # Use shortage invoice_number when violations invoice_number is null
                if "invoice_number_shortage" in df_merged.columns:
                    df_merged = df_merged.with_columns(
                        invoice_number=pl.col("invoice_number").fill_null(
                            pl.col("invoice_number_shortage")
                        )
                    )
                else:
                    df_merged = df_merged.with_columns(
                        invoice_number=pl.col("invoice_number").fill_null("")
                    )

                if "invoice_date_shortage" in df_merged.columns:
                    df_merged = df_merged.with_columns(
                        invoice_date=pl.col("invoice_date_shortage")
                    )
                elif "invoice_date" in df_merged.columns:
                    df_merged = df_merged.with_columns(
                        invoice_date=pl.col("invoice_date")
                    )
                else:
                    df_merged = df_merged.with_columns(
                        invoice_date=pl.lit("", dtype=pl.String)
                    )
        else:
            # Violations are base (TAS/LPG/Others)
            if df_violations.is_empty():
                df_merged = df_shortage_agg.clone()
            else:
                df_merged = df_violations.join(
                    df_shortage_agg,
                    left_on=['tl_number', 'invoice_match_key'],
                    right_on=['tl_number', 'invoice_match_key'],
                    how=merge_how,
                    suffix='_shortage',
                    coalesce=True
                )

        # Fill missing data from shortage for unmatched records
        if 'location_name' not in df_merged.columns:
            # Column missing → create using plant_nm or blank
            df_merged = df_merged.with_columns(
                location_name=pl.col("plant_nm").fill_null("") if "plant_nm" in df_merged.columns else pl.lit("")
            )
        else:
            # Column exists → fill NaN with plant_nm
            df_merged = df_merged.with_columns(
                location_name=pl.col("location_name").fill_null(pl.col("plant_nm")).fill_null("")
                if "plant_nm" in df_merged.columns else pl.col("location_name").fill_null("")
            )

        if 'zone' not in df_merged.columns:
            df_merged = df_merged.with_columns(
                zone=pl.col("zone_nm").fill_null("") if "zone_nm" in df_merged.columns else pl.lit("")
            )
        else:
            df_merged = df_merged.with_columns(
                zone=pl.col("zone").fill_null(pl.col("zone_nm")).fill_null("")
                if "zone_nm" in df_merged.columns else pl.col("zone").fill_null("")
            )

        needs_fill = (
                "invoice_number" not in df_merged.columns
                or df_merged["invoice_number"].null_count() > 0
        )

        if needs_fill:
            # ensure invoice_number exists (like get(..., ''))
            if "invoice_number" not in df_merged.columns:
                df_merged = df_merged.with_columns(
                    invoice_number=pl.lit("", dtype=pl.String)
                )

            # if invoice_match_key exists, fill nulls from it; else just fill with ""
            if "invoice_match_key" in df_merged.columns:
                df_merged = df_merged.with_columns(
                    invoice_number=pl.col("invoice_number").fill_null(
                        pl.col("invoice_match_key")
                    )
                )
            else:
                df_merged = df_merged.with_columns(
                    invoice_number=pl.col("invoice_number").fill_null("")
                )

        # Handle created_at/violation_date with fallback to invoice_date
        if 'created_at' not in df_merged.columns:
            if 'violation_date' in df_merged.columns:
                df_merged = df_merged.with_columns(
                    created_at=pl.col("violation_date").fill_null(
                        pl.col("invoice_date")
                    )
                )
            else:
                df_merged = df_merged.with_columns(
                    created_at=pl.col("invoice_date").fill_null("")
                )
        elif "created_at" in df_merged.columns and df_merged["created_at"].null_count() > 0:
            # decide fallback expression: violation_date if present, else invoice_date, else ""
            if "violation_date" in df_merged.columns:
                fallback = pl.col("violation_date")
            elif "invoice_date" in df_merged.columns:
                fallback = pl.col("invoice_date")
            else:
                fallback = pl.lit("", dtype=pl.String)

            df_merged = df_merged.with_columns(
                created_at=pl.col("created_at").fill_null(fallback)
            )

        df_merged = df_merged.with_columns(
            created_at=(
                pl.col("created_at")
                .cast(pl.String)
                .str.to_datetime(format="%Y%m%d", strict=False)
                .dt.strftime("%Y-%m-%d")
            ),
            qty_shortage=(
                pl.col("qty_shortage")
                .cast(pl.Float64, strict=False)
                .fill_null(0)
            ),
        )

    else:
        df_shortage_grouped = create_shortage_detail(
            df_shortage, bu_ic, violation_types
        )

        # Get metadata
        df_shortage_meta = (
            df_shortage
            .group_by(["tl_number", "invoice_match_key"])
            .agg(
                pl.col("plant_nm").first(),
                pl.col("zone_nm").first(),
                pl.col("invoice_number").first(),
                pl.col("invoice_date").first(),
            )
        )

        df_shortage_grouped = df_shortage_grouped.join(
            df_shortage_meta,
            on=['tl_number', 'invoice_match_key'],
            how='left', coalesce=True
        )

        # Perform merge based on strategy
        if base_is_shortage:
            # Shortage is base (I&C)
            if df_violations.is_empty() or 'invoice_number' not in df_violations.columns:
                df_merged = df_shortage_grouped.clone()
            else:
                df_merged = df_shortage_grouped.join(
                    df_violations,
                    left_on=['tl_number', 'invoice_match_key'],
                    right_on=['tl_number', 'invoice_match_key'],
                    how=merge_how,
                    suffix='_shortage',
                    coalesce=True
                )
                if "invoice_number_shortage" in df_merged.columns:
                    df_merged = df_merged.with_columns(
                        invoice_number=pl.col("invoice_number").fill_null(
                            pl.col("invoice_number_shortage")
                        )
                    )
                else:
                    df_merged = df_merged.with_columns(
                        invoice_number=pl.col("invoice_number").fill_null("")
                    )

                if "invoice_date_shortage" in df_merged.columns:
                    df_merged = df_merged.with_columns(
                        invoice_date=pl.col("invoice_date_shortage")
                    )
                elif "invoice_date" in df_merged.columns:
                    df_merged = df_merged.with_columns(
                        invoice_date=pl.col("invoice_date")
                    )
                else:
                    df_merged = df_merged.with_columns(
                        invoice_date=pl.lit("", dtype=pl.String)
                    )
        else:
            # Violations are base (TAS/LPG/Others)
            if df_violations.is_empty():
                df_merged = df_shortage_grouped.clone()
            else:
                df_merged = df_violations.join(
                    df_shortage_grouped,
                    left_on=['tl_number', 'invoice_match_key'],
                    right_on=['tl_number', 'invoice_match_key'],
                    how=merge_how,
                    suffix='_shortage',
                    coalesce=True
                )

        # Fill missing data
        if 'location_name' not in df_merged.columns:
            # Column missing → create using plant_nm or blank
            df_merged = df_merged.with_columns(
                location_name=pl.col("plant_nm").fill_null("") if "plant_nm" in df_merged.columns else pl.lit("")
            )
        else:
            df_merged = df_merged.with_columns(
                location_name=pl.col("location_name").fill_null(pl.col("plant_nm"))
                if "plant_nm" in df_merged.columns else pl.col("location_name").fill_null("")
            )

        if 'zone' not in df_merged.columns:
            df_merged = df_merged.with_columns(
                zone=pl.col("zone_nm").fill_null("") if "zone_nm" in df_merged.columns else pl.lit("")
            )
        else:
            df_merged = df_merged.with_columns(
                zone=pl.col("zone").fill_null(pl.col("zone_nm"))
                if "zone_nm" in df_merged.columns else pl.col("zone").fill_null("")
            )

        needs_fill = (
                "invoice_number" not in df_merged.columns
                or df_merged["invoice_number"].null_count() > 0
        )

        if needs_fill:
            # ensure invoice_number exists (like get('invoice_number', ''))
            if "invoice_number" not in df_merged.columns:
                df_merged = df_merged.with_columns(
                    invoice_number=pl.lit("", dtype=pl.String)
                )

            # if invoice_match_key exists, fill nulls from it; else just fill with ""
            if "invoice_match_key" in df_merged.columns:
                df_merged = df_merged.with_columns(
                    invoice_number=pl.col("invoice_number").fill_null(
                        pl.col("invoice_match_key")
                    )
                )
            else:
                df_merged = df_merged.with_columns(
                    invoice_number=pl.col("invoice_number").fill_null("")
                )

        # Handle created_at/violation_date with fallback to invoice_date
        if 'created_at' not in df_merged.columns:
            if 'violation_date' in df_merged.columns:
                df_merged = df_merged.with_columns(
                    created_at=pl.col("violation_date").fill_null(
                        pl.col("invoice_date") if "invoice_date" in df_merged.columns else pl.lit("")
                    )
                )
            else:
                df_merged = df_merged.with_columns(
                    created_at=pl.col("invoice_date") if "invoice_date" in df_merged.columns else pl.lit("")
                )
        elif "created_at" in df_merged.columns and df_merged["created_at"].null_count() > 0:
            # decide fallback: violation_date if present, else invoice_date, else ""
            if "violation_date" in df_merged.columns:
                fallback = pl.col("violation_date")
            elif "invoice_date" in df_merged.columns:
                fallback = pl.col("invoice_date")
            else:
                fallback = pl.lit("", dtype=pl.String)

            df_merged = df_merged.with_columns(
                created_at=pl.col("created_at").fill_null(fallback)
            )

        df_merged = df_merged.with_columns(
            created_at=(
                pl.col("created_at")
                .cast(pl.String)
                .str.to_datetime(format="%Y%m%d", strict=False)
                .dt.strftime("%Y-%m-%d")
            ),
            qty_shortage_detail=(
                pl.col("qty_shortage_detail")
                .cast(pl.String)
                .fill_null("")
            ),
        )

    # Map with truck master for missing transporter_name, location_name, zone
    if not df_truck_master.is_empty() and not df_merged.is_empty():
        # Prepare truck lookup keyed by tl_number
        truck_lookup = (
            df_truck_master
            .rename({"truck_no": "tl_number"})
            .select(["tl_number", "transporter_name", "location_name", "zone"])
        )

        # Ensure transporter_name exists if needed
        if "transporter_name" not in df_merged.columns:
            df_merged = df_merged.with_columns(
                transporter_name=pl.lit(None, dtype=pl.String)
            )

        # Join once to bring in truck info
        df_merged = df_merged.join(
            truck_lookup,
            on="tl_number",
            how="left",
            suffix="_truck",
        )  # gives transporter_name_truck, location_name_truck, zone_truck

        # Fill from truck master where still missing
        df_merged = df_merged.with_columns(
            transporter_name=pl.coalesce(
                "transporter_name", "transporter_name_truck"
            ),
            location_name=pl.coalesce(
                "location_name", "location_name_truck"
            ),
            zone=pl.coalesce(
                "zone", "zone_truck"
            ),
        )

        # Drop helper columns
        df_merged = df_merged.drop(
            ["transporter_name_truck", "location_name_truck", "zone_truck"]
        )

    cols_to_drop = [
        "invoice_match_key",
        "plant_nm",
        "zone_nm",
        "invoice_number_shortage",
        "invoice_date",
        "invoice_date_shortage",
    ]

    existing = [c for c in cols_to_drop if c in df_merged.columns]
    if existing:
        df_merged = df_merged.drop(existing)

    exprs = []
    for v in all_violations:
        if v in df_merged.columns:
            exprs.append(pl.col(v).fill_null(0).alias(v))

    if exprs:
        df_merged = df_merged.with_columns(*exprs)

    return df_merged

async def download_streaming_data(df: pl.DataFrame, filename='violations'):

    df = df.with_columns(
        cs.datetime(time_zone="*").dt.replace_time_zone(None)
    )

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"{filename}_{timestamp}.xlsx"

    output = io.BytesIO()
    df.write_excel(
        workbook=output,  # BytesIO
        worksheet=f"{filename}",  # sheet name
    )
    output.seek(0)
    headers = {
        "Content-Disposition": f'attachment; filename="{file_name}"'
    }
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )

async def streaming_data(df: pl.DataFrame):
    batch_size = 10000
    total = df.height

    async def json_generator():
        for i in range(0, total, batch_size):
            yield json.dumps(df.slice(i, batch_size).to_dicts())
            await asyncio.sleep(0.1)

    return StreamingResponse(
        json_generator(),
        media_type="application/json"
    )

class VTSAnalyticsActions:
    @staticmethod
    def transform_key(key, query=None):
        """Transform keys based on query context"""
        # if query and any(x in query.lower() for x in ["vts_alert_history", "vts_ongoing_trips"]) and key.lower() == "bu":
        #     return "location_type"
        # if query and "vts_alert_history" in query.lower() and key.lower() == "sap_id":
        #     return "location_id" 
        if query and 'sales_trips_till_date' in query.lower() and key.lower() in ('transporter_code','bu'):
            return None
         
        if query and 'sales_trips_till_date' in query.lower() and key.lower() == 'zone':
            return 'zone_nm'
        
        if query and 'sales_trips_till_date' in query.lower() and key.lower() == 'location_name':
            return 'plant_nm'
        
        return key
    
    @staticmethod
    def build_filter_conditions(filters, cross_filters, query):
        """Build WHERE conditions from filters and cross_filters"""
        all_conditions = []
        
        if 'sales_trips_till_date' in query.lower():
            bu_filter = None
            if filters:
                for f in filters:
                    if hasattr(f, 'key') and f.key.lower() == 'bu':
                        bu_filter = f
                        break
            
            bu_value = bu_filter.value if bu_filter else None

            if bu_value:
                bu = bu_value.upper()
                if bu == 'TAS':
                    all_conditions.append("division = '11'")
                    all_conditions.append("sales_org = '7000'")
                    all_conditions.append("(qty_shortage > '0')")
                elif bu == 'LPG':
                    all_conditions.append("division in ('20', '21')")
                    all_conditions.append("sales_org = '2000'")
                    all_conditions.append("(qty_shortage > '0')")
                elif bu == 'I&C':
                    all_conditions.append("distribution_channel = '12'")
                    all_conditions.append("sales_org = '3000'")
                    all_conditions.append("route <> 'EXW001'")
            
            else:
                all_conditions.append("division = '11'")
                all_conditions.append("sales_org = '7000'")
        
        elif 'sales_trips_till_date' not in query.lower():
            bu_filter = None
            if filters:
                for f in filters:
                    if hasattr(f, 'key') and f.key.lower() == 'bu':
                        bu_filter = f
                        break
            
            if bu_filter and bu_filter.value.upper() == 'I&C':
                all_conditions.append("bu = 'TAS'")
               
        # Process regular filters
        if filters:
            for rec in filters:
                key = VTSAnalyticsActions.transform_key(rec.key, query)
                if key is None:
                    continue

                if (key.lower() == 'bu' and 'sales_trips_till_date' not in query.lower() and
                     rec.value.upper() == 'I&C'):
                   continue

                val = rec.value
                
                condition = VTSAnalyticsActions.create_condition(key, val)
                if condition:
                    all_conditions.append(condition)

        # Process cross filters
        if cross_filters:
            for rec in cross_filters:
                key = rec.key
                val = rec.value

                if "DATE" in key.upper():
                    condition = VTSAnalyticsActions.create_date_condition(query,val)
                else:
                    condition = VTSAnalyticsActions.create_condition(key, val)
                
                if condition:
                    all_conditions.append(condition)
        
        return all_conditions

    @staticmethod
    def create_condition(key, val):
        """Create a single condition based on key and value"""
        if isinstance(val, str):
            if "," in val:  # Handle comma-separated values in string
                values = val.split(",")
                if len(values) == 1:
                    return f"{key} = '{values[0]}'"
                else:
                    return f"{key} IN {tuple(values)}"
            return f"{key} = '{val}'"
        elif isinstance(val, list):
            if len(val) == 1:
                return f"{key} = '{val[0]}'"
            else:
                return f"{key} IN {tuple(val)}"
        return None

    @staticmethod
    def create_date_condition(query,val):
        """Create date range condition"""
        start = val.split(",")[0]
        end_date = val.split(",")[-1]
        end = f"{end_date} 23:59:59"
       
        if "vts_alert_history" in query.lower():
            return f"vts_end_datetime BETWEEN '{start}' AND '{end}'"
        
        if "vts_tripauditmaster" in query.lower():
            return f"createdat BETWEEN '{start}' AND '{end}'"
        
        onging_trips = ["violation_type = 'wr'", "violation_type = 'tc'", "violation_type = 'hs'"]
        if any(ot in query.lower() for ot in onging_trips):
            return f"event_start_datetime BETWEEN '{start}' AND '{end}'"
        
        if "violation_type = 'rd'" in query.lower():
            return f"event_end_datetime BETWEEN '{start}' AND '{end}'"
        
        if "sales_trips_till_date" in query.lower():
            return f"invoice_date::DATE BETWEEN '{start}' AND '{end}'"
        
        if "completed_trips_risk_score" in query.lower():
            return f"scheduled_trip_start_datetime BETWEEN '{start}' AND '{end}'"
                
        queries = ["vts_device_removed", "vts_harsh_acceleration", "vts_harsh_braking", "vts_panic"]
        if any(q in query.lower() for q in queries):
            return f"event_date BETWEEN '{start}' AND '{end}'"
        
        
        return f"created_at BETWEEN '{start}' AND '{end}'"
    

    @staticmethod
    def apply_conditions_to_query(query, conditions):
        """Apply WHERE conditions to query while preserving GROUP BY and ORDER BY"""
        if not conditions:
            return query
        
        conditions_str = " AND ".join(conditions)
        query_lower = query.lower()

        if ") as history_data" in query_lower:
            # Split into two parts: before and after the subquery alias
            idx = query_lower.index(") as history_data")
            subquery_part = query[:idx]  # everything inside the subquery
            rest_part = query[idx:]      # the alias + group by etc.

            # Insert conditions inside the subquery WHERE clause
            if "where" in subquery_part.lower():
                subquery_part = subquery_part.rstrip() + f" AND {conditions_str}"
            else:
                subquery_part = subquery_part.rstrip() + f" WHERE {conditions_str}"

            return subquery_part + rest_part
        
        # Handle queries with GROUP BY
        if "group by" in query_lower:
            idx = query_lower.index("group by")
            base_part = query[:idx].strip()
            group_by_part = query[idx:].strip()
            
            if "where" not in base_part.lower():
                return f"{base_part} WHERE {conditions_str} {group_by_part}"
            else:
                return f"{base_part} AND {conditions_str} {group_by_part}"
        
        # Handle queries with ORDER BY (no GROUP BY)
        elif "order by" in query_lower:
            idx = query_lower.index("order by")
            base_part = query[:idx].strip()
            order_by_part = query[idx:].strip()
            
            if "where" not in base_part.lower():
                return f"{base_part} WHERE {conditions_str} {order_by_part}"
            else:
                return f"{base_part} AND {conditions_str} {order_by_part}"
        
        # Simple query without GROUP BY or ORDER BY
        else:
            if "where" not in query_lower:
                return f"{query} WHERE {conditions_str}"
            else:
                return f"{query} AND {conditions_str}"

    @staticmethod
    def add_alert_type_conditions(conditions, alert_type):
        """Add alert type specific conditions"""
        if not alert_type:
            return conditions
            
        if alert_type.lower() == "blocked":
            conditions.append("alert_status = 'Open' and vehicle_unblocked_date is null") 
        if alert_type.lower() == "acceptance_close":
            conditions.append("alert_status = 'Close' and vehicle_unblocked_date is null")
        elif alert_type.lower() == "auto_unblock":
            conditions.append("alert_status = 'Close' AND mark_as_false = false AND vehicle_unblocked_date is not null")
        elif alert_type.lower() == "manual_unblock":
            conditions.append("alert_status = 'Close' AND mark_as_false = true and vehicle_unblocked_date is not null")
        else:
        # all_alerts or unknown value → no extra condition
           pass
     
        return conditions

    @staticmethod
    def get_period_expression(drill_state):
        """Get period expression based on drill state"""
        if drill_state and drill_state.lower() == "day_wise":
            return "DATE(created_at)"
        elif drill_state and drill_state.lower() == "month_wise":
            return "DATE_TRUNC('month', created_at)"
        return ""

    @staticmethod
    def get_group_by_column(drill_state):
        """Get group by column based on drill state"""
        if drill_state and "location" in drill_state.lower():
            return "location_name"
        elif drill_state and "zone" in drill_state.lower():
            return "zone"
        return None

    @staticmethod
    def format_date(period, drill_state):
        """Format date based on drill state"""
        if drill_state and drill_state.lower() in ["day_wise", "month_wise"]:
            return pd.to_datetime(period).strftime("%b-%d-%Y")
        return str(period)

    @staticmethod
    async def execute_query(query, limit=0, engine='pandas'):
        """Execute query and return DataFrame"""
        try:
            resp = await urdhva_base.BasePostgresModel.get_aggr_data(query=query, limit=limit)
            data = resp.get("data", [])
            if engine == 'polars':
                return pl.DataFrame(data) if data else pl.DataFrame()
            if engine == 'dict':
                return data
            return pd.DataFrame(data) if data else pd.DataFrame()
        except Exception as e:
            print(f"Query execution error: {e}")
            if engine == 'polars':
                return pl.DataFrame()
            if engine == 'dict':
                return []
            return pd.DataFrame()

    @staticmethod
    async def vts_card_chart(filters, cross_filters, drill_state, payload):    
        try:
            # Get base query           
            card_query = vts_query.vts_query.get(drill_state.split(",")[0])

            # Build and apply conditions (pass the query for key transformation)
            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, card_query)
            final_query = VTSAnalyticsActions.apply_conditions_to_query(card_query, conditions)

            # Execute query
            df = await VTSAnalyticsActions.execute_query(final_query, engine='dict')
            return {"status": True, "message": "success", "data": df}

        except Exception as e:
            print("Exception in BigNumber Chart:", str(e))
            print("traceback:", traceback.format_exc())
            return {"status": False, "message": str(e), "data": []}
    
    @staticmethod
    async def vts_dashboard_card_download(filters, cross_filters, drill_state, payload):
        """
            Download VTS dashboard cards as an Excel file.

            This function retrieves the required data from the database and generates
            an Excel file containing the selected columns for download.

            :param filters: Filters applied based on location, SAP ID, and BU for data filtering.
            :param cross_filters: Additional filters, primarily used for date range selection.
            :param drill_state: Used to map and select the appropriate query
                                from the `vts_query` file in the same path.
            :param payload: Additional parameters passed from the frontend
                            for applying extra conditional checks if required.
            :return: Excel file response as a downloadable stream.
            :rtype: dict[str, Any] | StreamingResponse
        """

        try:
            # 1. Build and execute query
            download_card_query = vts_query.vts_query.get(drill_state.split(",")[0])
            conditions = VTSAnalyticsActions.build_filter_conditions(
                filters, cross_filters, download_card_query
            )
            final_query = VTSAnalyticsActions.apply_conditions_to_query(
                download_card_query, conditions
            )

            df = await VTSAnalyticsActions.execute_query(final_query, engine="polars")

            # 2. Extract creator_id and approver_id from alert_history (JSONB)
            df = df.with_columns([
                    # Creator ID → first employee_id where action_type == Justification
                    pl.col("alert_history").list.eval(pl.element().struct.field("employee_id").filter(
                            pl.element().struct.field("action_type") == "Justification")
                    ).list.first().alias("creator_id"),

                    # Approver ID → first employee_id where action_type == Approved
                    pl.col("alert_history").list.eval(pl.element().struct.field("employee_id").filter(
                            pl.element().struct.field("action_type") == "Approved")
                    ).list.first().alias("approver_id"),
                ])


            # 3. Select required columns for Excel
            df = (df.select(["zone", "sap_id", "location_name",  "vehicle_number", "violation_type",
                            "unique_id", "alert_status", "device_name", "severity", "created_at", 
                            "creator_id", "approver_id", "vehicle_unblocked_date", "vehicle_blocked_end_date" 
                ])
                .rename({
                    "unique_id": "Alert ID",
                    "sap_id" : "Location ID",
                    "vehicle_number" : "Truck Number",
                    "device_name" : "Instance ID",
                    "creator_id": "Creator ID",
                    "approver_id": "Approver ID",
                })
            )
            # 4. Return Excel download
            return await download_streaming_data(df, filename="itdgAlerts")
        
        except Exception as e:
            print("traceback:", traceback.format_exc())
            print("error",str(e))
            return {"status": False, "message": str(e), "data": []}

    @staticmethod
    async def pagination_df(df, payload):
        total_count = len(df)

        if str(payload.get("download", "")).lower() == "true":
            merged_df = pd.DataFrame(df)
            for col in merged_df.select_dtypes(include=["datetime64[ns, UTC]", "datetimetz"]).columns:
                merged_df[col] = merged_df[col].dt.tz_localize(None)

            merged_df = merged_df.dropna(axis=1, how="all")
            merged_df = merged_df.loc[:, (merged_df.astype(str).apply(lambda x: x.str.strip() != "").any())]

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_name = f"violations_{timestamp}.xlsx"

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                merged_df.to_excel(writer, index=False, sheet_name='violations')

            output.seek(0)
            headers = {
                "Content-Disposition": f'attachment; filename="{file_name}"'
            }
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers=headers
            )
        
        skip = int(payload.get("skip", 0))        # page number
        limit = int(payload.get("limit", 20))     # page size
        search_value = str(payload.get("search", "")).strip().lower()
        
        if search_value:
            mask = df.astype(str).apply(
                lambda col: col.str.lower().str.startswith(search_value, na=False)
                ).any(axis=1)
            df = df[mask]

        start = skip * limit
        end = start + limit

        paginated_df = df.iloc[start:end]

        return {
            "status": True,
            "message": "success",
            "total_count": total_count,
            "data": await safe_json(paginated_df)
        }

    @staticmethod
    async def vts_insite(filters, cross_filters, drill_state, payload):
        try:
            query = vts_query.vts_query.get(drill_state.split(",")[0])
            alert_type = payload.get("alert_type") if payload else None
    
            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, query)
            conditions = VTSAnalyticsActions.add_alert_type_conditions(conditions, alert_type)
            vts_insite_query = VTSAnalyticsActions.apply_conditions_to_query(query, conditions)
            print(vts_insite_query)
            df1 = await VTSAnalyticsActions.execute_query(vts_insite_query, engine='polars')

            truck_master_query = """SELECT truck_no, transporter_name FROM vts_truck_master"""
            df_truck_master = await VTSAnalyticsActions.execute_query(truck_master_query, engine='polars')

            merged_df = df1.join(
                    df_truck_master,
                    left_on="vehicle_number",
                    right_on="truck_no",
                    how="left"
                )
            
            merged_df = merged_df.drop(["truck_no"], strict=False)

            if payload.get("download") == "true":
                return await download_streaming_data(merged_df,filename='itdgAlerts')

            return await streaming_data(merged_df)

        except Exception as e:
            print("traceback:", traceback.format_exc())
            return {"status": False, "message": str(e), "data": []}

    @staticmethod
    async def vts_insite_violation(filters, cross_filters, drill_state, payload):
        """
        Main function to handle VTS violation queries with BU-specific merge logic:
        - I&C: Default=Shortage LEFT Violations, Filtered=Shortage INNER Violations
        - TAS/LPG: Default=Violations OUTER Shortage, Filtered=Violations LEFT Shortage
        - Others: Violations LEFT Shortage (always)
        """
        try:
            # Identify Business Unit
            bu_ic = any(getattr(f, 'key') == 'bu' and getattr(f, 'value') == 'I&C' for f in filters)
            bu_lpg = any(getattr(f, 'key') == 'bu' and getattr(f, 'value') == 'LPG' for f in filters)
            bu_tas = any(getattr(f, 'key') == 'bu' and getattr(f, 'value') == 'TAS' for f in filters)

            query = vts_query.vts_query.get(drill_state.split(",")[0])
            all_violations = vts_query.vts_query.get("all_violations", [])
            violation_types = payload.get("violation_type", []) if payload else []
          

            # Get truck master data (used across all flows)
            df_truck_master = await get_truck_master()

            # Get shortage data and filter for this truck
            df_shortage = await get_shortage_data(filters, cross_filters, bu_ic, violation_types)

            # ==================== DEFAULT CASE - ALL DATA ====================
            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, query)
            query = VTSAnalyticsActions.apply_conditions_to_query(query, conditions)
            df_history = await VTSAnalyticsActions.execute_query(query, engine="polars")

            if not df_history.is_empty and 'invoice_number' in df_history.columns:
                df_history = df_history.unique(subset=["invoice_number"], keep="first")

            
            # Merge
            final_df = merge_shortage_with_violations(
                df_history, df_shortage, df_truck_master,
                violation_types, bu_ic, bu_tas, bu_lpg,
                all_violations, aggregate=True
            )

            if "shortage_destination_code" in final_df.columns:
                final_df = final_df.with_columns(
                    destination_code=pl.coalesce([
                        pl.col("shortage_destination_code"),  # Prefer shortage
                        pl.col("destination_code")             # Fallback to violation
                    ])
                ).drop("shortage_destination_code")
            
            loc_df = await get_location_master()
            final_df = final_df.join(
                loc_df.select([
                    pl.col("sap_id").alias("destination_code"),
                    pl.col("name").alias("destination_name")
                ]),
                on="destination_code",
                how="left"
            )
             
            final_df = final_df.with_columns(
              pl.col("destination_name").fill_null("N/A")
            )
            
            if final_df.is_empty():
                return {"status": True, "message": "No data found", "data": []}
            # Ensure qty_shortage column exists
            if 'qty_shortage' not in final_df.columns:
                final_df = final_df.with_columns(
                    pl.lit(0).cast(pl.Int64).alias("qty_shortage")
                )
            else:
                final_df = final_df.with_columns(
                    qty_shortage=(
                        pl.col("qty_shortage")
                        .cast(pl.Float64, strict=False)
                        .fill_null(0)
                    )
                )

            # Filter out rows with no violations at all
            violation_cols = [v for v in all_violations if v in final_df.columns]
            violation_cols.append('qty_shortage')

            final_df = final_df.with_columns(total_violations=pl.sum_horizontal(pl.col(violation_cols)))

            if bu_ic and not violation_types:
                # For I&C default, include all rows (including zero violations)
                pass
            else:
                final_df = final_df.filter(pl.col("total_violations") > 0)

            if 'total_violations' in final_df.columns:
                final_df = final_df.drop("total_violations")

            if final_df.is_empty():
                return {"status": True, "message": "No violation data found", "data": []}

            # ==================== HANDLE GROUP_BY ====================
            group_by_col = payload.get("group_by") if payload else None
            if group_by_col and group_by_col in final_df.columns:
                violation_cols = [v for v in all_violations if v in final_df.columns]
                # violation_cols.append('qty_shortage')
                agg_df = (
                    final_df
                    .group_by(group_by_col)
                    .agg([pl.col(violation_cols).sum()])
                )

                # add total_count as horizontal sum across violation columns
                agg_df = agg_df.with_columns(
                    total_count=pl.sum_horizontal(pl.col(violation_cols))
                )

                return {
                    "status": True,
                    "message": "success",
                    "data": agg_df.to_dicts()
                }

            # ==================== HANDLE DRILL-DOWN ====================
            qlick_view = payload.get("qlick_view") if payload else None
            click_value = payload.get("click_value") if payload else None
            location_name = payload.get("location_name") if payload else None

            violation_cols = [v for v in all_violations if v in final_df.columns]
            violation_cols.append('qty_shortage')

            # ZONE VIEW
            if qlick_view == "zone" and not click_value:
                final_df = final_df.with_columns(zone=pl.col("zone").fill_null("Unknown"))
                agg_df = (
                    final_df
                    .group_by("zone")
                    .agg([pl.col(violation_cols).sum()])
                )
                agg_df = agg_df.with_columns(total_count=pl.sum_horizontal(pl.col(violation_cols)))
                return {
                    "status": True,
                    "message": "Zone-wise violations",
                    "data": agg_df.to_dicts()
                }

            # ZONE -> LOCATION DRILL
            if qlick_view == "zone" and click_value:
                final_df = final_df.filter(pl.col("zone") == click_value)
                if final_df.is_empty():
                    return {
                        "status": True,
                        "message": f"No data found for zone {click_value}",
                        "data": []
                    }

                agg_df = (
                    final_df
                    .group_by("location_name")
                    .agg([pl.col(violation_cols).sum()])
                )
                agg_df = agg_df.with_columns(total_count=pl.sum_horizontal(pl.col(violation_cols)))

                return {
                    "status": True,
                    "message": f"Violations for all plants in zone {click_value}",
                    "data": agg_df.to_dicts()
                }

            # LOCATION -> TRANSPORTER DRILL
            elif qlick_view == "location_name" and click_value:
                final_df = final_df.filter(pl.col("location_name") == click_value)
                if final_df.is_empty():
                    return {
                        "status": True,
                        "message": f"No data found for location {click_value}",
                        "data": []
                    }
                agg_df = (
                    final_df
                    .group_by("transporter_name")
                    .agg([pl.col(violation_cols).sum()])
                )
                agg_df = agg_df.with_columns(total_count=pl.sum_horizontal(pl.col(violation_cols)))

                return {
                    "status": True,
                    "message": f"Violations for all transporters in location {click_value}",
                    "data": agg_df.to_dicts()
                }

            # TRANSPORTER -> VEHICLE DRILL
            elif qlick_view == "transporter_name" and click_value and location_name:
                final_df = final_df.filter(
                    (pl.col("transporter_name") == click_value) &
                    (pl.col("location_name") == location_name)
                )
                if final_df.is_empty():
                    return {
                        "status": True,
                        "message": f"No data found for transporter {click_value}",
                        "data": []
                    }
                agg_df = (
                    final_df
                    .group_by("tl_number")
                    .agg([pl.col(violation_cols).sum()])
                )
                agg_df = agg_df.with_columns(total_count=pl.sum_horizontal(pl.col(violation_cols)))

                return {
                    "status": True,
                    "message": f"Vehicle-wise violations for transporter {click_value}",
                    "data": agg_df.to_dicts()
                }

            # VEHICLE -> INVOICE DRILL
            elif qlick_view == "tl_number" and click_value:
                final_df = final_df.filter(pl.col("tl_number") == click_value)
                if final_df.is_empty():
                    return {
                        "status": True,
                        "message": f"No data found for vehicle {click_value}",
                        "data": []
                    }

                if 'invoice_number' in final_df.columns:
                    group_cols = ["invoice_number"]
                    if "created_at" in final_df.columns:
                        group_cols.append("created_at")

                    agg_df = (
                        final_df
                        .group_by(group_cols)
                        .agg([pl.col(violation_cols).sum()])
                    )
                    agg_df = agg_df.with_columns(total_count=pl.sum_horizontal(pl.col(violation_cols)))
                else:
                    agg_df = final_df.clone()
                    agg_df = agg_df.with_columns(total_count=pl.sum_horizontal(pl.col(violation_cols)))

                return {
                    "status": True,
                    "message": f"Invoice-wise violations for vehicle {click_value}",
                    "data": agg_df.to_dicts()
                }
            if str(payload.get("download", "")).lower() == "true":
                return await download_streaming_data(final_df, filename='violations')
            return await streaming_data(final_df)

        except Exception as e:
            print("ERROR:", traceback.format_exc())
            return {"status": False, "message": str(e), "data": []}


    @staticmethod
    async def violation_percentages(filters, cross_filters, drill_state, payload):
        try:
            # Step 1: Get base query and apply filters
            query = vts_query.vts_query.get(drill_state.split(",")[0])
            emlock_open_query = vts_query.vts_query.get("emlock_open")

            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, query)
            query = VTSAnalyticsActions.apply_conditions_to_query(query, conditions)
            df = await VTSAnalyticsActions.execute_query(query)
            total_trip_count=len(df)          # Total vts_trip_count

            emlock_conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, emlock_open_query)
            emlock_query = VTSAnalyticsActions.apply_conditions_to_query(emlock_open_query, emlock_conditions)
            emlock_df = await VTSAnalyticsActions.execute_query(emlock_query)
            emlock_open = emlock_df["emlock_open"][0] if not emlock_df.empty else 0

            if df.empty:
                return {"status": True, "message": "No data found", "data": []}

            # Step 2: Define violation columns
            violation_cols = [
                "route_deviation_count",
                "stoppage_violations_count",
                "device_tamper_count",
                "speed_violation_count",
                "night_driving_count",
                "main_supply_removal_count",
                "continuous_driving_count"

            ]
            print("dfcolumns",df.columns)

            df_viol = df[["invoice_number"] + violation_cols].copy()
            df_viol.dropna(subset=["invoice_number"], inplace=True)

            # Step 3: Convert to binary (mark violations)
            for col in violation_cols:
                if col == "main_supply_removal_count":
                    df_viol[col] = df_viol[col].apply(lambda x: 1 if x and x >= 6 else 0)
                else:
                    df_viol[col] = df_viol[col].apply(lambda x: 1 if x and x != 0 else 0)

            # Step 4: Count each violation across all invoices
            violation_counts = {col: int(df_viol[col].sum()) for col in violation_cols}

            # Step 5: Add emlock_open
            violation_counts["emlock_open"] = int(emlock_open)

            # Step 6: Get shortage count
            shortage_result = await VTSAnalyticsActions.total_count_shortage(filters, cross_filters, drill_state, payload)
            shortage_count = shortage_result.get("trip_count", 0) if shortage_result.get("status") else 0
            violation_counts["shortage_count"] = shortage_count

            # --- PRINT COUNTS ---
            print("Violation counts (including emlock and shortage):")
            for key, count in violation_counts.items():
                print(f"{key}: {count}")

            # Step 7: Calculate total and percentages
            percentages = {}
            for key, count in violation_counts.items():
                percentages[key] = round(100 * count / total_trip_count, 2) if total_trip_count > 0 else 0

            
            return {
                "status": True,"message": "Violation percentages calculated",
                "data": { "counts": violation_counts,"percentages": percentages,"total_trip":total_trip_count
                }
            }

        except Exception as e:
            import traceback
            print("traceback:", traceback.format_exc())
            return {"status": False, "message": str(e), "data": []}



    @staticmethod
    async def total_count_shortage(filters, cross_filters, drill_state, payload):
        try:
            # Step 1: Build the base query to get shortage data
            shortage_query = """
            SELECT 
                sap_id, 
                vehicle_id, 
                invoice_no,
                SUM(qty_shortage::numeric) as qty_shortage 
            FROM sales_trips_till_date
            WHERE
                load_status = '6'
            GROUP BY vehicle_id, invoice_no, sap_id
            """
            
            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, shortage_query)
            shortage_query = VTSAnalyticsActions.apply_conditions_to_query(shortage_query, conditions)
            print("Shortage Query:", shortage_query)

            # REMOVED: shortage_query = VTSAnalyticsActions.apply_conditions_to_query(shortage_query, conditions_str)
            
            # Step 2: Get location master data
            location_query = """
            SELECT 
                sap_id,
                bu
            FROM location_master
            """
            
            # Execute both queries
            shortage_df = await VTSAnalyticsActions.execute_query(shortage_query)
            location_df = await VTSAnalyticsActions.execute_query(location_query)
            
            # Check if dataframes are empty
            if shortage_df.empty:
                return {"status": True, "message": "No shortage data found", "data": []}
            
            if location_df.empty:
                return {"status": False, "message": "Location master data not available", "data": []}
            
            # Step 3: Merge the dataframes on sap_id
            merged_df = shortage_df.merge(location_df, on='sap_id', how='left')
            
            filtered_df = merged_df.copy()
            
            if filters:
                for rec in filters:
                    if rec.key.lower() == 'bu':
                        filtered_df = filtered_df[filtered_df['bu'] == rec.value]
            
            total_count = len(filtered_df)
                
            return {
                "status": True, 
                "message": "Success", 
                "trip_count": total_count
            }
            
        except Exception as e:
            print("traceback:", traceback.format_exc())
            return {"status": False, "message": str(e), "data": []}

    @staticmethod
    async def vts_drill_down_violation(filters, cross_filters, drill_state, payload):
        try:
            # Step 1: Get base query and apply filters
            query = vts_query.vts_query.get(drill_state.split(",")[0])
            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, query)
            vts_drill_query = VTSAnalyticsActions.apply_conditions_to_query(query, conditions)
            print(vts_drill_query)

            vts_df = await VTSAnalyticsActions.execute_query(vts_drill_query)
            # vts_df = vts_df.drop_duplicates(subset=['invoice_number'], keep='first')
            vts_df.rename(columns={"vts_end_datetime": "created_at"}, inplace=True)
            vts_df["created_at"] = pd.to_datetime(vts_df["created_at"]).dt.date
            vts_df = vts_df.sort_values(by='created_at', ascending=True)
            # vts_df = vts_df.drop_duplicates(subset=['invoice_number', 'zone'], keep='first')


            if vts_df.empty:
                return {"status": True, "message": "No data found", "data": []}
            
            transporter_query = """SELECT distinct truck_no, transporter_name FROM vts_truck_master"""
            transporter_df = await VTSAnalyticsActions.execute_query(transporter_query)
            merged_df = vts_df.merge(transporter_df, left_on="tl_number", right_on="truck_no", how="left")

            # Step 5: Filter violation type
            violation_type = payload.get("violation_type")
            if not violation_type or violation_type not in merged_df.columns:
                return {"status": False, "message": f"Invalid violation type: {violation_type}", "data": []}

            violation_filtered_df = merged_df[merged_df[violation_type].fillna(0) != 0].copy()
            violation_filtered_df = violation_filtered_df.sort_values(by='created_at', ascending=True)
            violation_filtered_df = violation_filtered_df.drop_duplicates(subset=['invoice_number'], keep='first')

            # Step 6: Remove empty values for zone, location, transporter
            for key in ["zone", "location_name", "transporter_name"]:
                if payload.get(key):
                    violation_filtered_df = violation_filtered_df[violation_filtered_df[key] == payload[key]]

            if violation_filtered_df.empty:
                return {"status": True, "message": "No data found for the applied filters", "data": []}

            # Step 7: TL-level drill-down for invoice details
            selected_tl = payload.get("tl_number")
            if selected_tl:
                violation_filtered_df = violation_filtered_df[violation_filtered_df["tl_number"] == selected_tl]

                if violation_filtered_df.empty:
                    return {"status": True,  "message": f"No invoices found for vehicle {selected_tl}", "data": []  }

                # Return invoice details sorted by created_at
                invoice_df = violation_filtered_df.sort_values(by="created_at", ascending=True)
                invoice_df = invoice_df[["invoice_number", "created_at", violation_type]]

                # Rename columns for frontend
                invoice_df.rename(columns={
                    "invoice_number": "invoice_no",
                    "created_at": "created_at",
                    violation_type: f"{violation_type}"  
                }, inplace=True)

                result = invoice_df.to_dict(orient="records")
                return { "status": True,  "message": f"{violation_type} details for vehicle {selected_tl}",  "data": result }

            # Step 8: Determine grouping column for summaries
            if payload.get("transporter_name"):
                group_col = "tl_number"
            elif payload.get("location_name"):
                group_col = "transporter_name"
            elif payload.get("zone"):
                group_col = "location_name"
            else:
                group_col = "zone"
            if "zone" in violation_filtered_df.columns:
                violation_filtered_df["zone"] = violation_filtered_df["zone"].fillna("UNKNOWN")

            # Step 9: Summarize counts
            summary_df = (
                violation_filtered_df.groupby(group_col)
                .agg({"invoice_number": pd.Series.nunique})
                .reset_index()
            )

            summary_df[violation_type] = violation_filtered_df.groupby(group_col)[violation_type].sum().values
            if group_col != "tl_number":
                 summary_df["vehicle_count"] = violation_filtered_df.groupby(group_col)["tl_number"].nunique().values
                
            rename_mapping = {
                "invoice_number": "invoice_count",
            }
            summary_df.rename(columns=rename_mapping, inplace=True)

            # Also rename the dynamic violation column for clarity
            summary_df.rename(columns={violation_type: violation_type}, inplace=True)

            result = summary_df.to_dict(orient="records")
            return {"status": True,  "message": f"{violation_type} drill-down data", "data": result }

        except Exception as e:
            print("traceback:", traceback.format_exc())
            return {"status": False, "message": str(e), "data": []}
    @staticmethod
    async def vts_drill(filters, cross_filters, drill_state, payload):
        try:
            #  A) DOWNLOAD: MULTIPLE SHEETS BY VIOLATION
            if payload.get("download"):

                violation_query = """
                    SELECT
                        tl_number,
                        invoice_number,
                        location_name,
                        zone,
                        DATE(vts_end_datetime) AS created_at,
                        stoppage_violations_count,
                        route_deviation_count,
                        device_tamper_count,
                        main_supply_removal_count,
                        night_driving_count,
                        speed_violation_count,
                        continuous_driving_count
                    FROM vts_alert_history
                    WHERE 
                        (
                            stoppage_violations_count > 0 OR
                            route_deviation_count > 0 OR
                            device_tamper_count > 0 OR
                            main_supply_removal_count > 0 OR
                            night_driving_count > 0 OR
                            speed_violation_count > 0 OR
                            continuous_driving_count > 0
                        )
                """

                # Apply filters
                conditions = VTSAnalyticsActions.build_filter_conditions(
                    filters, cross_filters, violation_query
                )
                final_query = VTSAnalyticsActions.apply_conditions_to_query(
                    violation_query, conditions
                )

                print("Final violation query for download:", final_query)

                df = await VTSAnalyticsActions.execute_query(final_query)
                df = df.drop_duplicates(keep="first")

                if df is None or df.empty:
                    return {
                        "status": True,
                        "message": "No violation data found for download",
                        "data": []
                    }

                # Violation columns
                violation_cols = [
                    "stoppage_violations_count",
                    "route_deviation_count",
                    "device_tamper_count",
                    "main_supply_removal_count",
                    "night_driving_count",
                    "speed_violation_count",
                    "continuous_driving_count",
                ]

                # Create Excel file
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine="xlsxwriter") as writer:

                    # -----------------------------------------------------
                    #  1️⃣  E M L O C K   S H E E T
                    # -----------------------------------------------------
                    print("Generating Emlock Sheet (UI Logic)...")

                    # 1) Fetch UI EMLOCK query
                    emlock_query = vts_query.vts_query.get("get_emlock_open_data")

                    # 2) Apply filters
                    emlock_conditions = VTSAnalyticsActions.build_filter_conditions(
                        filters, cross_filters, emlock_query
                    )
                    emlock_query = VTSAnalyticsActions.apply_conditions_to_query(
                        emlock_query, emlock_conditions
                    )

                    # 3) Execute SQL
                    resp = await urdhva_base.BasePostgresModel.get_aggr_data(
                        query=emlock_query, limit=0
                    )
                    df2 = pd.DataFrame(resp.get("data", []))

                    print("Raw emlock_df count:", len(df2))

                    # 4) Convert TRUE/FALSE to lowercase
                    df2["swipeoutl1"] = df2["swipeoutl1"].fillna("").astype(str).str.lower()
                    df2["swipeoutl2"] = df2["swipeoutl2"].fillna("").astype(str).str.lower()
                    # swipe_out_l1 = df2[df2["swipeoutl1"] == "false"]
                    # swipe_out_l2 = df2[df2["swipeoutl2"] == "false"]

                    final_df_pending = df2[
                        (df2["swipeoutl1"] == "false") |
                        (df2["swipeoutl2"] == "false")
                    ].copy()

                    # final_df_pending = pd.concat([swipe_out_l1, swipe_out_l2]).reset_index(drop=True)
                    final_df_pending["violation_type"] = "open EM Lock"

                    print("Final EMLOCK rows:", len(final_df_pending))

                    # 7) Write to Excel
                    if final_df_pending.empty:
                        pd.DataFrame([{"message": "No data found for emlock_open"}]).to_excel(
                            
                            writer, index=False, sheet_name="emlock_open"
                        )
                    else:
                        final_df_pending.to_excel(writer, index=False, sheet_name="emlock_open")

                    print(f"Wrote {len(final_df_pending)} rows to emlock_open sheet.")
                    

                    # -----------------------------------------------------
                    #  2️⃣  S H O R T A G E   S H E E T
                    # -----------------------------------------------------
                    print("Generating Shortage Sheet...")

                    shortage_query = """
                        SELECT *
                        FROM sales_trips_till_date
                        WHERE load_status = '6'
                    """
                    shortage_conditions = VTSAnalyticsActions.build_filter_conditions(
                        filters, cross_filters, shortage_query
                    )
                    shortage_query = VTSAnalyticsActions.apply_conditions_to_query(
                        shortage_query, shortage_conditions
                    )

                    shortage_df = await VTSAnalyticsActions.execute_query(shortage_query)

                    if shortage_df is None or shortage_df.empty:
                        pd.DataFrame([{"message": "No data found for shortage_count"}]).to_excel(
                            writer, index=False, sheet_name="shortage_count"
                        )
                    else:
                        shortage_df.to_excel(writer, index=False, sheet_name="shortage_count")

                    # -----------------------------------------------------
                    #  3️⃣  A L L   V I O L A T I O N S   S H E E T
                    # -----------------------------------------------------
                    melt_df = df.melt(
                        id_vars=["tl_number", "invoice_number", "location_name", "zone", "created_at"],
                        value_vars=violation_cols,
                        var_name="violation_type",
                        value_name="violation_value"
                    )

                    # Keep only rows with violations > 0
                    melt_df = melt_df[melt_df["violation_value"] > 0]

                    # Deduplicate: 1 row per invoice per violation TYPE
                    melt_df = melt_df.drop_duplicates(
                        subset=["invoice_number", "violation_type"],
                        keep="first"
                    )

                    # Pivot back to wide format
                    all_violations_df = melt_df.pivot_table(
                        index=["tl_number", "invoice_number", "location_name", "zone", "created_at"],
                        columns="violation_type",
                        values="violation_value",
                        fill_value=0
                    ).reset_index()

                    # Ensure all violation columns are present
                    for v in violation_cols:
                        if v not in all_violations_df.columns:
                            all_violations_df[v] = 0

                    # Sort columns in correct order
                    all_violations_df = all_violations_df[
                        ["tl_number", "invoice_number", "location_name", "zone", "created_at"] + violation_cols
                    ]

                    # Write to Excel
                    all_violations_df.to_excel(writer, index=False, sheet_name="all_violations")

                    # # -----------------------------------------------------
                    # #  4️⃣  S E P A R A T E   V I O L A T I O N   S H E E T S
                    # # -----------------------------------------------------
                    # for col in violation_cols:
                    #     if col not in df.columns:
                    #         continue

                    #     col_df = df[df[col].fillna(0) > 0].copy()
                    #     if col_df.empty:
                    #         continue

                    #     col_df = col_df.sort_values(by="created_at")
                    #     col_df = col_df.drop_duplicates(subset=["invoice_number"], keep="first")

                    #     sheet_name = col[:31]
                    #     col_df.to_excel(writer, index=False, sheet_name=sheet_name)

                # Return file
                output.seek(0)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                file_name = f"vts_violations_{timestamp}.xlsx"

                headers = {"Content-Disposition": f'attachment; filename="{file_name}"'}
                return StreamingResponse(
                    output,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers=headers,
                )

            # ---------------------------------------
            #  B) NORMAL JSON FLOW
            # ---------------------------------------
            query = vts_query.vts_query.get(drill_state.split(",")[0])
            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, query)
            vts_drill_query = VTSAnalyticsActions.apply_conditions_to_query(query, conditions)
            print("Drill-down query (non-download):", vts_drill_query)

            vts_df = await VTSAnalyticsActions.execute_query(vts_drill_query)
            vts_df.rename(columns={"vts_end_datetime": "created_at"}, inplace=True)
            vts_df["created_at"] = pd.to_datetime(vts_df["created_at"]).dt.date
            vts_df = vts_df.sort_values(by="created_at", ascending=True)

            if vts_df.empty:
                return {"status": True, "message": "No data found", "data": []}

            transporter_query = """SELECT distinct truck_no, transporter_name FROM vts_truck_master"""
            transporter_df = await VTSAnalyticsActions.execute_query(transporter_query)
            merged_df = vts_df.merge(
                transporter_df, left_on="tl_number", right_on="truck_no", how="left"
            )

            violation_type = payload.get("violation_type")
            if violation_type and violation_type != "all":
                if violation_type not in merged_df.columns:
                    return {
                        "status": False,
                        "message": f"Invalid violation type: {violation_type}",
                        "data": [],
                    }

                violation_filtered_df = merged_df[merged_df[violation_type].fillna(0) != 0].copy()
                violation_filtered_df = violation_filtered_df.sort_values(by="created_at", ascending=True)
                violation_filtered_df = violation_filtered_df.drop_duplicates(
                    subset=["invoice_number"], keep="first"
                )
                data = violation_filtered_df.to_dict(orient="records")
            else:
                data = merged_df.to_dict(orient="records")

            return {"status": True, "message": "success", "data": data}

        except Exception as e:
            print("traceback:", traceback.format_exc())
            return {"status": False, "message": str(e), "data": []}
   
    
    @staticmethod
    async def vts_ongoing_trips(filters, cross_filters, drill_state, payload):
        try:
            # Step 1: Get base query and apply filters
            query = vts_query.vts_query.get(drill_state.split(",")[0])
            ongoing_trips_type = payload.get("ongoing_trips_type")

            if ongoing_trips_type:
                ongoing_trips_query = query.format(ongoing_trips_type=ongoing_trips_type)
                conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, ongoing_trips_query)
                query = VTSAnalyticsActions.apply_conditions_to_query(ongoing_trips_query, conditions)
            else:
                conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, query)
                query = VTSAnalyticsActions.apply_conditions_to_query(query, conditions)

            # Step 2: Execute base ongoing trips query
            df = await VTSAnalyticsActions.execute_query(query)

            for col in ["vehicle_latitude", "vehicle_longitude"]:
                if col in df.columns:
                    df[col] = df[col].astype(str)

            df = pl.DataFrame(df) if isinstance(df, pd.DataFrame) else df
            if df.is_empty():
                return {"status": True, "message": "No data found", "data": []}
            
            # Remove duplicate records
            df = df.unique(
                subset=["event_start_datetime", "event_end_datetime", "tt_number", "invoice_no"],
                keep="first"
            )

            # Step 3: Get location info
            sap_id_list = df.select("sap_id").drop_nulls().to_series().cast(str).to_list()
            
            if not sap_id_list:
                df = df.with_columns([
                    pl.lit("Unknown").alias("location_name"),
                    pl.lit("Unknown").alias("zone")
                ])
                merged_df = df.clone()
            else:
                sap_id_str = "', '".join(map(str, sap_id_list))
                location_query = f"""
                    SELECT sap_id, name AS location_name, zone 
                    FROM location_master 
                    WHERE sap_id IN ('{sap_id_str}')
                """
                loc_df = await VTSAnalyticsActions.execute_query(location_query, engine='polars')
                # Step 4: Merge location info (keep all trips)
                df = df.with_columns(pl.col("sap_id").cast(str).str.strip_chars())
                loc_df = loc_df.with_columns(pl.col("sap_id").cast(str).str.strip_chars())
                
                merged_df = df.join(loc_df, on="sap_id", how="left")

                # Fill missing values — don't drop any record
                merged_df = merged_df.with_columns([
                    pl.col("location_name").fill_null("Unknown"),
                    pl.col("zone").fill_null("Unknown")
                ])

            # Step 5: Fetch alert history only for ongoing trips' invoices
            invoice_list = merged_df.select("invoice_no").drop_nulls().to_series().cast(str).unique().to_list()
            
            if not invoice_list:
                return {"status": True, "message": "No invoices found in trips", "data": []}

            completed_invoice_set = set()

            conn = vts_ongoing_trips.get_db_connection()
            cursor = conn.cursor()

            CHUNK_SIZE = 1000  # Safe size to avoid 8623
            try:
                for i in range(0, len(invoice_list), CHUNK_SIZE):
                    chunk = invoice_list[i:i + CHUNK_SIZE]
                    invoices_str = "', '".join(chunk)

                    completed_query = f"""
                        SELECT DISTINCT CHALLAN_NO
                        FROM COMPLETED_TRIP
                        WHERE CHALLAN_NO IN ('{invoices_str}')
                    """

                    cursor.execute(completed_query)
                    rows = cursor.fetchall()

                    # completed_query = f"""
                    #             SELECT DISTINCT  invoice_no
                    #             FROM vts_completed_trip
                    #             WHERE invoice_no IN ('{invoices_str}')
                    #         """
                    # rows = await VTSAnalyticsActions.execute_query(completed_query, engine='polars')

                    completed_invoice_set.update(
                        str(r[0]).strip()
                        for r in rows
                        if r[0] is not None
                    )
            except Exception as e:
                print("Error fetching completed invoices:", str(e))

            finally:    
                cursor.close()
                conn.close()

            # Step 6: Filter by status
            status_filter = payload.get("status")

            if status_filter == "live":
                merged_df = merged_df.filter(
                    ~pl.col("invoice_no").is_in(list(completed_invoice_set))
                )
            elif status_filter == "closed":
                merged_df = merged_df.filter(
                    pl.col("invoice_no").is_in(list(completed_invoice_set))
                )

            if merged_df.height == 0:
                return {"status": True, "message": f"No {status_filter} trips found", "data": []}
            
            if payload.get("table") == "true":
                final_columns = [
                            "event_start_datetime", "event_end_datetime", "sap_id", "region", 
                            "zone", "location_type", "destination_code", "tt_number", "trip_id",
                            "invoice_no", "load_no", 
                            "vehicle_latitude", "vehicle_longitude", "vehicle_location", "transporter_name"
                        ]
                existing_columns = [col for col in final_columns if col in merged_df.columns]
                table_df = merged_df.select(existing_columns)
                result = table_df.to_dicts()
                return {"status": True, "message": "Data found", "total_records": table_df.height, "data": result}

            # Step 7: TT-level drill-down
            selected_tt = payload.get("tt_number")
            if selected_tt:
                merged_df = merged_df.filter(pl.col("tt_number") == selected_tt)
                
                if merged_df.height == 0:
                    return {"status": True, "message": f"No trips found for vehicle {selected_tt}", "data": []}

                # Create created_at column
                trip_df = merged_df.with_columns(
                    pl.coalesce([
                        pl.col("event_start_datetime"),
                        pl.col("event_end_datetime")
                    ]).cast(pl.Date).cast(str).alias("created_at")
                ).sort("created_at")
                
                trip_df = trip_df.select(["invoice_no", "created_at"])
                result = trip_df.to_dicts()
                
                return {"status": True, "message": f"Trip details for vehicle {selected_tt}", "data": result}

            # Step 8: Grouping logic (zone / location / transporter)
            merged_df = merged_df.with_columns([
                pl.col("zone").fill_null("Unknown"),
                pl.col("location_name").fill_null("Unknown"),
                pl.col("transporter_name").fill_null("Unknown")
            ])

            if payload.get("transporter_name"):
                group_col = "tt_number"
            elif payload.get("location_name"):
                group_col = "transporter_name"
            elif payload.get("zone"):
                group_col = "location_name"
            else:
                group_col = "zone"

            # Group by and aggregate
            summary_df = merged_df.group_by(group_col).agg([
                pl.col("invoice_no").n_unique().alias("invoice_count")
            ])

            if group_col != "tt_number":
                vehicle_counts = merged_df.group_by(group_col).agg([
                    pl.col("tt_number").n_unique().alias("vehicle_count")
                ])
                summary_df = summary_df.join(vehicle_counts, on=group_col, how="left")

            result = summary_df.to_dicts()

        
            # Step 9: Handle Excel download
            if payload.get("download") == "true":
                # Remove timezone info from datetime columns
                for col in merged_df.columns:
                    if merged_df[col].dtype in [pl.Datetime, pl.Datetime("ms"), pl.Datetime("us"), pl.Datetime("ns")]:
                        merged_df = merged_df.with_columns(
                            pl.col(col).dt.replace_time_zone(None)
                        )

                violation_mapping = {
                    "HS": "Hotspot",
                    "TC": "Trip not closed more than 2 hours",
                    "RD": "Route Deviation > 2km",
                    "WR": "Trip without route"
                }
                
                if "violation_type" in merged_df.columns:
                    merged_df = merged_df.with_columns(
                        pl.col("violation_type").replace(violation_mapping, default=pl.col("violation_type"))
                    )

                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                file_name = f"{ongoing_trips_type}{status_filter}{timestamp}.xlsx"
                output = io.BytesIO()
                
                # Convert to pandas for Excel writing (xlsxwriter works with pandas)
                merged_pd = merged_df.to_pandas()
                with pl.Config() as cfg:
                    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                        merged_pd.to_excel(writer, index=False, sheet_name='ongoing_trips')
               
                output.seek(0)
                headers = {"Content-Disposition": f'attachment; filename="{file_name}"'}
                return StreamingResponse(
                    output,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers=headers
                )

            return {"status": True, "message": f"{status_filter} data found", "data": result}

        except Exception as e:
            print("traceback:", traceback.format_exc())
            return {"status": False, "message": str(e), "data": []}


    @staticmethod
    async def safety_compliance(filters, cross_filters, drill_state, payload):                       
        try:
            if payload.get("download"):
                print("Download requested — generating Excel with multiple sheets (Polars only)")

                # Step 1: Parse table names
                table_names = [tbl.strip() for tbl in drill_state.split(",") if tbl.strip()]
                if not table_names:
                    return JSONResponse({"error": "Missing drill_state (table names)"}, status_code=400)

                # Create an in-memory Excel file
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
                    combined_violation_rows = []


                    # Step 2: Loop tables and add sheets dynamically
                    for table_name in table_names:

                        query = vts_query.vts_query.get("safety_compliance")
                        query = query.format(drill_state=table_name)

                        conditions = VTSAnalyticsActions.build_filter_conditions(
                            filters, cross_filters, query
                        )
                        final_query = VTSAnalyticsActions.apply_conditions_to_query(
                            query, conditions
                        )

                        print(f" Running query for table '{table_name}': {final_query}")

                        df = await VTSAnalyticsActions.execute_query(final_query)
                        df = df.drop_duplicates(keep="first")

                        if df is None or len(df) == 0:
                            print(f"No data found for {table_name}, skipping...")
                            continue

                        # if not isinstance(df, pl.DataFrame):
                        #     df = pl.DataFrame(df)

                        # # Convert Polars → Pandas for Excel writer
                        # pdf = df.to_pandas()

                        # Write to Excel sheet (sheet name from table_name)
                        sheet_name = table_name[:31]  # Excel max 31 chars
                        df.to_excel(writer, index=False, sheet_name=sheet_name)

                # Finalize BytesIO
                output.seek(0)

                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                file_name = f"safety_complaince_{timestamp}.xlsx"

                headers = {"Content-Disposition": f'attachment; filename="{file_name}"'}

                return StreamingResponse(
                    output,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers=headers
                )
            # Step 1: Get base query
            query = vts_query.vts_query.get('safety_compliance')
            drill_state_col = drill_state.split(",")[0]
            query = query.format(drill_state=drill_state_col)

            # Step 2: Apply filters
            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, query)
            query = VTSAnalyticsActions.apply_conditions_to_query(query, conditions)
            print(query)

            # Step 3: Execute query
            df = await VTSAnalyticsActions.execute_query(query)
            df = df.drop_duplicates(keep='first')

            if df.empty:
                return {"status": True, "message": "No data found", "data": []}

            # Step 4: Apply payload filters for zone, location_name, transporter_name
            for key in ["zone", "location_name", "transporter_name"]:
                if payload.get(key):
                    df = df[df[key] == payload[key]]

            if df.empty:
                return {"status": True, "message": "No data found for the applied filters", "data": []}

            # Step 5: If tt_number in payload, return per-trip details with timestamp
            selected_tt = payload.get("tt_number")
            if selected_tt:
                trip_df = df[df["tt_number"] == selected_tt].copy()
                if trip_df.empty:
                    return {"status": True, "message": f"No trips found for vehicle {selected_tt}", "data": []}

                # Keep full event_date with timestamp
                trip_df = trip_df.sort_values(by="event_date")
                result = trip_df[["invoice_no", "event_date"]].rename(columns={"event_date": "created_at"}).to_dict(orient="records")

                return {"status": True, "message": f"Trip details for vehicle {selected_tt}", "data": result}

            # Step 6: Determine grouping column
            if payload.get("transporter_name"):
                group_col = "tt_number"
            elif payload.get("location_name"):
                group_col = "transporter_name"
            elif payload.get("zone"):
                group_col = "location_name"
            else:
                group_col = "zone"

            # Step 7: Group by and summarize
            df[group_col] = df[group_col].fillna("Unknown")
            summary_df = df.groupby(group_col).agg(
                invoice_count=("invoice_no", "nunique"),
                vehicle_count=("tt_number", "nunique"),
                **{f"{drill_state_col}_count": ("event_date", "count")}
            ).reset_index()

            result = summary_df.to_dict(orient='records')
            return {"status": True, "message": "success", "data": result}

        except Exception as e:
            print("traceback:", traceback.format_exc())
            return {"status": False, "message": str(e), "data": []}
    
                
    @staticmethod
    async def safety_compliance_percentage(filters, cross_filters, drill_state, payload):
                           
            total_trips_count = vts_query.vts_query.get("total_trips")
            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, total_trips_count)
            total_trips_count = VTSAnalyticsActions.apply_conditions_to_query(total_trips_count, conditions)            
            df_total = await VTSAnalyticsActions.execute_query(total_trips_count)            
            total_length=int(df_total.iloc[0, 0])
                                      
            try:
                # Queries returning counts
                query_keys = [
                    "vts_panic",
                    "vts_harsh_braking",
                    "vts_harsh_acceleration",
                    "vts_device_removed"
                    
                ]                                                            
                counts = {}
                for key in query_keys:
                    query = vts_query.vts_query.get(key)
                    conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, query)
                    query = VTSAnalyticsActions.apply_conditions_to_query(query, conditions)
                
                    df = await VTSAnalyticsActions.execute_query(query)
                    counts[key] = int(df.iloc[0, 0]) if not df.empty else 0
                    
                                     
                percentages = {k: round((v / total_length) * 100, 2) for k, v in counts.items()}
                return {"status": True, "message": "Success", "data": { "percentages":
                    percentages,"total_trip":total_length,"counts":counts}}
            

            except Exception as e:
                print("traceback:", traceback.format_exc())
                return {"status": False, "message": str(e), "data": []}
        
            
    @staticmethod
    async def location_level_voilation_breakup(filters, cross_filters, drill_state, payload):
        try:
            # Get query from payload
            query_type = payload.get("query_type") if payload else None
            query = vts_query.vts_query.get(query_type)
            if not query:
                return {"status": False, "message": "Query not found", "data": []}

            # Build and apply conditions (pass the query for key transformation)
            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, query)
            final_query = VTSAnalyticsActions.apply_conditions_to_query(query, conditions)

            # Execute main query
            df = await VTSAnalyticsActions.execute_query(final_query)
            if df.empty:
                return {"status": True, "message": "no data", "data": {}}

            # Fetch location master data
            loc_master_query = "SELECT sap_id, name, zone FROM location_master"
            loc_master_df = await VTSAnalyticsActions.execute_query(loc_master_query)

            # Create location mapping
            loc_map = {}
            for _, row in loc_master_df.iterrows():
                loc_map[row["sap_id"]] = {
                    "zone": row["zone"],
                    "name": row["name"]
                }

            # Build nested data
            nested_data = defaultdict(lambda: defaultdict(int))

            for _, row in df.iterrows():
                location_id = row["location_id"]
                
                # Determine group key
                group_key = None
                if location_id in loc_map:
                    if drill_state and "location" in drill_state.lower():
                        group_key = loc_map[location_id]["name"]
                    elif drill_state and "zone" in drill_state.lower():
                        group_key = loc_map[location_id]["zone"]

                if not group_key:
                    continue

                # Aggregate violation counts
                for col in df.columns:
                    if col != "location_id" and row[col] > 0:
                        nested_data[group_key][col] += row[col]

            # Format final data
            final_data = {}
            for group_key, violations in nested_data.items():
                final_data[group_key] = [
                    {"violation_type": vtype, "count": count}
                    for vtype, count in violations.items()
                ]

            return {"status": True, "message": "success", "data": final_data}

        except Exception as e:
            print("Exception:", str(e))
            return {"status": False, "message": str(e), "data": []}

    @staticmethod
    async def vts_alerts_violations(filters, cross_filters, drill_state, payload):
        try:
            query_type = payload.get("query_type") if payload else None
            alert_type = payload.get("alert_type") if payload else None
            base_query = vts_query.vts_query.get(query_type)
            
            if not base_query:
                return {"status": False, "message": "Query not found", "data": [], "percentages": []}

            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, base_query)
            conditions = VTSAnalyticsActions.add_alert_type_conditions(conditions, alert_type)
            
            alerts_query = VTSAnalyticsActions.apply_conditions_to_query(base_query, conditions)

            # Execute queries
            alerts_df = await VTSAnalyticsActions.execute_query(alerts_query, engine='polars')
            
            if alerts_df.is_empty():
                return {"status": True, "message": "success", "data": [], "percentages": []}

            # Get group by column
            group_by_column = VTSAnalyticsActions.get_group_by_column(drill_state)
            if not group_by_column or group_by_column not in alerts_df.columns:
                return {"status": False, "message": f"Column '{group_by_column}' not found", "data": [], "percentages": []}
            
            if 'violation_type' not in alerts_df.columns:
                return {"status": False, "message": "violation_type column not found", "data": [], "percentages": []}
            
            alerts_df = alerts_df.filter((pl.col(group_by_column).is_not_null()) & (pl.col(group_by_column) != ""))
          
            if alerts_df.is_empty():
                return {"status": True, "message": "success", "data": [], "percentages": []}
            
            grouped = (alerts_df.group_by([group_by_column, "violation_type"]).agg(pl.count().alias("count")))
            
            if grouped.is_empty():
                 return {"status": True, "message": "success", "data": [], "percentages": []}
        
            # Prepare response data
            data_response = []
            for group_value in grouped[group_by_column].unique().to_list():
                group_data = grouped.filter(pl.col(group_by_column) == group_value)
                violations_list = [
                    {"violation_type": row['violation_type'], "count": int(row['count'])}
                    for row in group_data.to_dicts()
                ]
                
                if violations_list:
                    data_response.append({group_value: violations_list})

            # Calculate percentages
            violation_totals = (grouped.group_by('violation_type').agg(pl.sum("count")))
            grand_total = violation_totals["count"].sum()
            percentages = []
            if grand_total > 0:
                percentages = [
                    {
                    "violation_type": row["violation_type"],
                    "percentage": round((row["count"] / grand_total) * 100, 2)
                    }
                     for row in violation_totals.iter_rows(named=True)
                ]

            return {
                "status": True, "message": "success",
                "data": data_response, "percentages": percentages
            }

        except Exception as e:
            print("Exception in vts_alerts_violations:", str(e))
            print("traceback:", traceback.format_exc())
            return {"status": False, "message": str(e), "data": [], "percentages": []}

    @staticmethod
    async def violation_trends_over_time(filters, cross_filters, drill_state, payload):
        try:
            query_type = payload.get("query_type") if payload else None
            alert_type = payload.get("alert_type") if payload else None
            base_query = vts_query.vts_query.get(query_type)

            if not base_query:
                return {"status": False, "message": "Query not found", "data": []}

            # Build conditions with alert type (pass base_query for key transformation)
            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, base_query)
            conditions = VTSAnalyticsActions.add_alert_type_conditions(conditions, alert_type)

            # Get period expression and format query
            period_expr = VTSAnalyticsActions.get_period_expression(drill_state)
            alerts_query = base_query.format(period_expr=period_expr)
            alerts_query = VTSAnalyticsActions.apply_conditions_to_query(alerts_query, conditions)

            # Execute queries
            alerts_df = await VTSAnalyticsActions.execute_query(alerts_query,engine='polars')
            
            if alerts_df.height == 0:
                return {"status": True, "message": "success", "data": []}

            if not {"violation_type", "period"}.issubset(alerts_df.columns):
                 return {"status": False, "message": "Required columns not found", "data": []}
            
            alerts_df = alerts_df.filter(pl.col("period").is_not_null() & (pl.col("period").cast(pl.Utf8).str.strip_chars() != ""))
            
            if alerts_df.height == 0:
                 return {"status": True, "message": "success", "data": []}
            
            
            grouped = (alerts_df.group_by(['period', 'violation_type']).agg(pl.len().alias("count")))
            
            if grouped.height == 0:
                return {"status": True, "message": "success", "data": []}
            
            result = []
            for period in grouped.select('period').unique().to_series():
                period_data = grouped.filter(pl.col('period') == period)
                formatted_date = VTSAnalyticsActions.format_date(period, drill_state)
                
                values = [
                    {"violation_type": row['violation_type'], "count": int(row['count'])}
                    for row in period_data.iter_rows(named=True)
                ]
                
                if values:
                    result.append({"date": formatted_date, "records": values})
            
            # Sort by period (assuming period is sortable)
            result.sort(key=lambda x: x['date'])
            
            return {"status": True, "message": "success", "data": result}

        except Exception as e:
            print("Exception in violation_trends_over_time:", str(e))
            print("traceback:", traceback.format_exc())
            return {"status": False, "message": str(e), "data": []}
        
    @staticmethod
    async def violation_details(filters, cross_filters, drill_state, payload):
        try:
            query_type = payload.get("query_type") if payload else None
            violation_type = payload.get("violation_type")
            base_query = vts_query.vts_query.get(query_type)
            
            if not base_query:
                return {"status": False, "message": "Query not found", "data": []}

            # Build conditions and format query (pass base_query for key transformation)
            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, base_query)
            period_expr = VTSAnalyticsActions.get_period_expression(drill_state)
            
            query = base_query.format(period_expr=period_expr, violation_type=violation_type)
            final_query = VTSAnalyticsActions.apply_conditions_to_query(query, conditions)

            # Execute query
            df = await VTSAnalyticsActions.execute_query(final_query)
            if df.empty:
                return {"status": True, "message": "success", "data": []}

            # Process numeric columns
            numeric_cols = [col for col in df.columns if col != "period"]
            for col in numeric_cols:
                df[col] = df[col].astype(int)
            
            # Split columns into summary and instance
            summary_cols = numeric_cols[:4] 
            instance_cols = numeric_cols[4:]
            
            # Calculate summaries
            summary_counts = [{col: int(df[col].sum()) for col in summary_cols}]
            
            overall_instance_totals = {col: int(df[col].sum()) for col in instance_cols}
            grand_total = sum(overall_instance_totals.values())

            instance_breakup = {}
            for col, total_count in overall_instance_totals.items():
                instance_breakup[col] = {
                    "total_count": total_count,
                    "percentage": round((total_count / grand_total) * 100, 2) if grand_total > 0 else 0
                }

            # Format period data
            period_data = []
            for _, row in df.iterrows():
                counts = {col: int(row[col]) for col in instance_cols}
                formatted_date = VTSAnalyticsActions.format_date(row["period"], drill_state)
                
                period_data.append({
                    "date": formatted_date,
                    "value": {"counts": counts}
                })
        
            return {
                "status": True, "message": "success",
                "data": {
                    violation_type: summary_counts,
                    "period_wise": period_data,
                    "instance_breakup": instance_breakup
                }
            }
        
        except Exception as e:
            print("Exception:", str(e))
            print(traceback.format_exc())
            return {"status": False, "message": str(e), "data": {}}
                 

    @staticmethod
    async def alert_summary(filters, cross_filters, drill_state, payload):
        try:
            query_type = payload.get("query_type") if payload else None
            violation_type = payload.get("violation_type")
            query = vts_query.vts_query.get(query_type)
            
            if not query:
                return {"status": False, "message": "Query not found", "data": []}
            
            # Get group by column and build conditions (pass query for key transformation)
            group_by_column = VTSAnalyticsActions.get_group_by_column(drill_state)
            if not group_by_column:
                return {"status": False, "message": "Invalid drill state", "data": []}
                
            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, query)
            
            # Format and execute query
            formatted_query = query.format(group_by_column=group_by_column, violation_type=violation_type)
            final_query = VTSAnalyticsActions.apply_conditions_to_query(formatted_query, conditions)

            df = await VTSAnalyticsActions.execute_query(final_query)
            
            # Filter out null/empty values
            df = df[df[group_by_column].notna()]
            df = df[df[group_by_column].astype(str).str.strip() != ""]

            if df.empty:
                return {"status": True, "message": "no data", "data": {}}
           
            # Format results
            final_result = {}
            for _, row in df.iterrows():
                group_val = row[group_by_column] 
                instance = row["instance_level"]

                if group_val not in final_result:
                    final_result[group_val] = []

                final_result[group_val].append({
                    instance: [{
                        "Blocked": row["Blocked"],
                        "Auto Unblock": row["Auto Unblock"],
                        "Manual Unblock": row["Manual Unblock"],
                        "Total": row["Total"]
                    }]
                })

            return {"status": True, "message": "success", "data": final_result}

        except Exception as e:
            return {"status": False, "message": str(e), "data": {}}

    @staticmethod
    async def card_chart_shortage(filters, cross_filters, drill_state, payload):
        try:
            card_query = vts_query.vts_query.get(drill_state.split(",")[0])
            query = vts_query.vts_query.get(card_query)
            
            if not query:
                return {"status": False, "message": "Query not found", "data": []}

            # Build and apply conditions (pass query for key transformation)
            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, query)
            final_query = VTSAnalyticsActions.apply_conditions_to_query(query, conditions)
            
            # Execute VTS history query
            vts_history = await VTSAnalyticsActions.execute_query(final_query)

            # Execute Tibco query
            conn = await VTSAnalyticsActions.tibco_connection()
            if not conn:
                return {"status": False, "message": "Database connection failed", "data": {}}
                
            shortage_tibco_query = vts_query.vts_query.get("shortage_tibco")
            shortage_resp = await VTSAnalyticsActions.execute_tibco_query(conn, shortage_tibco_query)
            shortage_data = pd.DataFrame(shortage_resp.get("data", []))
            
            # Process and merge data
            vts_history = vts_history[
                pd.notnull(vts_history["invoice_number"]) & 
                (vts_history["invoice_number"] != "")
            ]
            vts_history["invoice_prefix"] = vts_history["invoice_number"].apply(lambda x: x.split("-")[0])

            merged_df = pd.merge(
                vts_history, shortage_data,
                left_on="invoice_prefix", right_on="INVOICE_NO",
                how="inner"
            )
            
            total_qty_shortage = int(merged_df["QTY_SHORTAGE"].sum())
            
            # Calculate total violations
            violation_cols = [col for col in vts_history.columns if col not in ["invoice_number", "invoice_prefix"]]
            total_violation_count = vts_history[violation_cols].sum().sum()

            shortage_percentage = round((total_qty_shortage / total_violation_count) * 100, 2) if total_violation_count > 0 else 0

            conn.close()

            return {
                "status": True, "message": "success",
                "data": {"shortage_percentage": shortage_percentage}
            }

        except Exception as e:
            return {"status": False, "message": str(e), "data": {}}
    
    

    @staticmethod
    async def tibco_connection():
        try:
            creds = credential_loader.get_credentials('TIBCO')
            print("creds --->", creds)
            
            params = {
                "host": creds['host'],
                "database": creds['database'],
                "user": creds['user'],
                "password": creds['password'],
                "port": creds['port']
            }
            
            conn = mysql.connector.connect(**params)
            return conn
        except Exception as e:
            print(f"DB connection failed: {e}")
            return None
    
    @staticmethod
    async def execute_tibco_query(conn, query):
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute(query)
            rows = cursor.fetchall()
            cursor.close()
            return {"data": rows}
        except mysql.connector.Error as e:
            print("Query execution failed:", e)
            return {"data": []}
        
    
    @staticmethod
    async def integrate_shortage_trips(filters, cross_filters, drill_state, payload):
        import pytz
        from datetime import datetime, timedelta
        
        # ----- 1. Filter Separation and Date Condition Preparation ----- 
        trips_filters = filters  # All filters apply to trips
        
        # Extract Transporter Filter for later Pandas application (must be done post-merge)
        transporter_filter = next((f for f in trips_filters if getattr(f, 'key') == 'transporter_name'), None)
        trips_query = f"""
            SELECT *     
            FROM 
                sales_trips_till_date T
            WHERE load_status = '6'
        """
        sql_filters = [f for f in filters if getattr(f, "key", None) != "transporter_name"]    
        conditions = VTSAnalyticsActions.build_filter_conditions(sql_filters, cross_filters, trips_query)
        trips_query = VTSAnalyticsActions.apply_conditions_to_query(trips_query, conditions)
        print("trips_query", trips_query)
        trips_df = await VTSAnalyticsActions.execute_query(trips_query)
        if trips_df.empty:
            return {"status": "success", "total_invoice_count": 0, "total_vehicle_count": 0,
                    "filtered_invoice_count": 0, "filtered_vehicle_count": 0, "zones": []}
            
        trips_df.columns = [c.lower() for c in trips_df.columns]
        # 4b. Merge email master
        email_query = "SELECT transporter_code, transporter_name FROM email_master"
        email_df = await VTSAnalyticsActions.execute_query(email_query)

        if not email_df.empty:
            email_df.columns = [c.lower() for c in email_df.columns]

            # Clean and normalize keys
            email_df['transporter_code'] = (
                email_df['transporter_code']
                .astype(str)
                .str.strip()
                .str.replace(r'^00', '', regex=True)
            )
            trips_df['carrier_no'] = (
                trips_df['carrier_no']
                .astype(str)
                .str.strip()
                .str.replace(r'^00', '', regex=True)
            )

            # Ensure transporter_code is unique
            email_df = email_df.drop_duplicates(subset=['transporter_code'])

            # SAFE mapping: no extra rows, just add transporter_name column
            email_map = email_df.set_index('transporter_code')['transporter_name']
            trips_df['transporter_name'] = trips_df['carrier_no'].map(email_map)

            # --- Debug export for missing transporter_name ---
            # trips_df.to_csv('/Users/algofusion/Downloads/missing_transporters.csv', index=False)

        # ----- 5. Filter valid trips (Original Logic) -----
        
        # trips_df['qty_shortage'] = pd.to_numeric(trips_df['qty_shortage'], errors='coerce')
        trips_df['qty_shortage'] = (
            trips_df['qty_shortage']
            .astype(str)
            .str.replace(r"[^0-9.]", "", regex=True)  # keep only digits & decimal
            .str.strip()
        )

        trips_df['qty_shortage'] = pd.to_numeric(trips_df['qty_shortage'], errors='coerce').fillna(0)

        filtered_trips_df = trips_df[
            # trips_df['transporter_name'].notnull() &
            # trips_df['transporter_code'].notnull() &
            (trips_df['qty_shortage'] > 0)
        ].copy()

        # ----- 6. Apply Transporter Filter (Pandas, Original Logic) -----
        
        if transporter_filter:
            key = getattr(transporter_filter, "key", None)
            val = getattr(transporter_filter, "value", None)
            cond = getattr(transporter_filter, "cond", None)
            
            if key and val and key.lower() == 'transporter_name':
                df_col = key.lower()
                if cond == "equals":
                    filtered_trips_df = filtered_trips_df[filtered_trips_df[df_col] == val]
                elif cond == "in":
                    if not isinstance(val, list):
                        val = [val]
                    filtered_trips_df = filtered_trips_df[filtered_trips_df[df_col].isin(val)]

        # ----- 7. Counts after filtering (Original Logic) -----
        filtered_vehicle_count = filtered_trips_df['vehicle_id'].nunique()
        filtered_invoice_count = filtered_trips_df['invoice_no'].nunique()

        # filtered_vehicle_count = len(filtered_trips_df['vehicle_id'])
        
        if filtered_trips_df.empty:
            return {"status": "success", "total_invoice_count": 0, "total_vehicle_count": 0,
                    "filtered_invoice_count": 0, "filtered_vehicle_count": 0, "zones": []}
                    
        # filtered_trips_df = filtered_trips_df.drop_duplicates()

        # ----- 8. Convert load_date to IST (CRITICAL FIX: Original Logic) -----
        
        ist = pytz.timezone("Asia/Kolkata")
        if 'load_date' in filtered_trips_df.columns:
            filtered_trips_df['load_date'] = pd.to_datetime(filtered_trips_df['load_date'])
            
            if filtered_trips_df['load_date'].dt.tz is None:
                filtered_trips_df['load_date'] = filtered_trips_df['load_date'].dt.tz_localize('UTC').dt.tz_convert(ist)
            else:
                filtered_trips_df['load_date'] = filtered_trips_df['load_date'].dt.tz_convert(ist)
                
            filtered_trips_df['load_date'] = filtered_trips_df['load_date'].dt.strftime("%Y-%m-%d %H:%M:%S%z")

        # ----- 9. Dynamic hierarchical grouping (Original Logic) -----
        
        def compute_group_summary(df, group_cols):
            if not group_cols:
                return None

            result = []
            current_col = group_cols[0]
            next_cols = group_cols[1:]

            for keys, group in df.groupby(current_col, dropna=False):
                item = {current_col: keys}
                # item["shortage"] = group["qty_shortage"].sum()
                item["shortage"] = group["qty_shortage"].astype(float).sum()

                item["invoice_count"] = group["invoice_no"].nunique()
                # item["vehicle_count"] = group["vehicle_id"].nunique()
                # item["invoice_count"] = len(group["invoice_no"])
                # print('item["invoice_count"]', item["invoice_count"])
                item["vehicle_count"] = group["vehicle_id"].nunique()  
                

                # --- Material Group Bifurcation Logic ---
                if "material_group_nm" in group.columns and "qty_shortage" in group.columns:
                    bif_df = (
                        group
                        .groupby("material_group_nm", dropna=False)["qty_shortage"]
                        .sum()
                        .reset_index()
                    )

                    item["item_bifurcation"] = [
                        {
                            "material_group_nm": row["material_group_nm"],
                            "shortage": round(float(row["qty_shortage"]), 2),
                        }
                        for _, row in bif_df.iterrows()
                    ]


                child = compute_group_summary(group, next_cols)
                if child:
                    if next_cols[0] == "plant_nm":
                        item["plants"] = child
                    elif next_cols[0] == "transporter_name":
                        item["transporters"] = child
                    elif next_cols[0] == "vehicle_id":
                        item["vehicles"] = child
                    elif next_cols[0] == "invoice_no":
                        item["invoices"] = child
                else:
                    if current_col == "invoice_no" and "load_date" in group.columns:
                        item["load_date"] = group["load_date"].iloc[0]
                    

                result.append(item)

            return result

        filter_keys = [getattr(f, "key", None) for f in filters] if filters else []
        if "vehicle_id" in filter_keys:
            group_cols = ["vehicle_id", "invoice_no"]
        elif "transporter_name" in filter_keys:
            group_cols = ["transporter_name", "vehicle_id"]
        elif "plant_nm" in filter_keys:
            group_cols = ["plant_nm", "transporter_name"]
        elif "zone_nm" in filter_keys:
            group_cols = ["zone_nm", "plant_nm"]
        else:
            group_cols = ["zone_nm"]
        if "material_group_nm" not in filtered_trips_df.columns and "item_no" in filtered_trips_df.columns:
            filtered_trips_df.rename(columns={"item_no": "material_group_nm"}, inplace=True)
        
        
        if payload.get('table') == "true":
            filtered_trips_df = filtered_trips_df.rename(
                columns={'material_group_nm': 'product_bifurcation', 'qty_shortage': 'shortage'}
            )
            filtered_trips_df['product_bifurcation'] = (
                filtered_trips_df['product_bifurcation'].astype(str)
                + ':' + filtered_trips_df['shortage'].astype(str)
            )

            table_df = (
                filtered_trips_df
                .groupby(['vehicle_id', 'invoice_no'], as_index=False)
                .agg({
                    'shortage': 'sum',  # sum shortages for same vehicle+invoice
                    'product_bifurcation': lambda x: ', '.join(x),
                    'plant_nm': 'first',
                    'zone_nm': 'first',
                    'transporter_name': 'first',
                    'load_date': 'first'
                })
            )

            # ---------- SHORTAGE FILTER ----------
            
            shortage_filter = payload.get("shortage_filter") # <=
            if shortage_filter:
                sf = str(shortage_filter).replace(" ", "")

                if "<" in sf:
                    limit = float(sf.split("<")[1])
                    table_df = table_df[table_df["shortage"] < limit]
                elif ">=" in sf or "≥" in sf:
                    limit = float(sf.split(">=")[1]) if ">=" in sf else float(sf.replace("≥", ""))
                    table_df = table_df[table_df["shortage"] >= limit]

            total_records = len(table_df)
            total_shortage = table_df["shortage"].sum()

            # ---------- SEARCH FILTER ----------
            search_text = payload.get("search_text") or payload.get("search")
            print("RAW SEARCH TEXT FROM PAYLOAD:", repr(search_text))

            if search_text:
                search_text = str(search_text).strip()
                print("SEARCH TEXT AFTER STRIP:", repr(search_text))

                if search_text:
                    # Optional: search only in specific columns
                    search_cols = [
                        "vehicle_id","invoice_no","plant_nm",
                        "zone_nm","transporter_name","product_bifurcation","load_date","shortage"
                    ]
                    search_cols = [c for c in search_cols if c in table_df.columns]

                    if search_cols:
                        mask = table_df[search_cols].astype(str).apply(
                            lambda col: col.str.contains(search_text, case=False, na=False),
                            axis=0).any(axis=1)

                        table_df = table_df[mask]

      
            page = int(payload.get("page", 1))
            page_size = int(payload.get("page_size", 100))  # default 100

            if page_size <= 0:
                page_size = total_records

            start = (page - 1) * page_size
            end = page * page_size

            paged_df = table_df.iloc[start:end]

            return {
                "status": "success",
                "message": "Table Data fetched successfully",
                "data": await safe_json(paged_df),
                "page": page,"page_size": page_size,
                "total_records": total_records,"total_shortage": total_shortage,
            } 
        

        filtered_trips_df = filtered_trips_df.replace([float('inf'), float('-inf')], None)
        filtered_trips_df = filtered_trips_df.where(pd.notnull(filtered_trips_df), None)
        print("TOTAL SHORTAGE BEFORE GROUPING =", filtered_trips_df["qty_shortage"].astype(float).sum())


        zones_list = compute_group_summary(filtered_trips_df, group_cols)

        
        def clean_for_json(obj):
            import math
            if isinstance(obj, dict):
                return {k: clean_for_json(v) for k, v in obj.items()}
            elif isinstance(obj, list):
                return [clean_for_json(v) for v in obj]
            elif obj is None:
                return ""        # convert None -> empty string
            elif isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)):
                return ""        # convert NaN or inf -> empty string
            return obj

        # Clean nested structure (zones_list)
        zones_list = clean_for_json(zones_list)
        # -------------------------------------------------------

        from fastapi.encoders import jsonable_encoder
        from fastapi.responses import JSONResponse


        response_data = {
            "status": "success",
            "filtered_invoice_count": filtered_invoice_count,
            "filtered_vehicle_count": filtered_vehicle_count,
            "zones": zones_list
        }

        return JSONResponse(content=jsonable_encoder(response_data))
    
    @staticmethod
    async def get_unblock_ageing(filters, cross_filters, drill_state, payload):
        try:
            # Cross filters
            _filters, daterange = await generate_cross_filter(cross_filters)
            current_date = datetime.now().strftime("%Y-%m-%d")

            closed_query = vts_query.vts_query.get("closed_alerts")
            shortage_query = vts_query.vts_query.get("unblocked_tt_shortage")

            # Drill Down filters for closed_query
            closed_query = await get_drill_down_filter(filters, closed_query)

            access_filters = [
                dashboard_studio_model.WidgetFiltersCreate(**rec)
                for rec in await hpcl_ceg_model.LpgOperationsSummary.get_clause_conditions(formated=True)
            ]

            closed_query = await widget_actions.WidgetActions.apply_filter_drilldown(closed_query, access_filters, drill_state)

            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, shortage_query)
            shortage_query = VTSAnalyticsActions.apply_conditions_to_query(shortage_query, conditions)
            shortage_query += " GROUP BY vehicle_id"

            # Date condition only for closed_query
            clause = "WHERE" if "where" not in closed_query.lower() else "AND"
            if daterange:
                closed_query += f" {clause} created_at BETWEEN {daterange}"
            else:
                closed_query += f" {clause} CAST(created_at AS DATE) = '{current_date}'"
            print("Final Shortage Query:", shortage_query)
            # Execute queries
            resp = await urdhva_base.BasePostgresModel.get_aggr_data(query=closed_query, limit=0)
            shortage_resp = await urdhva_base.BasePostgresModel.get_aggr_data(query=shortage_query, limit=0)

            # DataFrames
            df = pd.DataFrame(resp.get("data", []))
            

            df = await filter_data(df, _filters)

            shortage = pd.DataFrame(shortage_resp.get("data", []))

            # AVERAGE UNBLOCKING LOGIC (NEW)
            # Use blocked → unblocked dates from query
            df["vehicle_blocked_end_date"] = pd.to_datetime(df["vehicle_blocked_end_date"]).dt.tz_localize(None)
            df["vehicle_blocked_start_date"] = pd.to_datetime(df["vehicle_blocked_start_date"]).dt.tz_localize(None)
            
            df["created_at"] = pd.to_datetime(
                df["created_at"]
            ).dt.tz_localize(None)

            # consider only unblocked records
            df = df[df["vehicle_unblocked_date"].notna()]


            # Average Unblocking duration (days)
            df["unblocking_days"] = (
                (df["vehicle_unblocked_date"] - df["created_at"])
                .dt.total_seconds() / 86400
            ).clip(lower=0)
            # Violation counts
            violation_counts = (
                df.pivot_table(
                    index=["sap_id", "zone", "tt_number"],
                    columns="violation_type",
                    values="location_name",
                    aggfunc="count",
                    fill_value=0
                )
            )

            # Average Unblocking (group level)
            avg_unblocking = (
                df.groupby(
                    ["sap_id", "location_name", "transporter_code", "zone", "tt_number"],
                    as_index=False
                )
                .agg(
                    total_unblocking_days=("unblocking_days", "sum"),
                    total_alerts=("unblocking_days", "count")
                )
            )

            avg_unblocking["average_unblocking"] = (
                avg_unblocking["total_unblocking_days"]
                / avg_unblocking["total_alerts"]
            ).round(2)

            df = avg_unblocking.merge(
                violation_counts, on=["sap_id", "tt_number"], how="left"
            )

            df.columns.name = None

            # Merge shortage
            df = pd.merge(df, shortage, on="tt_number", how="left")
            df = df.fillna(0)

            # Ensure violation columns
            for col in [
                "continuous_driving_count", "device_tamper_count",
                "main_supply_removal_count", "night_driving_count",
                "route_deviation_count", "speed_violation_count",
                "stoppage_violations_count"
            ]:
                if col not in df.columns:
                    df[col] = 0

            df.rename(
                columns={
                    "continuous_driving_count": "CD",
                    "device_tamper_count": "DT",
                    "main_supply_removal_count": "PD",
                    "night_driving_count": "ND",
                    "route_deviation_count": "RD",
                    "speed_violation_count": "SV",
                    "stoppage_violations_count": "US"
                },
                inplace=True
            )

            # Drill-down aggregation
            if drill_state:
                group_by_keys = [drill_state]

                if filters:
                    filter_keys = [rec.key.strip('"') for rec in filters]

                    if "zone" in filter_keys and "location_name" not in filter_keys:
                        group_by_keys = ["zone", "location_name"]
                    elif (
                        "zone" in filter_keys and
                        "location_name" in filter_keys and
                        "transporter_code" not in filter_keys
                    ):
                        group_by_keys = ["zone", "location_name", "transporter_code"]
                    elif (
                        "zone" in filter_keys and
                        "location_name" in filter_keys and
                        "transporter_code" in filter_keys and
                        "tt_number" not in filter_keys
                    ):
                        group_by_keys = [
                            "zone", "location_name", "transporter_code", "tt_number"
                        ]

                df = df.groupby(group_by_keys, as_index=False).agg({
                    "CD": "sum",
                    "DT": "sum",
                    "PD": "sum",
                    "ND": "sum",
                    "RD": "sum",
                    "SV": "sum",
                    "US": "sum",
                    "average_unblocking": "mean",
                    "shortage": "sum"
                })
                closing_cols = ["CD", "DT", "PD", "ND", "RD", "SV", "US"]

                df["average_closing"] = (
                    df[closing_cols].sum(axis=1) / len(closing_cols)
                ).round(2)

            return {
                "status": True,
                "message": "success",
                "data": df.to_dict(orient="records")
            }

        except Exception as e:
            print("-- Exception in get unblock ageing widget --")
            print("traceback :", traceback.format_exc())
            return {"status": False, "message": str(e), "data": []}
    
    async def get_emlock_open_data(filters, cross_filters, drill_state, payload):
        """
        Retrieve and process emlock open data with filters and drill-down.
        """
        try:
            _filters, daterange = await generate_cross_filter(cross_filters)
            current_date = datetime.now().strftime("%Y-%m-%d")

            query = vts_query.vts_query.get("get_emlock_open_data")
            print("Base query from config:\n", query)

            query = await get_drill_down_filter(filters, query)
            print("Query after drill-down filters:\n", query)

            access_filters = [
                dashboard_studio_model.WidgetFiltersCreate(**rec)
                for rec in await hpcl_ceg_model.LpgOperationsSummary
                .get_clause_conditions(formated=True)
            ]
            print("Access filters applied:", access_filters)

            query = await widget_actions.WidgetActions.apply_filter_drilldown(
                query, access_filters, drill_state
            )
            clause = "WHERE" if "where" not in query.lower() else "AND"
            query += (
                f" {clause} createdat BETWEEN {daterange}" if daterange
                else f" {clause} CAST(createdat AS DATE) = '{current_date}'"
            )

            print("FINAL SQL QUERY (emlock open data):\n", query)

            resp = await urdhva_base.BasePostgresModel.get_aggr_data(
                query=query, limit=0
            )
            print("Raw DB response keys:", resp.keys())
            print("Total rows fetched from DB:", len(resp.get("data", [])))

            df = pd.DataFrame(resp.get("data", []))
            print("DataFrame shape after fetch:", df.shape)

            if df.empty:
                print("No data returned from DB.")
                return {"status": True, "message": "No data found", "data": []}
            
            df = await filter_data(df, _filters)
            # Keep first occurrence of duplicates based on key columns
            dedup_cols = [col for col in ["invoice_number", "trucknumber", "zone", "location_name"] if col in df.columns]
            if dedup_cols:
                df = df.drop_duplicates(subset=dedup_cols, keep='first')
                print(f"After deduplication: {len(df)} rows")
            
            # Helper function to get completed invoices
            async def get_completed_invoices(invoice_df):
                invoice_list = invoice_df["invoice_number"].dropna().astype(str).unique().tolist()
                if not invoice_list:
                    return []
                
                invoices_str = "', '".join(invoice_list)
                completed_query = f"""
                    SELECT DISTINCT invoice_no
                    FROM vts_completed_trip
                    WHERE invoice_no IN ('{invoices_str}')
                """
                completed_df = await VTSAnalyticsActions.execute_query(
                    completed_query, 
                    engine='polars' if payload.get("download", "").lower() == "true" else None
                )
                
                if payload.get("download", "").lower() == "true":
                    return completed_df["invoice_no"].to_list() if not completed_df.is_empty() else []
                else:
                    return completed_df["invoice_no"].astype(str).tolist() if not completed_df.empty else []
            
            # Helper function to apply status filter
            def apply_status_filter(data_df, status, completed_list):
                if not status:
                    return data_df
                
                status = status.lower().strip()
                if status == "close":
                    status = "closed"
                
                invoice_col = data_df["invoice_number"].astype(str)
                
                if status == "live":
                    return data_df[~invoice_col.isin(completed_list)]
                elif status == "closed":
                    return data_df[invoice_col.isin(completed_list)]
                
                return data_df
            
            # Helper function to add swipe columns
            def add_swipe_columns(data_df):
                data_df["has_swipeoutl1"] = (
                    data_df["swipeoutl1"].fillna("").astype(str).str.lower().eq("false")
                )
                data_df["has_swipeoutl2"] = (
                    data_df["swipeoutl2"].fillna("").astype(str).str.lower().eq("false")
                )
                return data_df
            
            # Handle download mode
            if payload.get("download", "").lower() == "true":
                print("Download mode enabled. Preparing data for Excel export.")
                
                completed_invoices = await get_completed_invoices(df)
                df = apply_status_filter(df, payload.get("status"), completed_invoices)
                df = add_swipe_columns(df)
                
                download_df = df[df["has_swipeoutl1"] | df["has_swipeoutl2"]].copy()
                download_df = download_df.drop(columns=['has_swipeoutl1', 'has_swipeoutl2'], errors='ignore')
                
                pl_df = pl.from_pandas(download_df)
                return await download_streaming_data(pl_df, filename='emlock_open_data')

            if payload.get("search") == "true":
                print("Search mode enabled. Returning raw filtered data.")
                return {
                    "status": True,
                    "message": "success",
                    "data": df.to_dict(orient="records")
                }
            
            if payload.get("table") == "true":
                print("Table mode enabled. Returning table data with drill down filters.")
                
                completed_invoices = await get_completed_invoices(df)
                df = apply_status_filter(df, payload.get("status"), completed_invoices)
                df = add_swipe_columns(df)
                df = df[df["has_swipeoutl1"] | df["has_swipeoutl2"]]
                
                df_clean = df.replace([np.nan, np.inf, -np.inf], None)
                
                status_msg = payload.get("status", "").lower().strip()
                if status_msg == "close":
                    status_msg = "closed"

                return {
                    "status": True,
                    "message": f"success - {status_msg} data" if status_msg else "success",
                    "data": df_clean.to_dict(orient="records"),
                    "total_records": len(df_clean)
                }
            
            # Default aggregation mode
            df = add_swipe_columns(df)
            
            # Group by base columns
            base_group_cols = ["zone", "region", "location_name", "invoice_number", "trucknumber"]
            available_cols = [col for col in base_group_cols if col in df.columns]
            print("Grouping columns used:", available_cols)

            base = df.groupby(available_cols, as_index=False).agg(
                has_swipeoutl1=("has_swipeoutl1", "any"),
                has_swipeoutl2=("has_swipeoutl2", "any"),
            )

            # Get completed invoices and apply status filter
            completed_invoices = await get_completed_invoices(base)
            print("Invoice list count for status check:", len(base["invoice_number"].dropna().unique()))
            print("Completed trip rows fetched:", len(completed_invoices))
            
            base = apply_status_filter(base, payload.get("status"), completed_invoices)
            
            status_filter = payload.get("status")
            if base.empty:
                print("No data after applying status filter.")
                return {
                    "status": True,
                    "message": f"No {status_filter} data found" if status_filter else "No data found",
                    "data": [],
                    "swipe_out_l1_count": 0,
                    "swipe_out_l2_count": 0,
                    "distinct_invoice_count": 0,
                    "distinct_vehicle_count": 0,
                }
            
            # Calculate totals
            total_swipe_l1 = base[base["has_swipeoutl1"]]["invoice_number"].nunique()
            total_swipe_l2 = base[base["has_swipeoutl2"]]["invoice_number"].nunique()

            filtered_base = base[base["has_swipeoutl1"] | base["has_swipeoutl2"]]
            total_invoice = filtered_base["invoice_number"].nunique()
            total_vehicle = filtered_base["trucknumber"].nunique()

            # Determine grouping keys based on filters
            group_by_keys = ["zone"]
            
            if filters:
                filter_keys = [str(rec.key).lower() for rec in filters]
                
                if "zone" in filter_keys:
                    if "region" not in filter_keys:
                        group_by_keys = ["region"]
                    elif "location_name" not in filter_keys:
                        group_by_keys = ["location_name"]
                    elif "trucknumber" not in filter_keys:
                        group_by_keys = ["trucknumber"]
                    else:
                        group_by_keys = ["invoice_number"]

            group_by_keys = [col for col in group_by_keys if col in base.columns]
            print("Final group_by keys:", group_by_keys)

            # Aggregate data
            grouped_df = base.groupby(group_by_keys, as_index=False).agg(
                swipeoutl1_count=(
                    "invoice_number",
                    lambda x: x[base.loc[x.index, "has_swipeoutl1"]].nunique()
                ),
                swipeoutl2_count=(
                    "invoice_number",
                    lambda x: x[base.loc[x.index, "has_swipeoutl2"]].nunique()
                ),
                distinct_invoice_count=(
                    "invoice_number",
                    lambda x: x[
                        base.loc[x.index, "has_swipeoutl1"] | base.loc[x.index, "has_swipeoutl2"]
                    ].nunique()
                ),
                distinct_vehicle_count=(
                    "trucknumber",
                    lambda x: x[
                        base.loc[x.index, "has_swipeoutl1"] | base.loc[x.index, "has_swipeoutl2"]
                    ].nunique()
                ),
            )

            grouped_df = grouped_df.fillna(0)
            
            grouped_df = grouped_df[grouped_df["distinct_invoice_count"] > 0]

            return {
                "status": True,
                "message": "success",
                "swipe_out_l1_count": int(total_swipe_l1),
                "swipe_out_l2_count": int(total_swipe_l2),
                "distinct_invoice_count": int(total_invoice),
                "distinct_vehicle_count": int(total_vehicle),
                "data": grouped_df.to_dict(orient="records"),
            }

        except Exception:
            print("traceback:", traceback.format_exc())
            return {"status": False, "message": "Internal error", "data": []}


    @staticmethod
    async def power_disconnection(filters, cross_filters, drill_state, payload):
        try:
            #  Get base query and apply filters
            query = vts_query.vts_query.get(drill_state.split(",")[0])
            if not query:
                return {"status": False, "message": "Query not found", "data": []}
            
            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, query)
            final_query = VTSAnalyticsActions.apply_conditions_to_query(query, conditions)
            vts_df = await VTSAnalyticsActions.execute_query(final_query)
            vts_df = vts_df.drop_duplicates(subset=['invoice_number'], keep='first')
            
            if vts_df.empty:
                return {"status": True, "message": "No power disconnection alerts found", "data": []}

            trans_query = """SELECT truck_no, transporter_name from vts_truck_master"""
            df_transporter = await VTSAnalyticsActions.execute_query(trans_query)
            
            if df_transporter.empty:
                return {"status": False, "message": "No matching vehicle details found in alerts", "data": []}
            
            merged_df = vts_df.merge(df_transporter, left_on="tl_number", right_on="truck_no", how="left")
            
            if merged_df.empty:
                return {"status": False, "message": "No valid zone data found after merging with alerts", "data": []}
           
            # Filter for power disconnection violations (>= 6)
            violation_type = "main_supply_removal_count"
            violation_filtered_df = merged_df[merged_df[violation_type].fillna(0) >= 6].copy()
            
            # Remove empty values for zone, location, transporter
            for key in ["zone", "location_name", "transporter_name"]:
                if payload.get(key):
                    violation_filtered_df = violation_filtered_df[violation_filtered_df[key] == payload[key]]
            
            if violation_filtered_df.empty:
                return {"status": True, "message": "No data found for the applied filters", "data": []}
            
            #  TL-level drill-down for invoice details
            selected_tl = payload.get("tl_number")
            if selected_tl:
                violation_filtered_df = violation_filtered_df[violation_filtered_df["tl_number"] == selected_tl]
                
                if violation_filtered_df.empty:
                    return {"status": True, "message": f"No invoices found for vehicle {selected_tl}", "data": []}
                
                # Return invoice details sorted by created_at
                invoice_df = violation_filtered_df.sort_values(by="created_at", ascending=True)
                invoice_df = invoice_df[["invoice_number", "created_at", violation_type]]
                
                # Rename columns for frontend
                invoice_df.rename(columns={
                    "invoice_number": "invoice_no",
                    "created_at": "created_at"
                }, inplace=True)
                
                result = invoice_df.to_dict(orient="records")
                return {"status": True, "message": f"{violation_type} details for vehicle {selected_tl}", "data": result}
            
            # Determine grouping column for summaries
            if payload.get("transporter_name"):
                group_col = "tl_number"
            elif payload.get("location_name"):
                group_col = "transporter_name"
            elif payload.get("zone"):
                group_col = "location_name"
            else:
                group_col = "zone"
            
            # Summarize counts
            # violation_count_more_than_6: Count of invoices where main_supply_removal_count >= 6
            # total_violations: Sum of actual main_supply_removal_count values
            summary_df = (
                violation_filtered_df.groupby(group_col)
                .agg({
                    "invoice_number": pd.Series.nunique,  # invoice_count
                    violation_type: ['count', 'sum']  # count of records >= 6, and sum of actual values
                })
                .reset_index()
            )
            
            # Flatten multi-level columns
            summary_df.columns = [group_col, 'invoice_count', 'violation_count_more_than_6', 'total_violations']
            
            if group_col != "tl_number":
                summary_df["vehicle_count"] = violation_filtered_df.groupby(group_col)["tl_number"].nunique().values
            
            result = summary_df.to_dict(orient="records")
            return {"status": True, "message": f"{violation_type} drill-down data", "data": result}
        
        except Exception as e:
            print("Exception in power_disconnection:", str(e))
            print("traceback:", traceback.format_exc())
            return {"status": False, "message": str(e), "data": []}
    

    @staticmethod
    async def risk_score(filters, cross_filters, drill_state, payload):
        """
        Fetch paginated data from the specified risk score table and also support downloading.
        """
        try:
            table_name = payload.get("table_name")
            columns = payload.get("columns")
            limit = 0 if payload.get("download") == "true" else payload.get("page_size", 100)
            conditions = []
            # Pagination parameters from payload
            page = int(payload.get("page", 0))

            if not table_name:
                return {"status": False, "message": "table_name not provided in payload", "data": []}

            print(f"Fetching data from table: {table_name}")

            if columns and isinstance(columns, list) and columns:
                select_columns = ", ".join([f'"{col}"' for col in columns])
                base_query = f'SELECT {select_columns} FROM public."{table_name}"'
            else:
                base_query = f'SELECT * FROM public."{table_name}"'

            access_filters = [
                dashboard_studio_model.WidgetFiltersCreate(**rec)
                for rec in await hpcl_ceg_model.LpgOperationsSummary
                .get_clause_conditions(formated=True)
            ]
            base_query = await widget_actions.WidgetActions.apply_filter_drilldown(
                base_query, access_filters, drill_state
            )

            # Build and apply conditions
            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, base_query)

            # Add search filter condition if a search term is provided
            search_term = payload.get("search")
            if search_term and columns:
                search_conditions = []
                for col in columns:
                    search_conditions.append(f'CAST("{col}" AS TEXT) ILIKE \'%{search_term}%\'')
                if search_conditions:
                    conditions.append(f"({' OR '.join(search_conditions)})")

            # Add column-specific search filters
            column_filters = payload.get("column_filters")
            if column_filters and isinstance(column_filters, dict):
                for col, search_val in column_filters.items():
                    if search_val:  # Ensure there is a value to search for
                        # Add a case-insensitive search condition for the specific column
                        conditions.append(f'CAST("{col}" AS TEXT) ILIKE \'%{search_val}%\'')
            
            # Add column-specific range filters (>=, <=)
            range_filters = payload.get("range_filters")
            if range_filters and isinstance(range_filters, list):
                for r_filter in range_filters:
                    col = r_filter.get("column")
                    op = r_filter.get("operator")
                    val = r_filter.get("value")

                    if col and op and val is not None:
                        supported_operators = ['>=', '<=', '>', '<', '=', '!=']
                        if op not in supported_operators:
                            continue

                        conditions.append(f'CAST("{col}" AS NUMERIC) {op} {val}')


            # Build and execute count query first
            count_query = f'SELECT COUNT(*) FROM public."{table_name}"'
            filtered_count_query = VTSAnalyticsActions.apply_conditions_to_query(count_query, conditions)
            count_resp = await urdhva_base.BasePostgresModel.get_aggr_data(query=filtered_count_query)
            total_records = count_resp['data'][0]['count'] if count_resp.get('data') else 0

                        # Add sorting
            sort_by = payload.get("sort_by")
            sort_direction = payload.get("sort_direction", "asc").upper()
            if sort_by and sort_direction in ["ASC", "DESC"]:
                # Add ORDER BY to the base query before applying other conditions
                # Note: apply_conditions_to_query handles placing this correctly
                base_query += f' ORDER BY "{sort_by}" {sort_direction}'

            # Build and execute data query with pagination
            filtered_data_query = VTSAnalyticsActions.apply_conditions_to_query(base_query, conditions)
            resp = await urdhva_base.BasePostgresModel.get_aggr_data(query=filtered_data_query, limit=limit, skip=page, skip_total=True)

        
            if not resp['data']:
                return {"status": True, "message": "No data found", "data": [], "total_records": 0}

            if payload.get("download") == "true":
                df = pd.DataFrame(resp['data'])
                for col in df.select_dtypes(include=["datetime64[ns, UTC]", "datetimetz"]).columns:
                    df[col] = df[col].dt.tz_localize(None)

                df = df.dropna(axis=1, how="all")
                df = df.loc[:, (df.astype(str).apply(lambda x: x.str.strip() != "").any())]

                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                file_name = f"{table_name}_{timestamp}.xlsx"

                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                    df.to_excel(writer, index=False, sheet_name=table_name[:31])

                output.seek(0)
                headers = {"Content-Disposition": f'attachment; filename="{file_name}"'}
                return StreamingResponse(
                    output,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers=headers
                )

            clicked_invoice_no = payload.get("clicked_invoice_no")
            if table_name == "completed_trips_risk_score" and clicked_invoice_no:
                safe_invoice = str(clicked_invoice_no).replace("'", "''")
                combo_query = f"SELECT * FROM public.combo_alerts WHERE invoice_no = '{safe_invoice}'"
                combo_resp = await urdhva_base.BasePostgresModel.get_aggr_data(query=combo_query, skip_total=True)
                combo_alerts_data = combo_resp.get('data', [])
                return {
                    "status": True,
                    "message": f"Combo alerts for invoice {clicked_invoice_no}",
                    "data": combo_alerts_data,
                    "total_records": len(combo_alerts_data)
                }

            return {
                "status": True,
                "message": f"Successfully fetched {len(resp['data'])} records from {table_name}",
                "data": resp['data'],
                "page": page,
                "page_size": limit,
                "total_records": total_records
            }
        except Exception as e:
            print("Exception in risk_score:", str(e))
            print("traceback:", traceback.format_exc())
            return {"status": False, "message": str(e), "data": [], "total_records": 0}

    @staticmethod
    async def adding_device(filters, cross_filters, drill_state, payload):
            try:
                sap_tt_val = payload.get("sap_tt_no")
                print(sap_tt_val)

                if sap_tt_val:
                    safe_val = str(sap_tt_val).replace("'", "''")
                    query = f"""
                        SELECT  truck_no, location_name, transporter_code, transporter_name, bu
                        FROM vts_truck_master
                        WHERE truck_no = '{safe_val}'
                        LIMIT 1
                    """
                    print("Query", query)
                    df = await VTSAnalyticsActions.execute_query(query)

                    if df.empty:
                        return {"status": False, "message": f"No data found  SAP TT No. {sap_tt_val}", "data": []}
    

                # transporter_code 
                if "transporter_code" in df.columns and "transporter_name" in df.columns:
                    df["Transporter"] = (
                            df["transporter_code"].astype(str).replace(["nan", "None", ""], "None")  + " : " +
                            df["transporter_name"].astype(str).replace(["nan", "None", ""], "None")
                        )
                else:
                    df["Transporter"] = ""

                df = df.rename(columns={"truck_no": "SAP TT No.","bu": "Select Business","location_name": "Location"})

                df.drop(columns=["transporter_code", "transporter_name"], inplace=True, errors="ignore")
                df = df.replace([np.nan, np.inf, -np.inf], None)
                print(df)

                # return matched single row
                return {
                    "status": True,"message": f"Data found for SAP TT No. {sap_tt_val}","data": df.to_dict(orient="records")}

            except Exception as e:
                print("Exception in adding_device:", str(e))
                print("traceback:", traceback.format_exc())
                return {"status": False, "message": str(e), "data": []}
    
    @staticmethod
    async def device_commissioning_table(filters, cross_filters, drill_state, payload):
        try:
            query = """select * from device_installation"""

            if not query:
                return {"status": False, "message": "Query not found", "data": []}

            # Execute VTS history query
            df = await VTSAnalyticsActions.execute_query(query)
            # print(df)

            return{"status" :True , "message":"success","data":df.to_dict(orient="records")}
        except Exception as e:
            return {"status": False, "message": str(e), "data": {}}
    
    @staticmethod
    async def vts_accept_and_block(filters, cross_filters, drill_state, payload):
        try:
            base_query = vts_query.vts_query.get("accept_and_block")
            condition = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, base_query)
            
            if isinstance(condition, list):
                condition = " AND ".join(condition)

            final_condition = ""
            if condition:
                condition = (
                    condition.replace("bu", "a.bu").replace("created_at", "a.created_at"))

                final_condition = " AND " + condition

            final_query = base_query.format(final_condition=final_condition)
            print("Final Query: ", final_query)

            merged_df = await VTSAnalyticsActions.execute_query(final_query, engine="polars")
            merged_df = (merged_df.explode("notices").unnest("notices"))
            
            # Unique alert_id sets
            system_ids = (
                merged_df.filter(pl.col("doc_type") == "System Generated").select("alert_id").unique())
                        
            user_ids = (
                merged_df.filter(pl.col("doc_type") == "User Created").select("alert_id").unique())

            # Compare alert_id sets
            system_only_ids = system_ids.join( user_ids, on = "alert_id", how = "anti")                                 
            system_only_df = (merged_df.join(system_only_ids, on = "alert_id", how = "inner").unique(subset=["alert_id"]))
            
            both_ids = system_ids.join( user_ids, on="alert_id", how="inner")
            both_df = ( merged_df.join(both_ids, on="alert_id", how="inner").unique(subset=["alert_id"]))
            
            system_only_df, both_df = (
                    system_only_df.drop("file_path","report_type", strict=False) , both_df.drop("file_path","report_type", strict=False))
            
            if payload.get("download") == "true":
                s = system_only_df.with_columns(pl.lit("no").alias("Show_Cause_Notice"))
                b = both_df.with_columns(pl.lit("yes").alias("Show_Cause_Notice"))
                combined = pl.concat([s, b], how="vertical")
                
                return await download_streaming_data(combined, filename='Show_Cause_Notice')
                
                        
            return {"status": True, "message": "success","data":{ "system_only" :system_only_df.height,"system_and_user":both_ids.height}}
        except Exception as e:
            print("Exception in vts_accept_and_block:", str(e))
            print("traceback:", traceback.format_exc())
            return {"status": False,"message": str(e),"data": []}

    @staticmethod
    async def action_device_vts(filters, cross_filters, drill_state, payload):
        """

        If only sap_tt_no given → return row from DB (no update)
        If sap_tt_no + action/remarks given → update that row and return updated data

        """
        try:
            sap_tt_no = payload.get("sap_tt_no")
            
            params = urdhva_base.queryparams.QueryParams()
            params.q = f"sap_tt_no='{sap_tt_no}'"
            print(params.q)
            params.limit = 1
            params.fields = []   # all fields
            

            existing = await DeviceInstallation.get_all(params, resp_type="plain") 
            # in list the data_will be {'data': [{'tt_chassis_no': 'CHS1234567890',....},'count':1,'ss':1]
            rows = existing.get("data")
            if not rows:
                return {"status": False,"message": f"No record found for SAP TT No {sap_tt_no}","data": []}

            row = rows[0] 

            action = payload.get("action")
            remarks = payload.get("remarks")
            reason_for_cancel = payload.get("reason_for_cancel") # remarks_2
            if (not action) and (remarks is None or remarks == "") and (not reason_for_cancel ):
                return {"status": True,"message": f"Data found for SAP TT No {sap_tt_no}","data": [row]}
        
            # 3) Get integer id from row (primary key) for update
            device_id = row.get("id")

            action = (action or "")
            print("status",action)
            status = "Approved" if action == "Accepted" else "Rejected"
            print("Approved")
            remarks = (remarks or "").strip()
            reason_for_cancel = (reason_for_cancel or "").strip()
            
            
            if reason_for_cancel:
                status_decommissioning = "Request For Approval"
                
            # update dict from existing row
            data_dict = dict(row)     
            data_dict.pop("id", None)  
            data_dict["status"] = status
            data_dict["remarks"] = remarks
            
            data_dict["reason_for_cancel"] = reason_for_cancel
            data_dict["status_decommissioning"]=status_decommissioning

            await DeviceInstallation(id=device_id, **data_dict).modify()
            updated = await DeviceInstallation.get_all(params, resp_type="plain")
            updated_rows = updated.get("data")

            return {
                "status": True,"message": f"Status updated to '{status}' for SAP TT No {sap_tt_no}","data": updated_rows}

        except Exception as e:
            print("Exception in action_device_vtss:", str(e))
            print("traceback:", traceback.format_exc())
            return {"status": False,"message": str(e),"data": [] }
