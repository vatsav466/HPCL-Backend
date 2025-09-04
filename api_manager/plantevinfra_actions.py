from hpcl_ceg_enum import *
from hpcl_ceg_model import *
import urdhva_base
import fastapi
import pandas as pd
import datetime
import traceback
from http.client import HTTPException
import orchestrator.dashboard.chart_factory.infra_functions as infra_functions
import orchestrator.dashboard.chart_factory.plant_retail_functions as plant_retail_functions
from fastapi.responses import FileResponse, JSONResponse
import polars as pl
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd
import os
import io
import traceback

router = fastapi.APIRouter(prefix='/plantevinfra')


# Action upload_plant_ev_file
@router.post('/upload_plant_ev_file', tags=['PlantEvInfra'])
async def plantevinfra_upload_plant_ev_file(file: fastapi.UploadFile):
    try:
        save_path = "/opt/ceg/algo/orchestrator/masterdata/infra_inputs"
        os.makedirs(save_path, exist_ok=True)
        dt_str = datetime.now().strftime("%Y%m%d_%H-%M-%S") + f"_{datetime.now().microsecond}"
        filename = f"PlantEV_{dt_str}.xlsx"
        file_location = os.path.join(save_path, filename)

        # Read file bytes once
        file_bytes = await file.read()

        # Save original file
        with open(file_location, "wb") as f:
            f.write(file_bytes)
        print("Saved uploaded file at:", file_location)

        def extract_date(ws):
            for cell in ['F1', 'G1', 'F2', 'G2']:
                value = ws[cell].value
                if isinstance(value, datetime):
                    return value.strftime('%d/%m/%Y')
                elif isinstance(value, str):
                    try:
                        return datetime.strptime(value.strip(), '%d/%b/%y').strftime('%d/%m/%Y')
                    except Exception:
                        continue
            return None

        state_zone_mapping = {
            "ANDHRA PRADESH": "SCR", "ARUNACHAL PRADESH": "EAS", "ASSAM": "EAS",
            "BIHAR": "ECZ", "CHHATTISGARH": "CEN", "DELHI": "NOR", "GOA": "WES",
            "GUJARAT": "WES", "HARYANA": "NOR", "HIMACHAL PRADESH": "NFZ",
            "JHARKAND": "ECZ", "KARNATAKA": "SWZ", "KERALA": "SWZ", "MADHYA PRADESH": "WIS",
            "MAHARASHTRA": "WIS", "MANIPUR": "EAS", "MEGHALAYA": "EAS", "MIZORAM": "EAS",
            "NAGALAND": "EAS", "ODISHA": "ECZ", "PUNJAB": "NFZ", "RAJASTHAN": "NRW",
            "SIKKIM": "EAS", "TAMIL NADU": "SR", "TELANGANA": "SCR", "TRIPURA": "EAS",
            "UTTAR PRADESH": "NRW", "UTTARAKHAND": "NRW", "WEST BENGAL": "EAS",
            "CHANDIGARH": "NFZ", "PUDUCHERRY": "SR", "LADAKH": "NOR", "JAMMU AND KASHMIR": "NOR"
        }
        omc_mapping = {'BPC': 'BPCL', 'IOC': 'IOCL', 'HPC': 'HPCL'}

        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        date_value = extract_date(ws)

        df = pd.read_excel(io.BytesIO(file_bytes), skiprows=2)
        df.columns = df.columns.str.strip().str.title()
        df = df[~(df['Omc'].astype(str).str.upper().str.contains('OMC') &
                  df.iloc[:, 1].astype(str).str.upper().str.contains('TOTAL'))]
        df['State'] = df['State'].astype(str).str.strip().str.upper()
        print(df['State'].unique())
        df['City'] = df['City'].astype(str)

        df_transformed = pd.DataFrame({
            'date': date_value if date_value else '',
            'zone': df['State'].map(state_zone_mapping).fillna(''),
            'state': df['State'].str.title(),
            'region': 'State',
            'company': df['Omc'].map(omc_mapping).fillna(df['Omc']),
            'ro_name': df['Ro Name'],
            'city': df['City'],
            'battery_swapping_stations': pd.to_numeric(df.get('No. Of Battery Swapping Stations', 0),
                                                       errors='coerce').fillna(0).astype(int),
            'vehicle_charging_stations': pd.to_numeric(df.get('No. Of Vehicle Charging Stations', 0),
                                                       errors='coerce').fillna(0).astype(int),
            'fame_status': df.get('Fame', '').fillna(''),
            'non_fame_status': df.get('Non Fame', '').fillna(''),
            'status': 'Active',
            'filename': filename
        })
        df_transformed['sbu'] = 'EV'
        final_records = df_transformed.fillna("").to_dict(orient="records")
        print('final_records: ',final_records)

        query = '''DELETE FROM plant_ev_infra'''
        await urdhva_base.BasePostgresModel.execute_query(query)
        await PlantEvInfra.bulk_update(final_records, upsert=False)
        return {"message": "Uploaded successfully", "rows_processed": len(final_records)}

    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")



# Action get_plant_ev_count_infra
@router.post('/get_plant_ev_count_infra', tags=['PlantEvInfra'])
async def plantevinfra_get_plant_ev_count_infra(data: Plantevinfra_Get_Plant_Ev_Count_InfraParams):
    return await plant_retail_functions.get_plant_ev_count_info(filters=data.filters, cross_filters=data.cross_filters,
                                                         drill_state=data.drill_state, limit=data.limit,
                                                         time_grain=data.time_grain)


# Action get_all_ev_infra
@router.post('/get_all_ev_infra', tags=['PlantEvInfra'])
async def plantevinfra_get_all_ev_infra(data: Plantevinfra_Get_All_Ev_InfraParams):
    return await plant_retail_functions.get_all_ev_info(filters=data.filters, cross_filters=data.cross_filters,
                                                         drill_state=data.drill_state, limit=data.limit,
                                                         time_grain=data.time_grain)


# Action get_ev_company_infra
@router.post('/get_ev_company_infra', tags=['PlantEvInfra'])
async def plantevinfra_get_ev_company_infra(data: Plantevinfra_Get_Ev_Company_InfraParams):
    return await plant_retail_functions.get_ev_company_info(filters=data.filters, cross_filters=data.cross_filters,
                                                        drill_state=data.drill_state, limit=data.limit,
                                                        time_grain=data.time_grain)


# Action get_distinct_ev_retail_infra
@router.post('/get_distinct_ev_retail_infra', tags=['PlantEvInfra'])
async def plantevinfra_get_distinct_ev_retail_infra(data: Plantevinfra_Get_Distinct_Ev_Retail_InfraParams):
    return await plant_retail_functions.get_distinct_ev_retail_info(data.sbu, data.company, data.zone, data.state,
                                                                    data.status)


# Action get_zone_wise_ev_infra
@router.post('/get_zone_wise_ev_infra', tags=['PlantEvInfra'])
async def plantevinfra_get_zone_wise_ev_infra(data: Plantevinfra_Get_Zone_Wise_Ev_InfraParams):
    return await plant_retail_functions.get_zone_wise_ev_info(filters=data.filters, cross_filters=data.cross_filters,
                                                            drill_state=data.drill_state, limit=data.limit,
                                                            time_grain=data.time_grain)
