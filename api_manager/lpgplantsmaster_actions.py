import urdhva_base
import hpcl_ceg_model
from hpcl_ceg_enum import *
from hpcl_ceg_model import *
import fastapi
import json
import psycopg2
import mysql.connector
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from fastapi.responses import StreamingResponse
import traceback

router = fastapi.APIRouter(prefix='/lpgplantsmaster')


def excel_safe(value):
    """
    Convert unsupported Excel types (list/dict) into JSON string.
    """
    if isinstance(value, (dict, list)):
        return json.dumps(value, default=str)
    return value


# Action create_location
@router.post('/create_location', tags=['LpgPlantsMaster'])
async def lpgplantsmaster_create_location(data: Lpgplantsmaster_Create_LocationParams):
    try:
        print("Data received:", data.__dict__)
        missing = [k for k, v in {"sap_id": data.sap_id, "ip_address": data.ip_address, "port_no": data.port_no, "username": data.username, "password": data.password, "db_name": data.db_name, "db_type": data.db_type}.items() if v in [None, ""]]
        if missing:
            return {"status": False, "message": f"Fields are missing: {', '.join(missing)}", "data": None}

        existing = await hpcl_ceg_model.LpgPlantsMaster.get_all(urdhva_base.QueryParams(q=f"sap_id={data.sap_id}"), resp_type="plain")
        if existing.get("data"):
            return {"status": False, "message": f"Location with SAP ID {data.sap_id} already exists", "data": None}

        data_dict = data.__dict__.copy()
        location = await hpcl_ceg_model.LocationMaster.get_all(urdhva_base.QueryParams(q=f"sap_id='{data.sap_id}' and bu='LPG'"), resp_type="plain")
        if location.get("data"):
            loc = location["data"][0]
            data_dict["plant_name"] = loc.get("name")
            data_dict["region"] = loc.get("region")
            data_dict["zone"] = loc.get("zone")
        else:
            if not data.name:
                return {
                    "status": False,
                    "message": "Plant name is required when SAP ID is not available in Location Master",
                    "data": None
                }

            data_dict["plant_name"] = data.name.strip()
            data_dict["region"] = None
            data_dict["zone"] = None
        resp = await hpcl_ceg_model.LpgPlantsMasterCreate(**data_dict).create()
        rpt = urdhva_base.context.context.get("rpt", {})
        print("resp->", resp)
        if resp:
            print("Creating audit log for plant creation")
            await SystemAuditLogCreate(
                **{
                    "employee_id": rpt.get("username"),
                    "role": rpt.get("novex_role", []),
                    "email": rpt.get("email", ""),
                    "bu": "LPG",
                    "action": "CREATE",
                    "section": "LPG Action",
                    "action_model": "LpgPlantsMaster",
                    "remarks": f"LPG Plant '{data_dict.get('plant_name')}' with SAP ID {data.sap_id} created successfully",                    "raw_data": {
                        "new_data": data_dict
                    }
                }
            ).create()

        return {"status": True, "message": "Location created successfully", "data": resp}

    except Exception as e:
        print(traceback.format_exc())
        return {"status": False, "message": str(e), "data": None}


# Action update_location
@router.post('/update_location', tags=['LpgPlantsMaster'])
async def lpgplantsmaster_update_location(data: Lpgplantsmaster_Update_LocationParams):
    try:
        rpt = urdhva_base.context.context.get("rpt", {})
        existing = await hpcl_ceg_model.LpgPlantsMaster.get_all(urdhva_base.QueryParams(q=f"sap_id = {data.sap_id}"), resp_type="plain")

        if not existing.get("data"):
            return {
                "status": False,
                "message": f"Location with SAP ID {data.sap_id} does not exist",
                "data": None
            }

        record = existing["data"][0]
        old_data = record.copy()
        plant_name = record.get("plant_name") or record.get("name") or "Unknown"
        changes = []
        for key, value in data.__dict__.items():
            if value is None:
                continue
            old_value = record.get(key)
            # Password comparison
            if key == "password":
                old_pwd = (urdhva_base.types.Secret(str(old_value)).get_secret() if old_value and str(old_value).startswith("enc#_") else old_value)
                new_pwd = (
                    value.get_secret()
                    if hasattr(value, "get_secret")
                    else (urdhva_base.types.Secret(str(value)).get_secret() if value and str(value).startswith("enc#_") else value))

                if old_pwd != new_pwd:
                    changes.append("Password changed")
                    record[key] = value
                continue

            # Normal field comparison
            if old_value != value:
                changes.append(f"{key.replace('_', ' ').title()} changed from '{old_value}' to '{value}'")
                record[key] = value

        if not changes:
            return {
                "status": True,
                "message": "No changes detected",
                "data": existing["data"][0]
            }

        resp = await hpcl_ceg_model.LpgPlantsMaster(**record).modify()
        audit_old_data = {k: v for k, v in old_data.items() if k != "password"}
        audit_new_data = {k: v for k, v in record.items() if k != "password"}
        if resp:
            await hpcl_ceg_model.SystemAuditLogCreate(
                **{
                    "employee_id": rpt.get("username"),
                    "role": rpt.get("novex_role", []),
                    "email": rpt.get("email", ""),
                    "bu": "LPG",
                    "action": "UPDATE",
                    "section": "LPG Action",
                    "action_model": "LpgPlantsMaster",
                    "remarks": (
                        f"Plant {plant_name} "
                        f"[SAP ID: {data.sap_id}] updated. "
                        f"Changes: {'; '.join(changes)}"
                    ),
                    "raw_data": {
                        "old_data": audit_old_data,
                        "new_data": audit_new_data
                    }
                }
            ).create()

        return {
            "status": True,
            "message": "Location updated successfully",
            "data": resp
        }

    except Exception as e:
        return {
            "status": False,
            "message": str(e),
            "data": None
        }

# Action delete_location
@router.post('/delete_location', tags=['LpgPlantsMaster'])
async def lpgplantsmaster_delete_location(data: Lpgplantsmaster_Delete_LocationParams):
    try:
        rpt = urdhva_base.context.context.get("rpt", {})
        resp = await hpcl_ceg_model.LpgPlantsMaster.get_all(urdhva_base.QueryParams(q=f"sap_id={data.sap_id}"), resp_type="plain")
        if not resp["data"]:
            return {"status": False, "message": "Location not found", "data": None}

        # Store record before delete
        old_data = resp["data"][0].copy()
        await hpcl_ceg_model.LpgPlantsMaster.delete(old_data["id"])
        await hpcl_ceg_model.SystemAuditLogCreate(
            **{
                "employee_id": rpt.get("username"),
                "role": rpt.get("novex_role", []),
                "email": rpt.get("email", ""),
                "bu": "LPG",
                "action": "DELETE",
                "section": "LPG Action",
                "action_model": "LpgPlantsMaster",
                "remarks": f"Location {data.sap_id} deleted successfully",
                "raw_data": {
                    "old_data": old_data
                }
            }
        ).create()

        return {"status": True, "message": "Location deleted successfully", "data": None}

    except Exception as e:
        return {"status": False, "message": str(e), "data": None}


# Action plant_details
@router.post('/plant_details', tags=['LpgPlantsMaster'])
async def lpgplantsmaster_plant_details(data: Lpgplantsmaster_Plant_DetailsParams):
    try:
        query = """
            SELECT p.sap_id, p.plant_name, p.ip_address, p.port_no AS port, p.username, p.mail_recipients,
                p.db_type, p.db_name,
                COUNT(c.id) AS carousals,
                MAX(CASE WHEN usc.log_kind='event' THEN usc.last_process_date END) AS last_event_sync,
                MAX(CASE WHEN usc.log_kind='production' THEN usc.last_process_date END) AS last_production_sync
            FROM lpg_plants_master p
            LEFT JOIN lpg_carousals c ON p.sap_id = c.sap_id
            LEFT JOIN lpg_unified_sync_cursor usc ON p.plant_name ILIKE '%' || usc.plant_name || '%'
            GROUP BY p.sap_id, p.plant_name, p.ip_address, p.port_no, p.username, p.mail_recipients, p.db_type, p.db_name
            ORDER BY p.plant_name
        """

        result = await hpcl_ceg_model.LpgPlantsMaster.get_aggr_data(query=query, limit=0)
        return { "status": True, "message": "success", "data": result.get("data", [])}

    except Exception as e:
        return {"status": False, "message": str(e), "data": []}


# Action download_plant_and_carousal_details
@router.post('/download_plant_and_carousal_details', tags=['LpgPlantsMaster'])
async def lpgplantsmaster_download_plant_and_carousal_details(data: Lpgplantsmaster_Download_Plant_And_Carousal_DetailsParams):
    try:
        plants = await hpcl_ceg_model.LpgPlantsMaster.get_all(urdhva_base.QueryParams(), resp_type="plain")
        carousels = await hpcl_ceg_model.LpgCarousals.get_all(urdhva_base.QueryParams(), resp_type="plain")
        plant_data = plants.get("data", [])
        carousel_data = carousels.get("data", [])

        wb = Workbook()
        # ==========================================
        # Sheet 1 - LPG Plant Details
        # ==========================================
        ws1 = wb.active
        ws1.title = "LPG Plant Details"
        if plant_data:
            # plant_headers = list(plant_data[0].keys())
            plant_headers = ["sap_id", "ip_address", "port_no", "username", "password", "db_name", "db_type", "name", "plant_name", "region", "zone"]
            ws1.append(plant_headers)
            for row in plant_data:
                ws1.append([excel_safe(row.get(col)) for col in plant_headers])

        ws1.freeze_panes = "A2"
        # ==========================================
        # Sheet 2 - Carousel Details
        # ==========================================
        ws2 = wb.create_sheet("Carousel Details")
        if carousel_data:
            # carousel_headers = list(carousel_data[0].keys())
            carousel_headers = ["sap_id", "carousal_id", "heads", "rated_productivity", "production_hrs", "breaks"]
            ws2.append(carousel_headers)
            for row in carousel_data:
                ws2.append([excel_safe(row.get(col)) for col in carousel_headers])

        ws2.freeze_panes = "A2"
        # ==========================================
        # Auto Fit Columns
        # ==========================================
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                ws.column_dimensions[column_letter].width = max_length + 2

        # ==========================================
        # Generate Excel
        # ==========================================
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = (f"LPG_Master_Data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition":
                f'attachment; filename="{filename}"'
            }
        )

    except Exception as e:

        return {
            "status": False,
            "message": str(e),
            "data": None
        }


# Action test_connection
@router.post('/test_connection', tags=['LpgPlantsMaster'])
async def lpgplantsmaster_test_connection(data: Lpgplantsmaster_Test_ConnectionParams):
    try:
        missing = [k for k, v in {"ip_address": data.ip_address, "port_no": data.port_no, "username": data.username, 
                                    "db_name": data.db_name, "db_type": data.db_type}.items() if v in [None, ""]]
        if missing:
            return {"status": False, "message": f"Mandatory fields missing: {', '.join(missing)}", "data": None}

        password = data.password
        if not password:
            if not data.sap_id:
                return {"status": False, "message": "Password or SAP ID required", "data": None}

            query=f"SELECT password FROM lpg_plants_master WHERE sap_id={int(data.sap_id)}"
            result = await hpcl_ceg_model.LpgPlantsMaster.get_aggr_data(query=query, limit=1)
            rows = result.get("data", [])
            if not rows:
                return {"status": False, "message": f"SAP ID {data.sap_id} not found", "data": None}

            password = rows[0].get("password")
        if not password:
            return {"status": False, "message": f"Password not found for SAP ID {data.sap_id}", "data": None}

        if str(password).startswith("enc#_"):
            password = urdhva_base.types.Secret(password).get_secret()
            print("Password decrypted successfully")

        print(f"Testing {data.db_type} | Host={data.ip_address} | Port={data.port_no} | User={data.username} | DB={data.db_name}")

        db_type = str(data.db_type).lower().strip()
        if db_type == "mysql":
            conn = mysql.connector.connect(host=data.ip_address, port=data.port_no, user=data.username, password=password, database=data.db_name, connection_timeout=10)

        elif db_type in ["postgres", "postgresql"]:
            conn = psycopg2.connect(host=data.ip_address, port=data.port_no, user=data.username, password=password, dbname=data.db_name, connect_timeout=10)

        else:
            return {"status": False, "message": f"Unsupported DB Type: {data.db_type}", "data": None}

        conn.close()

        return {"status": True, "message": "Connection successful", "data": {"connected": True, "sap_id": data.sap_id}}

    except Exception as e:
        print("Connection Error:", str(e))
        print(traceback.format_exc())
        return {
            "status": False,
            "message": str(e),
            "data": {"connected": False}
        }
