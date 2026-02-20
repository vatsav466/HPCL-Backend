import urdhva_base
import time
import httpx
import base64
import asyncio
import json
import string
import asyncio
import hashlib
import datetime
import traceback
import pandas as pd
import numpy as np
import urdhva_base.redispool
from openpyxl import Workbook
from calendar import monthrange
try:
    from secrets import choice
except ImportError:
    from random import choice
from utilities import sales_mapping
from collections import defaultdict
from dateutil.relativedelta import relativedelta
from utilities import interlock_category_mapping
import Thingsboard.bu_asset_master_new as tb_master
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def month_short_to_number(short_name):
    # Parses the short month name (%b) and extracts the month as an integer.
    return datetime.datetime.strptime(short_name, '%b').month


def password_generator(password_length=16, special_characters_allowed=True, case_sensitive=True):
    """
    @description: function to generate random password
    @param password_length: length of the password
    @param special_characters_allowed: whether to allow special characters
    @param case_sensitive: whether to allow case sensitive(ascii uppercase allowed or not)
    @return: generated password of length password_length
    """
    password_chars = list(string.digits) + list(string.ascii_lowercase)
    if case_sensitive:
        password_chars += list(string.ascii_uppercase)
    if special_characters_allowed:
        password_chars += ["!", "#", "$", "%", "^", "&", "*", "(", ")", ",", ".", "-", "_", "+", "=", "<", ">", "?"]
    random_pass = "".join([choice(password_chars) for i in range(password_length)])
    return random_pass


def get_time_stamp_by_delta(dt=None, months=0, days=0, years=0, with_month_start_day=True,
                            date_time_format="%Y-%m-%d", ascending=False, with_month_end_day=False):
    """
    Get the timestamp by descending or ascending a specified number of months from the current date.
    :param dt: datetime object
    :param months: Total months to descend
    :param days: Total days to descend
    :param years: Total years to descend
    :param with_month_start_day: whether date should start from day 1 or present day
    :param date_time_format: Format to return the date
    :param ascending: To use in incremental or decremental format
    :param with_month_end_day: whether date should be actual or month end date
    :return: Formatted date string
    Example:
    on 2025-01-19
      Case 1
        input:- utilities.helpers.get_time_stamp_by_delta(days=1, with_month_start_day=False, year=1, ascending=True)
        response:- '2026-01-20'
      Case 2
        input:- utilities.helpers.get_time_stamp_by_delta(days=1, with_month_start_day=False, year=1, ascending=False)
        response:- '2024-01-18'
      Case 3
        input:- utilities.helpers.get_time_stamp_by_delta(days=1, with_month_start_day=True, year=1, ascending=False)
        response:- '2023-12-31'
      Case 4
        input:- utilities.helpers.get_time_stamp_by_delta(with_month_start_day=True, year=1, ascending=False)
        response:- '2024-01-01'
      Case 5
        input:- utilities.helpers.get_time_stamp_by_delta(days=1, year=0, date_time_format=None, ascending=False)
        response:- datetime.datetime(2024, 12, 31, 7, 56, 43, 663410, tzinfo=datetime.timezone.utc)
    """
    # Todo:- Need to add default timezone from settings file and requested timezone as input for changes
    if not dt:
        dt = datetime.datetime.now(tz=datetime.timezone.utc)

    # Set the day to 1 if with_month_start_day is True
    if with_month_start_day:
        dt = dt.replace(day=1)
    elif not months and with_month_end_day:
        _, months_days = monthrange(dt.year, dt.month)
        dt = dt.replace(day=months_days)

    # Subtract the specified number of months
    if months > 0:
        dt = dt - relativedelta(months=months) if not ascending else dt + relativedelta(months=months)
        if with_month_end_day:
            _, months_days = monthrange(dt.year, dt.month)
            dt = dt.replace(day=months_days)
    elif years > 0:
        day_filter = 0
        if days > 0:
            day_filter = days
        dt = dt - relativedelta(year=dt.year-years, days=day_filter) if not ascending \
            else dt + relativedelta(year=dt.year+years, days=day_filter)
    elif days > 0:
        dt = dt - relativedelta(days=days) if not ascending else dt + relativedelta(days=days)

    # Format the date
    if date_time_format:
        return dt.strftime(date_time_format)

    return dt


def generate_hash(list_of_strings, bit_size=64):
    """
    Generates unique hash key for the given inputs
    :param list_of_strings:
    :param bit_size:
    :return: unique string
    """
    # Convert the list of strings to a single string
    combined_string = ''.join(list_of_strings)
    # Create a hash object
    if bit_size > 40:
        # SHA256 for 64-bit key
        hash_object = hashlib.sha256()
    elif bit_size > 32:
        # SHA1 for 40-bit key
        hash_object = hashlib.sha1()
    else:
        # SHA1 for 32-bit key
        hash_object = hashlib.md5()
    # Update the hash object with the combined string
    hash_object.update(combined_string.encode('utf-8'))
    # Get the hexadecimal digest of the hash
    hash_value = hash_object.hexdigest()
    return hash_value


def encrypt_file(file_path):
    """
    Encrypt a file using the provided encryption key.
    
    Args:
        file_path (str): Path to the file to be encrypted.
        encryption_key (bytes): Encryption key.

    Returns:
        str: Path to the encrypted file.
    """
    encrypted_file_path = f"{file_path}.enc"

    with open(file_path, "rb") as file:
        file_data = file.read()  # Read file content
        with open(encrypted_file_path, "wb") as encrypted_file:
            encrypted_file.write(file_data)  # Save encrypted data
    file_path = str(urdhva_base.types.Secret().validate(encrypted_file_path, ''))
    return base64.b64encode(file_path.encode()).decode()


def normalize_string(input_value):
    """
    Normalizes provided string, If binary convert to string and return else return string
    :param input_value:
    :return:
    """
    if isinstance(input_value, bytes):
        return input_value.decode()
    return input_value


async def generate_filter_query(filters, query, where_clause=False):
    try:
        conditions = []
        _key = None
        if filters:
            for rec in filters:
                values = rec.value.split(",")
                if len(values) == 1:
                    if rec.cond == "not_equals":
                        conditions.append(f'{rec.key} != \'{values[0]}\'')
                    elif rec.key in ["DATE", "created_at"]:
                        conditions.append(f'DATE({rec.key}) = \'{values[0]}\'')
                    else:
                        conditions.append(f'{rec.key} = \'{values[0]}\'')
                elif len(values) == 2 and rec.key in ["DATE", "created_at"]:
                    from_date = values[0]
                    to_date = values[-1]
                    conditions.append(f"{rec.key} BETWEEN '{from_date} 00:00:00' AND '{to_date} 23:59:59' ")
                else:
                    conditions.append(f"{rec.key} IN {tuple(values)}")
        if conditions:
            if where_clause:
                query += " WHERE " + " AND ".join(conditions)
            else:
                query += " AND " + " AND " .join(conditions)
        if _key:
            return query, _key
        return query
    except Exception as e:
        print("--- Exception in drill down filters ---")
        print("Exception :", str(e))
        return query


async def get_location_details(bu, sap_id):
    """
    Retrieves location details based on the provided business unit and SAP ID.

    Parameters:
    bu (str): The business unit identifier.
    sap_id (str): The SAP ID of the location.

    Returns:
    dict: Location details, including name, address, coordinates, etc., or None if not found.
    """
    MAX_RETRIES = 5
    RETRY_DELAY = 2
    for attempt in range(MAX_RETRIES):
        try:
            if not bu or not sap_id:
                return False, {"msg": "Invalid parameters: 'bu' and 'sap_id' are required."}
            async with httpx.AsyncClient(verify=False) as client:
                base_url = f"http://{urdhva_base.settings.cache_gateway_host}:{urdhva_base.settings.cache_gateway_port}"
                resp = await client.get(f"{base_url}/api_cache/v1/get_location_data", params={"bu": bu,
                                                                                            'location_id': sap_id})
                if resp.status_code // 100 == 2:
                    return resp.json()
                else:
                    print(resp.status_code, resp.text)
        except Exception as e:
            print(f"Error in getting location details: {e}, BU: {bu}, Location ID: {sap_id}")
        if attempt < MAX_RETRIES - 1:
            time.sleep(RETRY_DELAY * (2 ** attempt))
        else:
            return False, {}
    return False, {}


async def get_alert_camunda_url(alert_id, base_url):
    """
    API to get camunda based on the alertid
    :param alert_id:
    :return:
    """
    redis_ins = urdhva_base.redispool.get_synchronous_redis_connection()
    try:
        if redis_ins.hexists("alert_camunda_url", f"{alert_id}"):
            url = redis_ins.hget("alert_camunda_url", f"{alert_id}")
            return url.decode() if isinstance(url, bytes) else url
        return base_url
    except:
        return base_url
    finally:
        try:
            redis_ins.close()
        except:
            ...


def validate_camunda_settings_rule(camunda_settings, location_id, bu):
    """
    Verifying rule configuring for camunda, Validating odd/even settings
    :param camunda_settings:
    :param location_id:
    :param bu:
    :return:
    """
    if not camunda_settings.get("rule"):
        return True
    try:
        if camunda_settings['rule'] == "even":
            if int(location_id) % 2 == 0:
                return True
        elif camunda_settings['rule'] == "odd":
            if int(location_id) % 2 != 0:
                return True
    except Exception as e:
        print(f"Exception while handling rule {e}, Traceback {traceback.format_exc()}")
    return False


async def get_camunda_url(bu, sap_id, alert_section, location_data={}):
    """
    Logic to decide serving camunda url for given bu and sap_id
    :param bu:
    :param sap_id:
    :param alert_section:
    :return:
    """
    camunda_config = urdhva_base.settings.camunda_configuration
    default_url = urdhva_base.settings.camunda_url

    # If configuration is missing or BU is not in the config, return default
    if not camunda_config or bu not in camunda_config:
        return default_url

    # status, location_data = await get_location_details(bu, sap_id)
    # if not status:
    #     return default_url

    if not location_data:
        location_data = {'sap_id': sap_id}
    # Fields to check in settings
    match_keys = ['sap_id']
    #, 'sales_area', 'region', 'zone']

    # Checking ones having alert section
    for settings in camunda_config[bu]:
        if settings.get('alert_section') == alert_section:
            if (any(settings.get(k) and location_data.get(k, "") in settings[k] for k in match_keys) and
                    validate_camunda_settings_rule(settings, sap_id, bu)):
                return settings['url']

    # Checking ones not having alert section
    for settings in camunda_config[bu]:
        if not settings.get('alert_section'):
            if (any(settings.get(k) and location_data.get(k, "") in settings[k] for k in match_keys) and
                    validate_camunda_settings_rule(settings, sap_id, bu)):
                return settings['url']

    # Checking for single URL with alert section
    for settings in camunda_config[bu]:
        if settings.get('alert_section') == alert_section and validate_camunda_settings_rule(settings, sap_id, bu):
            return settings['url']

    # Checking for single URL without alert section
    for settings in camunda_config[bu]:
        if not settings.get('alert_section') and validate_camunda_settings_rule(settings, sap_id, bu):
            return settings['url']

    # Checking for global match
    for settings in camunda_config[bu]:
        if (any(not settings.get(k) or "*" in settings[k] for k in match_keys) and
                validate_camunda_settings_rule(settings, sap_id, bu)):
            return settings['url']

    return default_url

async def get_doc_link(file_name: str):
    server_ip = urdhva_base.settings.server_ip
    return f"http://{server_ip}:8080/api/alerts/stored_document?file_name={file_name}"

def map_device_category(interlock_name):
    for category, interlocks in interlock_category_mapping.interlock_to_category.items():
        if interlock_name in interlocks:
            return category
    return "Unknown"

# def fetch_oi_devices(page_size=100, page=0):
#     """
#     Fetch devices from ThingsBoard instance.

#     Parameters
#     ----------
#     page_size : int
#         Page size of the devices list.
#     page : int
#         Page number of the devices list.

#     Returns
#     -------
#     list
#         List of devices.
#     """
#     params = {'pageSize': 100, 'page': 0}
#     print("params --> ", params)
#     response = tb_master.ThingsBoardInterface().api_handler("GET", "/api/tenant/devices", {}, params)
#     return response.get("data", []) if response else []

def fetch_oi_devices(page_size=100):
    """
    Fetch all OI devices from ThingsBoard using pagination.

    Parameters
    ----------
    page_size : int
        Number of devices per page.

    Returns
    -------
    list
        List of all OI devices.
    """
    all_devices = []
    page = 0

    while True:
        params = {
            'pageSize': page_size,
            'page': page
        }

        print(f"[INFO] Fetching page {page} with size {page_size}...")

        response = tb_master.ThingsBoardInterface().api_handler(
            "GET", "/api/tenant/devices", {}, params
        )

        if not response:
            print("[WARNING] No response received.")
            break

        devices = response.get("data", [])
        has_next = response.get("hasNext", False)

        print(f"[INFO] Retrieved {len(devices)} devices.")

        # Filter only OI devices if needed
        oi_devices = [d for d in devices if d.get("type") == "OI"]
        all_devices.extend(oi_devices)

        if not has_next:
            break

        page += 1

    print(f"[SUCCESS] Total OI devices fetched: {len(all_devices)}")
    return all_devices


def fetch_device_data(device_id, key="water"):
    # Fetching server attributes
    attr_url = f"/api/plugins/telemetry/DEVICE/{device_id}/values/attributes/SERVER_SCOPE"
    attr_data = tb_master.ThingsBoardInterface().api_handler('GET', attr_url, {}, {})

    required_kls = None
    target_volume = None

    # Extract required attributes from the server data
    if isinstance(attr_data, list):
        for item in attr_data:
            # Handle water attributes
            if key == "Water Volume":
                if item.get('key') == 'Required Water Volume':
                    required_kls = float(item.get('value'))
                elif item.get('key') == 'Water Volume':
                    target_volume = float(item.get('value'))
            
            # Handle foam attributes
            elif key == "Foam Volume":
                if item.get('key') == 'Required Foam Volume':
                    required_kls = float(item.get('value'))
                elif item.get('key') == 'Foam Volume':
                    target_volume = float(item.get('value'))
    elif isinstance(attr_data, dict):
        if key == "Water Volume":
            required_kls = float(attr_data.get('Required Water Volume')) if 'Required Water Volume' in attr_data else None
            target_volume = float(attr_data.get('Water Volume')) if 'Water Volume' in attr_data else None
        elif key == "Foam Volume":
            required_kls = float(attr_data.get('Required Foam Volume')) if 'Required Foam Volume' in attr_data else None
            target_volume = float(attr_data.get('Foam Volume')) if 'Foam Volume' in attr_data else None

    # Raise an exception if either value is missing
    if required_kls is None or target_volume is None:
        raise Exception(f"Missing '{key}' or 'Required {key}' in server attributes")

    # Fetching telemetry data for the device
    telemetry_url = f"/api/plugins/telemetry/DEVICE/{device_id}/values/timeseries"
    telemetry_params = {'keys': f'{key}'}
    telemetry_data = tb_master.ThingsBoardInterface().api_handler('GET', telemetry_url, {}, telemetry_params)

    volume = None
    if f'{key}' in telemetry_data:
        latest_entry = telemetry_data[f'{key}'][-1]  # Get the latest entry
        volume = float(latest_entry.get('value'))

    # Raise an exception if volume is missing
    if volume is None:
        raise Exception(f"Missing '{key}' telemetry")

    return required_kls, target_volume, volume



def fetch_alarm_data(device_id):
    alarm_url = f"/api/alarm/DEVICE/{device_id}?searchStatus=ACTIVE"
    params = {'pageSize': 100, 'page': 0}
    alarm_data = tb_master.ThingsBoardInterface().api_handler('GET', alarm_url, {}, params)
    return alarm_data

# TODO: Check the devices for saftey and process plc for all the location by comparing them sap_id 
# by comparing server attributes and latest telemetry which device is active for particular location
async def fetch_plc_devices(tb, page_size=100):
    """Fetch all PLC devices from ThingsBoard."""
    all_devices = []
    page = 0
    
    while True:
        params = {
            'pageSize': page_size,
            'page': page,
            'sortOrder': 'ASC',
            'sortProperty': 'name'
        }
        
        response = tb.api_handler("GET", "/api/tenant/devices", {}, params)
        if not response:
            break
        
        devices = response.get("data", [])
        has_next = response.get("hasNext", False)
        
        plc_devices = [d for d in devices if d.get("type") == "PLC"]
        all_devices.extend(plc_devices)
        
        if not has_next:
            break
        page += 1
    
    return all_devices

async def check_plc_status():
    """Check PLC device status and return results."""
    # Initialize ThingsBoard
    tb = tb_master.ThingsBoardInterface()
    
    # Get all PLC devices
    all_devices = await fetch_plc_devices(tb)
    
    results = []
    
    for device in all_devices:
        device_id = device.get("id", {}).get("id")
        device_name = device.get("name")
        
        if not device_id or not device_name:
            continue
        
        # Get attributes and telemetry
        attributes = tb.api_handler("GET", f"/api/plugins/telemetry/DEVICE/{device_id}/values/attributes", {}, {})
        telemetry = tb.api_handler("GET", f"/api/plugins/telemetry/DEVICE/{device_id}/values/timeseries", {}, {})
        
        # Process data
        attr_data = {item['key']: item['value'] for item in attributes} if attributes else {}
        tele_data = {k: v[0]['value'] for k, v in telemetry.items()} if telemetry else {}
        
        # Get SAPID
        sap_id = attr_data.get("SAPID")
        if not sap_id:
            continue
        
        # Check PLC A status
        plc_a_attr = attr_data.get("PLC A IS MASTER")
        plc_a_tele = tele_data.get("PLC A IS MASTER")
        
        # Check PLC B status
        plc_b_attr = attr_data.get("PLC B IS MASTER")
        plc_b_tele = tele_data.get("PLC B IS MASTER")
        
        # Convert to int
        try:
            plc_a_attr = int(plc_a_attr) if plc_a_attr is not None else None
            plc_a_tele = int(plc_a_tele) if plc_a_tele is not None else None
            plc_b_attr = int(plc_b_attr) if plc_b_attr is not None else None
            plc_b_tele = int(plc_b_tele) if plc_b_tele is not None else None
        except:
            continue
        
        # Determine status
        plc_a_status = "master" if (plc_a_attr == plc_a_tele and plc_a_attr is not None) else "slave"
        plc_b_status = "master" if (plc_b_attr == plc_b_tele and plc_b_attr is not None) else "slave"
        
        # Add result
        results.append({
            "device_name": device_name,
            "sap_id": sap_id,
            "plc_a_status": plc_a_status,
            "plc_b_status": plc_b_status
        })
    
    return results

def get_user_details(where_clause):
    where_clause = []
    if not urdhva_base.ctx.exists():
        return where_clause
    rpt = urdhva_base.context.context.get('rpt', {})
    print("rpt: ", rpt)
    user_bu = rpt.get("bu", [])
    user_zone = rpt.get("zone", [])
    user_region = rpt.get("region", [])
    user_sales_area = rpt.get("sales_area", [])

    user_zone = [sales_mapping.sales_zone_map.get(zone, zone) for zone in user_zone]

    if not user_region:
        user_region = [x['value'] for x in where_clause if x.get('key') == 'Region_Name']   
    where_clause = []

    if user_bu:
        if len(user_bu) == 1:
            where_clause.append({
                "key": "SBU_Name",
                "cond": "=",
                "value": user_bu[0]
            })
            print("where_clause: ", where_clause)
        else:
            where_clause.append({
                "key": "SBU_Name",
                "cond": "IN",
                "value": user_bu
            })
            print("where_clause: ", where_clause)

    if user_zone:
        if len(user_zone) == 1:
            where_clause.append({
                "key": "ORGZONECD",
                "cond": "=",
                "value": user_zone[0]
            })
            print("where_clause: ", where_clause)
        else:
            where_clause.append({
                "key": "ORGZONECD",
                "cond": "IN",
                "value": user_zone
            })
            print("where_clause: ", where_clause)

    if user_region:
        if len(user_region) == 1:
            where_clause.append({
                "key": "Region_Name",
                "cond": "=",
                "value": user_region[0]
            })
            print("where_clause: ", where_clause)
        else:
            where_clause.append({
                "key": "Region_Name",
                "cond": "IN",
                "value": user_region
            })
            print("where_clause: ", where_clause)

    if user_sales_area:
        if len(user_sales_area) == 1:
            where_clause.append({
                "key": "SalesArea_Name",
                "cond": "=",
                "value": user_sales_area[0]
            })
            print("where_clause: ", where_clause)
        else:
            where_clause.append({
                "key": "SalesArea_Name",
                "cond": "IN",
                "value": user_sales_area
            })
            print("where_clause: ", where_clause)

    return where_clause

async def generate_trends_report(resp, output_file, report_type="daily", filters=None):
    print("resp --> ", resp)
    print("filters --> ", filters)
    filters = filters or {}
    selected_equipment = filters.get("equipment_name", None)

    data_key = "daily_data" if report_type.lower() == "daily" else "monthly_data"
    data = resp.get(data_key, {})
    if not data:
        print("No data found.")
        return

    # Apply case-insensitive filter
    if selected_equipment:
        selected_equipment_lower = selected_equipment.lower()
        matched_key = next((key for key in data.keys() if key.lower() == selected_equipment_lower), None)
        if matched_key:
            filtered_data = {matched_key: data[matched_key]}
        else:
            filtered_data = {}
    else:
        filtered_data = data

    print("filtered_data:", filtered_data)

    # ---- Setup styling ----
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ---- Organize data ----
    all_data = []
    for equipment_name, entries in filtered_data.items():
        for entry in entries:
            entry["equipment_name"] = equipment_name
            all_data.append(entry)

    if not all_data:
        print("No matching entries after filtering.")
        return

    # ---- Sort data by date (descending) ----
    if report_type.lower() == "daily":
        date_key = "date"
        date_format = "%Y-%m-%d"  # This matches '2025-04-15'
    else:
        date_key = "month"
        date_format = "%b-%Y"     # This matches 'Apr-2025'

    # Handle potential parsing issues
    def parse_date_safe(entry):
        date_str = entry.get(date_key, "")
        try:
            return datetime.datetime.strptime(date_str, date_format)
        except ValueError:
            return datetime.datetime.min  # Default for invalid format

    all_data.sort(key=parse_date_safe, reverse=True)
    # ---- Setup Excel ----
    wb = Workbook()
    ws = wb.active
    ws.title = "Trends Report"

    # ---- Title ----
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"{report_type.upper()} REPORT"
    title_cell.font = Font(bold=True, size=14)
    title_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = thin_border

    row = 3
    first_entry = all_data[0]

    metadata = {
        "Export Date and Time": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Total Alert Count": sum(entry.get("count", 0) for entry in all_data),
        "Data Type": f"{report_type.capitalize()} Equipment Alert",
        "Equipment Name Selected": selected_equipment or "All",
        "Zone": first_entry.get("zone", ""),
        "Plant Name": first_entry.get("location_name", ""),
        "SAP ID": first_entry.get("sap_id", "")
    }

    for key, val in metadata.items():
        ws[f"A{row}"] = key
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        ws[f"A{row}"].border = thin_border

        ws[f"B{row}"] = val
        ws[f"B{row}"].alignment = Alignment(horizontal="center")
        ws[f"B{row}"].border = thin_border
        row += 1

    row += 1

    # ---- Table Header ----
    date_label = "Date" if report_type.lower() == "daily" else "Month"
    headers = [date_label, "Equipment Name", "Alert Category", "Alert Type", "Count"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    row += 1

    # ---- Write Data ----
    for entry in all_data:
        ws.cell(row=row, column=1).value = entry.get("date" if report_type == "daily" else "month", "")
        ws.cell(row=row, column=2).value = entry.get("equipment_name", "")
        ws.cell(row=row, column=3).value = entry.get("alert_category", "")
        ws.cell(row=row, column=4).value = entry.get("alert_type", "")
        ws.cell(row=row, column=5).value = entry.get("count", 0)
        row += 1

    # ---- Save ----
    wb.save(output_file)
    print(f"Saved {report_type} report to {output_file}")

async def generate_equipment_report(resp, output_file, report_type="daily", filters=None):
    print("resp --> ", resp)

    # ---- Extract data from response ----
    data = resp.get(f"{report_type}_data", {})
    process_data = data.get("process", {})
    gantry_data = data.get("gantry", {})
    safety_data = data.get("safety", {})

    all_data = {}

    # ---- Styling ----
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ---- Helpers ----
    def init_date_entry(date):
        if date not in all_data:
            all_data[date] = {
                "Equipment": {"details": [], "total": 0}
            }

    # def merge_section(section_data):
    #     for date, content in section_data.items():
    #         init_date_entry(date)
    #         if "Equipment" in content:
    #             all_data[date]["Equipment"]["details"].extend(content["Equipment"]["details"])
    #             # all_data[date]["Equipment"]["total"] += content["Equipment"]["total"]
    #             all_data[date]["Equipment"]["open_alerts_current_carry_count"] += \
    #                 content["Equipment"].get("open_alerts_current_carry_count", 0)
    
    def merge_section(section_data):
        for date, content in section_data.items():

            if date not in all_data:
                all_data[date] = {
                    "Equipment": {
                        "open_alerts_current_carry_count": 0,
                        "open_alerts_current_day": 0,
                        "close_alerts_current_day": 0,
                        "details": []
                    }
                }

            # Ensure keys exist (extra safety)
            equipment_data = all_data[date]["Equipment"]

            equipment_data.setdefault("open_alerts_current_carry_count", 0)
            equipment_data.setdefault("open_alerts_current_day", 0)
            equipment_data.setdefault("close_alerts_current_day", 0)
            equipment_data.setdefault("details", [])

            # Now safely add
            equipment_data["open_alerts_current_carry_count"] += \
                content["Equipment"].get("open_alerts_current_carry_count", 0)

            equipment_data["open_alerts_current_day"] += \
                content["Equipment"].get("open_alerts_current_day", 0)

            equipment_data["close_alerts_current_day"] += \
                content["Equipment"].get("close_alerts_current_day", 0)

            equipment_data["details"].extend(
                content["Equipment"].get("details", [])
            )



    # ---- Merge all sections ----
    merge_section(process_data)
    merge_section(gantry_data)
    merge_section(safety_data)

    # ---- Guard clause ----
    if not all_data:
        print("No data found.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment Alerts"

    def get_period_label(date_str, report_type):
        if report_type == "daily":
            return date_str
        date_formats = ["%Y-%m-%d", "%b-%Y", "%B-%Y", "%Y-%m", "%m-%Y"]
        for fmt in date_formats:
            try:
                parsed_date = datetime.datetime.strptime(date_str, fmt)
                return parsed_date.strftime("%B %Y")
            except ValueError:
                continue
        return date_str

    filters = filters or {}

    def matches_filters(detail):
        for key in ["equipment_name", "sensor_id"]:
            if key in filters and filters[key] is not None:
                if str(detail.get(key)) != str(filters[key]):
                    return False
        return True

    # ---- Sample record for metadata ----
    sample = None
    for entry in all_data.values():
        if entry["Equipment"]["details"]:
            sample = entry["Equipment"]["details"][0]
            break

    sample = sample or {}

    # ---- Report Title ----
    report_title = "WEEKLY / DATE RANGE REPORT" if report_type == "daily" else "MONTHLY REPORT"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = report_title
    title_cell.font = Font(bold=True, size=14, color="000000")
    title_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = thin_border

    row = 3

    # ---- Metadata ----
    total_records = sum(
        len([d for d in entry["Equipment"]["details"] if matches_filters(d)])
        for entry in all_data.values()
    )
    total_alert_count = sum(
        sum(d["count"] for d in entry["Equipment"]["details"] if d.get("type") != "carry_forward")
        for entry in all_data.values()
    )

    selected_name = filters.get("equipment_name", "All")
    selected_id = filters.get("sensor_id", "All")

    metadata = {
        "Export Date and Time": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Total Records": total_records,
        "Total Alert Count": total_alert_count,
        "Data Type": "Equipment Alert",
        "Equipment Name Selected": selected_name,
        "Equipment ID Selected": selected_id,
        "Zone": sample.get("zone", ""),
        "Plant Name": sample.get("location_name", ""),
        "SAP ID": sample.get("sap_id", ""),
        "Date Range": f"{min(all_data)} to {max(all_data)}" if all_data else "N/A"
    }

    for key, val in metadata.items():
        cell_key = ws[f"A{row}"]
        cell_val = ws[f"B{row}"]
        cell_key.value = key
        cell_key.font = Font(bold=True, color="000000")
        cell_key.fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
        cell_key.alignment = Alignment(horizontal="center")
        cell_key.border = thin_border
        cell_val.value = val
        cell_val.alignment = Alignment(horizontal="center")
        cell_val.border = thin_border
        row += 1

    row += 1

    # ---- Table Headers ----
    headers = [
        "Date" if report_type == "daily" else "Month",
        "Equipment Name",
        "Equipment ID",
        "Alert Count"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="000000")
        cell.fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    row += 1

    # ---- Sort Dates (actual date sort) ----
    def sort_key(date_str):
        for fmt in ("%Y-%m-%d", "%Y-%m", "%b-%Y", "%B-%Y", "%m-%Y"):
            try:
                return datetime.datetime.strptime(date_str, fmt)
            except ValueError:
                continue
        return date_str  # fallback as string

    sorted_dates = sorted(all_data.keys(), key=sort_key)

    summary = defaultdict(int)

    for date_str in sorted_dates:
        period_label = get_period_label(date_str, report_type)
        equipment_details = all_data[date_str]["Equipment"]["details"]
        total_cf_count = all_data[date_str]["Equipment"]["total"] if not equipment_details else 0

        if equipment_details:
            for detail in equipment_details:
                if not matches_filters(detail):
                    continue

                alert_type = detail.get("type", "")
                count = detail.get("count", 0)

                ws.cell(row=row, column=1).value = period_label
                ws.cell(row=row, column=2).value = detail.get("equipment_name", "")
                ws.cell(row=row, column=3).value = detail.get("sensor_id", "")
                ws.cell(row=row, column=4).value = count

                if alert_type != "carry_forward":
                    summary[(detail.get("equipment_name", ""), detail.get("sensor_id", ""))] += count

                row += 1
        else:
            # Carry forward placeholder (even in monthly)
            ws.cell(row=row, column=1).value = period_label
            ws.cell(row=row, column=2).value = "Carry Forward"
            ws.cell(row=row, column=3).value = ""
            ws.cell(row=row, column=4).value = total_cf_count
            row += 1

    row += 2

    # ---- Summary Section ----
    ws.cell(row=row, column=1).value = "Equipment Summary"
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1

    ws.cell(row=row, column=1).value = "Equipment Name"
    ws.cell(row=row, column=2).value = "Equipment ID"
    ws.cell(row=row, column=3).value = "Total Alert Count"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=3).font = Font(bold=True)
    row += 1

    for (equip_name, equip_id), count in sorted(summary.items(), key=lambda x: x[1], reverse=True):
        ws.cell(row=row, column=1).value = equip_name
        ws.cell(row=row, column=2).value = equip_id
        ws.cell(row=row, column=3).value = count
        row += 1

    wb.save(output_file)
    print(f"Saved Equipment report to {output_file}")


async def write_interlock_excel(resp, output_file, report_type="daily", filters=None):
    """Write interlock_name_count response to Excel in daily or monthly format."""
    data = resp.get(f"{report_type}_data", {})
    process_data = data.get("process", {})
    gantry_data = data.get("gantry", {})
    safety_data = data.get("safety", {})

    all_data = {}

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def init_date_entry(date):
        if date not in all_data:
            all_data[date] = {
                "Normal": {"details": [], "total": 0},
                "Fault": {"details": [], "total": 0},
                "Maintenance": {"details": [], "total": 0},
            }

    def merge_section(section_data):
        for date, content in section_data.items():
            init_date_entry(date)
            for status in ["Normal", "Fault", "Maintenance"]:
                if status in content:
                    all_data[date][status]["details"].extend(content[status]["details"])
                    all_data[date][status]["total"] += content[status]["total"]

    merge_section(process_data)
    merge_section(gantry_data)
    merge_section(safety_data)

    if not all_data:
        print("No data found in any section.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Interlock Alerts"

    # Determine a sample record
    sample = None
    for entry in all_data.values():
        for status in ["Normal", "Maintenance", "Fault"]:
            if entry[status]["details"]:
                sample = entry[status]["details"][0]
                break
        if sample:
            break

    if not sample:
        raise ValueError("No valid details found")

    filters = filters or {}

    def matches_filters(detail):
        for key in ["bcu_number", "equipment_name", "equipment_id", "assigned_bay"]:
            if key in filters and filters[key] is not None:
                if str(detail.get(key)) != str(filters[key]):
                    return False
        return True

    # Dynamically detect primary and secondary keys
    primary_key = None
    secondary_key = None

    for key in ["bcu_number", "equipment_name", "equipment_id", "assigned_bay"]:
        found = any(
            detail.get(key)
            for entry in all_data.values()
            for status in ["Normal", "Fault", "Maintenance"]
            for detail in entry[status]["details"]
        )
        if found:
            if not primary_key:
                primary_key = key
            elif not secondary_key:
                secondary_key = key
            if primary_key and secondary_key:
                break

    primary_key = primary_key or "bcu_number"
    secondary_key = secondary_key or "interlock_name"

    # ---- Report Title ----
    report_title = "WEEKLY / DATE RANGE REPORT" if report_type == "daily" else "MONTHLY REPORT"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = report_title
    title_cell.font = Font(bold=True, size=14, color="000000")
    title_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = thin_border

    row = 3

    # ---- Metadata Construction ----
    total_records = sum(
        len([d for d in entry[status]["details"] if matches_filters(d)])
        for entry in all_data.values() for status in ["Normal", "Maintenance", "Fault"]
    )
    total_alert_count = sum(
        entry[status]["total"]
        for entry in all_data.values() for status in ["Normal", "Maintenance", "Fault"]
    )

    # Display selected key in metadata
    selected_key = next((k for k in ["bcu_number", "equipment_name", "equipment_id", "assigned_bay"] if filters.get(k)), None)
    selected_label = selected_key.replace("_", " ").title() if selected_key else primary_key.replace("_", " ").title()
    selected_value = filters.get(selected_key, "All") if selected_key else "All"

    metadata = {
        "Export Date and Time": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Total Records": total_records,
        "Total Alert Count": total_alert_count,
        "Data Type": "Interlock Alert",
        f"{selected_label} Selected": selected_value,
        "Zone": sample.get("zone", ""),
        "Plant Name": sample.get("location_name", ""),
        "SAP ID": sample.get("sap_id", ""),
        "Date Range": f"{min(all_data)} to {max(all_data)}"
    }


    for key, val in metadata.items():
        cell_key = ws[f"A{row}"]
        cell_val = ws[f"B{row}"]
        cell_key.value = key
        cell_key.font = Font(bold=True, color="000000")
        cell_key.fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
        cell_key.alignment = Alignment(horizontal="center")
        cell_key.border = thin_border
        cell_val.value = val
        cell_val.alignment = Alignment(horizontal="center")
        cell_val.border = thin_border
        row += 1

    row += 1

    # ---- Table Headers ----
    headers = [
        "Date" if report_type == "daily" else "Month",
        primary_key.replace("_", " ").title(),
        secondary_key.replace("_", " ").title(),
        "Alert count"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="000000")
        cell.fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    row += 1

    # ---- Main Table ----
    summary = defaultdict(int)
    interlock_summary = defaultdict(int)
    sorted_dates = sorted(all_data.keys())

    def get_period_label(date_str, report_type):
        if report_type == "daily":
            return date_str
        
        # Try different date formats for monthly conversion
        date_formats = [
            "%Y-%m-%d",    # 2025-06-10
            "%b-%Y",       # Apr-2025
            "%B-%Y",       # April-2025
            "%Y-%m",       # 2025-04
            "%m-%Y"        # 04-2025
        ]
        
        for fmt in date_formats:
            try:
                parsed_date = datetime.datetime.strptime(date_str, fmt)
                return parsed_date.strftime("%B %Y")  # Return "April 2025" format
            except ValueError:
                continue
        
        # If no format matches, return original string
        return date_str
    
    for date_str in sorted_dates:
        period_label = get_period_label(date_str, report_type)
        
        for detail in all_data[date_str].get("Normal", {}).get("details", []):
            ws.cell(row=row, column=1).value = period_label
            ws.cell(row=row, column=2).value = detail["bcu_number"]
            ws.cell(row=row, column=3).value = detail["interlock_name"]
            ws.cell(row=row, column=4).value = detail["count"]
            
            # Update summaries
            summary[detail["bcu_number"]] += detail["count"]
            interlock_summary[detail["interlock_name"]] += detail["count"]
            row += 1

    row += 2

    # ---- BCU Summary ----
    ws.cell(row=row, column=1).value = f"{primary_key.replace('_', ' ').title()} Summary"
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1).value = primary_key.replace("_", " ").title()
    ws.cell(row=row, column=2).value = "Total count of alerts"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).font = Font(bold=True)
    row += 1

    for key_val in sorted(summary.keys()):
        ws.cell(row=row, column=1).value = key_val
        ws.cell(row=row, column=2).value = summary[key_val]
        row += 1

    row += 2

    # ---- Interlock Summary ----
    ws.cell(row=row, column=1).value = f"{secondary_key.replace('_', ' ').title()} Summary"
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1).value = secondary_key.replace("_", " ").title()
    ws.cell(row=row, column=2).value = "Total count of alerts"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).font = Font(bold=True)
    row += 1

    for key_val in sorted(interlock_summary.keys(), key=lambda x: interlock_summary[x], reverse=True):
        ws.cell(row=row, column=1).value = key_val
        ws.cell(row=row, column=2).value = interlock_summary[key_val]
        row += 1

    wb.save(output_file)
    print(f"Saved file to {output_file}")


async def critical_parameters_excel(resp, output_file, report_type="daily", filters=None):
    data = resp.get(f"{report_type}_data", {})
    if not data:
        print("No data found in response.")
        return

    if isinstance(filters, list):
        try:
            filters = {f.key: f.value for f in filters if hasattr(f, 'key') and hasattr(f, 'value')}
        except Exception:
            filters = {}
    filters = filters or {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Interlock Alerts"

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    report_title = "WEEKLY / DATE RANGE REPORT" if report_type == "daily" else "MONTHLY REPORT"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = report_title
    title_cell.font = Font(bold=True, size=14, color="000000")
    title_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = thin_border

    row = 3

    first_record = next((entry[0] for entry in data.values() if entry), None)
    if not first_record:
        raise ValueError("No valid records found in the response data.")
    
    def record_matches_filters(detail):
        for key in ["bcu_number", "assigned_bay"]:
            if key in filters and filters[key] is not None:
                if str(detail.get(key)) != str(filters[key]):
                    return False
        return True

    selected_key = next((k for k in ["bcu_number", "equipment_name", "equipment_id", "assigned_bay"] if filters.get(k)), None)
    selected_label = selected_key.replace("_", " ").title() if selected_key else "BCU Number"
    selected_value = filters.get(selected_key, "All") if selected_key else "All"

    total_records = sum(
        1 for period in data.values() for record in period if record_matches_filters(record)
    )
    total_alerts = sum(
        record["total_alerts"]
        for period in data.values()
        for record in period
        if record_matches_filters(record)
    )
    metadata = {
        "Export Date and Time": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Total Records": total_records,
        "Total Alert Count": total_alerts,
        "Data Type": "Interlock Alert",
        f"{selected_label} Selected": selected_value,
        "Zone": first_record.get("zone", ""),
        "Plant Name": first_record.get("location_name", ""),
        "SAP ID": first_record.get("sap_id", ""),
        "Date Range": f"{min(data)} to {max(data)}"
    }

    for key, val in metadata.items():
        cell_key = ws[f"A{row}"]
        cell_val = ws[f"B{row}"]
        cell_key.value = key
        cell_key.font = Font(bold=True)
        cell_key.fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
        cell_key.alignment = Alignment(horizontal="center")
        cell_key.border = thin_border
        cell_val.value = val
        cell_val.alignment = Alignment(horizontal="center")
        cell_val.border = thin_border
        row += 1

    row += 1

    headers = [
        "Date" if report_type == "daily" else "Month",
        "BCU Number",
        "Alert Count"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    row += 1

    summary = defaultdict(int)
    wrote_data = False

    for period, records in sorted(data.items()):
        for record in records:
            if not record_matches_filters(record):
                continue

            try:
                col1_val = period
                col2_val = record.get("bcu_number", "N/A")
                col3_val = record.get("total_alerts", "N/A")

                ws.cell(row=row, column=1).value = col1_val
                ws.cell(row=row, column=2).value = col2_val
                ws.cell(row=row, column=3).value = col3_val

                summary[col2_val] += col3_val if isinstance(col3_val, int) else 0
                row += 1
                wrote_data = True
            except Exception as e:
                print(f"Error writing row: {e}")

    if not wrote_data:
        print("No data rows written — possibly due to filters or missing fields.")

    row += 2

    ws.cell(row=row, column=1).value = "BCU Number Summary"
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1).value = "BCU Number"
    ws.cell(row=row, column=2).value = "Total Alerts"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).font = Font(bold=True)
    row += 1

    for bcu in sorted(summary):
        ws.cell(row=row, column=1).value = bcu
        ws.cell(row=row, column=2).value = summary[bcu]
        row += 1

    wb.save(output_file)
    print(f"Excel saved to {output_file}")

def get_interlock_name_and_instance_name_vts(interlock_name, instance_count):
    # special cases: 11, 12, 13 always end with "th"
    if 10 <= instance_count % 100 <= 13:
        suffix = "th"
    else:
        last_digit = instance_count % 10
        if last_digit == 1:
            suffix = "st"
        elif last_digit == 2:
            suffix = "nd"
        elif last_digit == 3:
            suffix = "rd"
        else:
            suffix = "th"
    
    # Split and remove last word (ThirdTime, SecondTime, etc.)
    parts = interlock_name.split()
    base_name = " ".join(parts[:-1])
    interlock_name = f"{base_name} {instance_count}{suffix}Time"
    instance_name = f"Instance - {instance_count}"
    return interlock_name, instance_name

async def calculate_productivity(productivity):
    try:
        rows = []
        for carousal, phases in productivity.items():
            row = {"carousal": int(carousal)}
            for phase, metrics in phases.items():
                for key, value in metrics.items():
                    row[f"{phase}_{key}"] = value
            rows.append(row)
        df = pd.DataFrame(rows)

        net_hours_column = ["normal_net_hours", "break_net_hours", "overtime_net_hours"]
        production_columns = ["normal_total_production", "break_total_production", "overtime_total_production"]
        
        for col in net_hours_column + production_columns:
            if col in df.columns:
                df[col] = df[col].fillna(0).astype(np.float64).abs()

        df["total_net_hours"] = df["normal_net_hours"] + df["break_net_hours"] + df["overtime_net_hours"]
        df["total_production"] = df["normal_total_production"] + df["break_total_production"] + df["overtime_total_production"]
        df["total_productivity"] = df["total_production"] / df["total_net_hours"]
        print("*"*20)
        print("--- productivity ---")
        print(df[["carousal", "total_production", "total_net_hours", "total_productivity"]])
        print("*"*20)
        return df
    except Exception as e:
        return pd.DataFrame()
    
