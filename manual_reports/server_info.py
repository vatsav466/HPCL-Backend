import json
import paramiko
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from copy import copy

# ======================================================
# CONFIG
# ======================================================
TEMPLATE_PATH = "ServerDetails.xlsx"
OUTPUT_PATH = "server_inventory_output.xlsx"

ranges = [
    ("10.90.38", 161, 167, "...."),
    ("10.90.38", 171, 177, "...."),
    ("10.90.38", 211, 222, "...."),
]

CREDENTIALS = []
for base, start, end, password in ranges:
    for ip in range(start, end + 1):
        CREDENTIALS.append({
            "IPAddress": f"{base}.{ip}",
            "UserName": "novex",
            "Password": password,
            "port": 22
        })


# ======================================================
# EXCEL HELPERS
# ======================================================
def safe_set_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        cell.value = value
        return
    for merged in ws.merged_cells.ranges:
        if cell.coordinate in merged:
            ws.cell(merged.min_row, merged.min_col).value = value
            return


def copy_row(ws, src_row, tgt_row):
    ws.insert_rows(tgt_row)
    for col in range(1, ws.max_column + 1):
        src = ws.cell(src_row, col)
        tgt = ws.cell(tgt_row, col)
        tgt.value = None
        if src.has_style:
            tgt.font = copy(src.font)
            tgt.border = copy(src.border)
            tgt.fill = copy(src.fill)
            tgt.number_format = copy(src.number_format)
            tgt.protection = copy(src.protection)
            tgt.alignment = copy(src.alignment)


def build_column_map(ws):
    """
    Builds 2-level header mapping:
    Storage.Name, Partitions.MountPoint, etc.
    """
    col_map = {}
    parent = None
    for col in range(1, ws.max_column + 1):
        p = ws.cell(1, col).value
        c = ws.cell(2, col).value
        if p:
            parent = p.strip()
        if c:
            col_map[f"{parent}.{c.strip()}"] = col
        else:
            col_map[parent] = col
    return col_map


# ======================================================
# SSH HELPERS
# ======================================================
def run_cmd(ssh, cmd):
    return ssh.exec_command(cmd)[1].read().decode().strip()


# ======================================================
# DATA COLLECTION (CORRECT MODEL)
# ======================================================
def collect_server_data(cred):
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    try:
        ssh.connect(
            hostname=cred["IPAddress"],
            username=cred["UserName"],
            password=cred["Password"],
            port=cred["port"],
            timeout=8
        )

        server = {
            "IPAddress": cred["IPAddress"],
            "HostName": run_cmd(ssh, "hostname"),
            "OsName": run_cmd(ssh, "grep ^NAME= /etc/os-release | cut -d= -f2 | tr -d '\"'"),
            "OsVersion": run_cmd(ssh, "grep ^VERSION_ID= /etc/os-release | cut -d= -f2 | tr -d '\"'"),
            "CPU Cores": int(run_cmd(ssh, "nproc")),
            "Memory(GB)": int(run_cmd(ssh, "free -g | awk '/Mem:/ {print $2}'"))
        }
        storage = []

        # Disk ↔ Partition structure
        lsblk_cmd = "lsblk -ndo NAME,TYPE,SIZE,PKNAME | awk '$2==\"disk\" || $2==\"part\"'"
        lsblk_raw = run_cmd(ssh, lsblk_cmd)

        for line in lsblk_raw.splitlines():
            name, typ, size, parent = (line.split() + [""])[:4]
            path = f"/dev/{name}"

            if typ == "disk":
                storage.append({
                    "partition": path,
                    "mount": '-',
                    "size": size,
                    "used": '-',
                    "free": '-',
                    "type": "Disk"
                })

        # Partition usage
        df_cmd = "df -h --output=source,target,size,used,avail | awk '$1 ~ \"^/dev/\"'"
        df_raw = run_cmd(ssh, df_cmd)

        for line in df_raw.splitlines():
            dev, mount, size, used, free = line.split()
            storage.append({
                "partition": dev,
                "mount": mount,
                "size": size,
                "used": used,
                "free": free,
                "type": "Partition"
            })

        return server, storage

    finally:
        ssh.close()


# ======================================================
# EXCEL GENERATION
# ======================================================
def generate_excel():
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    col_map = build_column_map(ws)
    row = 4  # data starts after headers + spacer

    for cred in CREDENTIALS:
        try:
            server, storage = collect_server_data(cred)
        except Exception as e:
            print(f"{cred['IPAddress']} failed: {e}")
            continue

        print("*" * 30)
        print(json.dumps(server))
        print(json.dumps(storage))
        print("*" * 30)

        safe_set_cell(ws, row, col_map["IPAddress"], server["IPAddress"])
        safe_set_cell(ws, row, col_map["HostName"], server["HostName"])
        safe_set_cell(ws, row, col_map["OsName"], server["OsName"])
        safe_set_cell(ws, row, col_map["OsVersion"], server["OsVersion"])
        safe_set_cell(ws, row, col_map["CPU Cores"], server["CPU Cores"])
        safe_set_cell(ws, row, col_map["Memory(GB)"], server["Memory(GB)"])

        for storage_ in storage:
            safe_set_cell(ws, row, col_map["Storage.Partition"], storage_["partition"])
            safe_set_cell(ws, row, col_map["Storage.Type"], storage_["type"])
            safe_set_cell(ws, row, col_map["Storage.MountPoint"], storage_["mount"])
            safe_set_cell(ws, row, col_map["Storage.Size"], storage_["size"])
            safe_set_cell(ws, row, col_map["Storage.Used"], storage_["used"])
            safe_set_cell(ws, row, col_map["Storage.Free"], storage_["free"])
            row += 1

        # spacer row
        copy_row(ws, src_row=3, tgt_row=row)
        row += 1

    wb.save(OUTPUT_PATH)
    print(f"✅ Excel generated successfully: {OUTPUT_PATH}")


# ======================================================
# MAIN
# ======================================================
if __name__ == "__main__":
    generate_excel()
